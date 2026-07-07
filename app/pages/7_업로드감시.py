"""업로드감시 (upload-monitor) — 채널 등록 갭/품절 정합.

박스재고가 있는데 각 채널에 아직 등록(업로드) 안 된 관리코드를 탐지
(+ 재고0인데 채널엔 살아있는 = 품절처리 대상). 등록현황 = 채널마진모니터 listing 스냅샷 공유,
재고·매입가 = 상품관리(product_master). 키=상품코드, 우선순위=재고금액(박스재고×박스매입가).
근거 = KB workflows/upload-monitor.md (ADR 0017).
"""
import base64
import json
import sys
import urllib.request
from io import BytesIO
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent.parent))  # repo root

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from core.workflows import upload_monitor as um

_REF = Path(__file__).parent.parent.parent / "reference"
_APP_REPO = "OURPROJECTDAO/work-automation-app"
_SKIP_PATH = "reference/upload_skip.csv"
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_CODE_COLS = {"상품코드", "관리코드"}  # 엑셀 자동변환(앞0 제거·날짜화) 방지 → 텍스트 서식


def _to_xlsx(df, text_cols, sheet_name="업로드감시") -> bytes:
    """DataFrame → xlsx bytes. text_cols 컬럼은 문자열+@ 서식(상품코드 앞0·관리코드 날짜화 방지)."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append([(v.item() if hasattr(v, "item") else v) for v in row.tolist()])
    for ci, col in enumerate(df.columns, start=1):
        if col in text_cols:
            for cell in ws[get_column_letter(ci)]:
                cell.number_format = "@"
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _pat():
    return st.secrets.get("GITHUB_PAT", "")


def _gh(path, method="GET", data=None, raw=False):
    url = f"https://api.github.com/repos/{_APP_REPO}/contents/{path}"
    req = urllib.request.Request(url, method=method)
    if _pat():
        req.add_header("Authorization", f"Bearer {_pat()}")
    req.add_header("Accept", "application/vnd.github.raw" if raw else "application/vnd.github+json")
    if data is not None:
        req.data = json.dumps(data).encode()
        req.add_header("Content-Type", "application/json")
    try:
        with urllib.request.urlopen(req) as r:
            return r.status, (r.read().decode() if raw else json.load(r))
    except urllib.error.HTTPError as e:
        return e.code, (e.read().decode() if raw else json.loads(e.read().decode() or "{}"))


def _commit_skip(new_text, msg):
    if not _pat():
        return False, "저장용 PAT(st.secrets GITHUB_PAT)가 없어 커밋할 수 없습니다."
    code, meta = _gh(_SKIP_PATH)
    sha = meta.get("sha") if code == 200 else None
    body = {"message": msg, "content": base64.b64encode(new_text.encode()).decode("ascii")}
    if sha:
        body["sha"] = sha
    code2, _ = _gh(_SKIP_PATH, "PUT", body)
    return code2 in (200, 201), (None if code2 in (200, 201) else f"커밋 실패: {code2}")

st.title("📦 업로드감시")
st.caption(
    "박스재고가 있는데 채널에 아직 등록 안 된 상품(업로드필요)과, 재고0인데 채널엔 살아있는 "
    "상품(품절처리필요)을 채널별로 점검합니다. 등록현황은 채널마진모니터 상품관리 스냅샷 기준, "
    "재고·매입가는 상품관리(product_master) 기준. 우선순위는 재고금액(박스재고×박스매입가)입니다. "
    "낱개·박스·소분·합포 중 하나라도 올라가 있으면 등록된 것으로 봅니다."
)


@st.cache_data(ttl=60, show_spinner=False)
def _skip_text():
    code, text = _gh(_SKIP_PATH, raw=True)
    return text if code == 200 else ""


skip_set = um.parse_skip_text(_skip_text())


@st.cache_data(ttl=600, show_spinner="재고 · 채널 등록현황 대조 중...")
def _load(skip_key):
    return um.build_gap_table(str(_REF), skip_pairs=set(skip_key))


rows = _load(frozenset(skip_set))
if not rows:
    st.warning("데이터가 없습니다. 상품관리(product_master)와 채널 listing 저장본을 확인하세요.")
    st.stop()

df = pd.DataFrame(rows)
KEYS = um.CHANNEL_KEYS

# ── KPI ──────────────────────────────────────────────────────────────────────
need_any = int((df[KEYS] == um.ST_NEED_UP).any(axis=1).sum())
sold_any = int((df[KEYS] == um.ST_NEED_SOLD).any(axis=1).sum())
k1, k2, k3 = st.columns(3)
k1.metric("감시대상 상품", f"{len(df):,}")
k2.metric("업로드필요", f"{need_any:,}", help="한 채널 이상에서 업로드필요")
k3.metric("품절처리필요", f"{sold_any:,}", help="재고0인데 채널엔 등록·판매 중")

st.markdown("**채널별 미업로드 / 품절처리 / 업로드제외 건수**")
sdf = pd.DataFrame(um.channel_summary(rows))[
    ["label", "업로드필요", "품절처리필요", "업로드제외"]].rename(columns={"label": "채널"})
st.dataframe(sdf, hide_index=True, width="stretch",
             column_config={"업로드필요": st.column_config.NumberColumn(format="localized"),
                            "품절처리필요": st.column_config.NumberColumn(format="localized"),
                            "업로드제외": st.column_config.NumberColumn(format="localized")})

# ── 채널 선택 (체크박스 + 전체 선택/해제) ────────────────────────────────────
ALL_STATUS = ["(전체)", um.ST_NEED_UP, um.ST_OK, um.ST_NEED_SOLD, um.ST_SKIP_CH, um.ST_SKIP]


def _toggle_all_channels():
    all_on = all(st.session_state.get(f"um_ch_{k}", True) for k in KEYS)
    for k in KEYS:
        st.session_state[f"um_ch_{k}"] = not all_on


tcol, _ = st.columns([1, 5])
tcol.button("전체 선택/해제", on_click=_toggle_all_channels, width="stretch")
ch_boxes = st.columns(len(um.CHANNELS))
for (k, label, _au), col in zip(um.CHANNELS, ch_boxes):
    col.checkbox(label, value=True, key=f"um_ch_{k}")
selected = [k for k in KEYS if st.session_state.get(f"um_ch_{k}", True)]

# ── 컬럼별(헤더) 상태 필터 — 선택 채널만 ──────────────────────────────────────
col_status: dict[str, str] = {}
if selected:
    st.markdown("**🔎 채널 컬럼별 상태 필터** — 채널마다 보고 싶은 상태를 고르세요 "
                "(예: 스마트스토어 = `이상없음`). 여러 채널이면 AND로 걸립니다.")
    fcols = st.columns(len(selected))
    for k, col in zip(selected, fcols):
        col_status[k] = col.selectbox(
            f"{um.CHANNEL_LABEL[k]} 상태", ALL_STATUS, index=0, key=f"um_st_{k}",
            help=f"표의 '{um.CHANNEL_LABEL[k]}' 컬럼을 이 상태인 행만 보이게 필터합니다.")
else:
    st.caption("채널을 하나 이상 선택하면 해당 상태 컬럼과 컬럼 필터가 표시됩니다.")

sc1, sc2, sc3, sc4 = st.columns([3, 1.5, 1.6, 1.4])
search = sc1.text_input("🔍 검색", placeholder="관리코드 · 상품코드 · 상품명 (부분일치)",
                        label_visibility="collapsed")
min_stock = sc2.number_input("박스재고 ≥", min_value=0, value=0, step=1,
                             help="박스재고가 이 값 이상인 상품만 (0=미적용)")
min_amount = sc3.number_input("재고금액 ≥ (원)", min_value=0, value=0, step=100000,
                              help="재고금액(박스재고×박스매입가)이 이 값 이상만 (0=미적용)")
show_excluded = sc4.checkbox("전채널 제외 포함", value=False,
                             help="모든 채널이 '업로드제외'인 상품(비대상)은 기본 숨김")

# ── 행 필터 (전채널제외 숨김 + 임계값 + 선택 채널 상태 AND + 검색) ─────────────
view = df.copy()
if not show_excluded:
    view = view[~(view[KEYS] == um.ST_SKIP_CH).all(axis=1)]
if min_stock > 0:
    view = view[view["박스재고"] >= min_stock]
if min_amount > 0:
    view = view[view["재고금액"] >= min_amount]
for k in selected:
    s = col_status.get(k, "(전체)")
    if s != "(전체)":
        view = view[view[k] == s]
if search:
    q = um._nfc(search).lower()
    hay = (view["관리코드"].astype(str) + " ||| " + view["상품코드"].astype(str)
           + " ||| " + view["상품명"].astype(str)).str.lower()
    view = view[hay.str.contains(q, regex=False, na=False)]

# 표시: 기본정보 + 선택 채널 컬럼(재고금액 옆). 채널키→라벨 헤더.
base_cols = ["상품코드", "관리코드", "상품명", "박스재고", "박스매입가", "재고금액"]
view_reset = view.reset_index(drop=True)
disp = view_reset[base_cols + selected].rename(columns=um.CHANNEL_LABEL)

filter_sig = hash((tuple(selected),
                   tuple(col_status.get(k, "(전체)") for k in selected),
                   um._nfc(search) if search else "", show_excluded,
                   min_stock, min_amount, len(view_reset)))
event = st.dataframe(
    disp,
    width="stretch",
    hide_index=True,
    on_select="rerun",
    selection_mode="multi-row",
    key=f"um_df_{filter_sig}",
    column_config={
        "박스재고": st.column_config.NumberColumn(format="localized"),
        "박스매입가": st.column_config.NumberColumn(format="localized"),
        "재고금액": st.column_config.NumberColumn(format="localized"),
    },
)
_sel = event.selection.rows if event and getattr(event, "selection", None) else []
sel_rows = [i for i in _sel if 0 <= i < len(view_reset)]
sel_codes = view_reset.iloc[sel_rows]["상품코드"].tolist() if sel_rows else []

amt = int(view["재고금액"].sum()) if len(view) else 0
st.caption(f"표시 {len(view):,} / 전체 {len(df):,}건 · 재고금액합 {amt:,}원 · "
           f"✅ 선택 **{len(sel_codes):,}건** (표 왼쪽 체크박스로 개별, 헤더로 화면 전체)")

# ── 내보내기 (XLSX — 상품코드/관리코드 텍스트 서식, 엑셀 자동변환 방지) ──────────
csv_src = view_reset.iloc[sel_rows] if sel_rows else view_reset
exp = csv_src[base_cols + selected].rename(columns=um.CHANNEL_LABEL)
if len(selected) == 1:
    tag = um.CHANNEL_LABEL[selected[0]]
elif len(selected) == len(KEYS) or not selected:
    tag = "전체"
else:
    tag = f"{len(selected)}채널"
st.download_button("📥 엑셀(XLSX) 다운로드 (선택 또는 현재 화면)",
                   _to_xlsx(exp, _CODE_COLS),
                   file_name=f"업로드감시_{tag}.xlsx", mime=_XLSX_MIME, key="um_xlsx_dl")

# ── 이미지 포함 CSV (대표 A1 / 상세 B1 실검사) ───────────────────────────────
if "um_img_cache" not in st.session_state:
    st.session_state.um_img_cache = {}


def _probe_view(codes):
    cache = st.session_state.um_img_cache
    todo = sorted({um._nfc(c) for c in codes if um._nfc(c) and um._nfc(c) not in cache})
    if todo:
        with st.spinner(f"{len(todo)}개 관리코드 이미지 실검사 중..."):
            cache.update(um.probe_images(todo))
    return cache


st.caption("🖼 대표(A1)·상세(B1) 이미지 유무·확장자·URL을 gi.esmplus.com에서 실검사해 엑셀에 추가합니다 "
           "(관리코드 기준, 빈 관리코드는 제외). 같은 관리코드는 세션 내 재검사 안 함.")
if st.button("🖼 이미지 확인 후 엑셀 만들기 (선택 또는 현재 화면)"):
    cache = _probe_view(csv_src["관리코드"].tolist())
    enr = csv_src[base_cols].copy()
    for ic in um.IMG_COLS:
        enr[ic] = csv_src["관리코드"].map(lambda m, ic=ic: cache.get(um._nfc(m), {}).get(ic, ""))
    for ch in selected:
        enr[um.CHANNEL_LABEL[ch]] = csv_src[ch].values
    st.session_state["um_img_csv"] = _to_xlsx(enr, _CODE_COLS)
    st.session_state["um_img_tag"] = tag
    a_o = int((enr["대표이미지유무"] == "O").sum())
    b_o = int((enr["상세이미지유무"] == "O").sum())
    st.session_state["um_img_summary"] = (a_o, len(enr) - a_o, b_o, len(enr) - b_o)

if st.session_state.get("um_img_csv"):
    s = st.session_state["um_img_summary"]
    st.caption(f"대표(A1) 있음 **{s[0]}** · 없음 {s[1]}  /  상세(B1) 있음 **{s[2]}** · 없음 {s[3]} "
               "— 전부 '없음'이면 배포 환경에서 gi.esmplus.com 접근이 막힌 것일 수 있습니다.")
    st.download_button("📥 이미지 포함 엑셀(XLSX) 다운로드", st.session_state["um_img_csv"],
                       file_name=f"업로드감시_{st.session_state.get('um_img_tag','전체')}_이미지.xlsx",
                       mime=_XLSX_MIME, key="um_img_dl")

st.divider()
st.subheader("🚫 채널별 업로드제외 (등록 / 해제)")
st.caption("표에서 상품을 선택하고 위 채널 체크박스로 대상 채널을 고른 뒤 등록/해제합니다. "
           "업로드제외는 '업로드필요'보다 우선 표시되며, 모든 채널이 제외된 상품은 기본 숨김됩니다. "
           "해제하면 다시 '업로드필요'로 돌아옵니다.")

if not _pat():
    st.warning("저장용 PAT(st.secrets GITHUB_PAT)가 없어 등록/해제를 커밋할 수 없습니다.")
elif not sel_codes:
    st.info("표에서 상품을 1개 이상 선택하세요.")
elif not selected:
    st.info("위 채널 체크박스에서 대상 채널을 1개 이상 선택하세요.")
else:
    pairs = [(sc, ch) for sc in sel_codes for ch in selected]
    chips = " · ".join(um.CHANNEL_LABEL[c] for c in selected)
    st.write(f"대상: **{len(sel_codes)}개 상품 × {len(selected)}채널 = {len(pairs)}쌍**  ({chips})")
    with st.expander(f"대상 (상품코드 × 채널) 미리보기 — {len(pairs)}쌍"):
        st.dataframe(pd.DataFrame([{"상품코드": sc, "채널": um.CHANNEL_LABEL[ch]} for sc, ch in pairs]),
                     hide_index=True, width="stretch")
    bcol1, bcol2 = st.columns(2)
    if bcol1.button("🚫 업로드제외 등록", width="stretch", type="primary"):
        ok, err = _commit_skip(um.build_skip_text(skip_set | set(pairs)),
                               f"upload-monitor: 업로드제외 등록 {len(pairs)}쌍")
        if ok:
            _skip_text.clear(); _load.clear()
            st.success(f"{len(pairs)}쌍 업로드제외 등록 완료.")
            st.rerun()
        else:
            st.error(err)
    if bcol2.button("↩️ 업로드제외 해제 (다시 업로드모드)", width="stretch"):
        ok, err = _commit_skip(um.build_skip_text(skip_set - set(pairs)),
                               f"upload-monitor: 업로드제외 해제 {len(pairs)}쌍")
        if ok:
            _skip_text.clear(); _load.clear()
            st.success(f"{len(pairs)}쌍 해제 완료 — 다시 업로드모드.")
            st.rerun()
        else:
            st.error(err)

st.info("스마트스토어·ESM 등록폼 자동생성(L4)은 다음 단계입니다. 지금은 채널별 갭을 CSV로 활용하세요.")
