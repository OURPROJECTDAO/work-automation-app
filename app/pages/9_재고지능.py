"""🔮 재고 지능 (두뇌② 입고·품절 예측) — 현재고 ÷ 소진율 → 재발주 시점.

intelligence-layer.md §6 ②. 현재고(product_master 최종재고·낱개) ÷ 일소진(매출자료 전채널 낱개)
= 소진예측일. 입고주기(매입현황 중앙 입고간격)를 리드타임 proxy로 재발주 시점 판정.
4구간: 🔴 품절임박(소진예측일≤리드) · 🟡 곧재발주 · 🟢 충분 · ⚪ 사장재고(매출0·재고묶임).
★소진은 base 관리코드 직조인(매출자료가 이미 낱개 분해라 합성코드 문제 없음). 음수재고=0 클램프.
"""
import csv
import io
import json
import sys
import urllib.error
import urllib.request
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from core.intelligence import stockout as so
from core.intelligence import purchases as buy_mod
from core.workflows import upload_monitor as um

_APP_API = "https://api.github.com/repos/OURPROJECTDAO/work-automation-app/contents"
_CODE_COLS = {"관리코드"}


def _pat() -> str:
    return st.secrets.get("GITHUB_PAT", "")


def _data_secret():
    repo = "OURPROJECTDAO/work-automation-data"
    try:
        d = st.secrets["data"]
        return d["pat"], d.get("repo", repo)
    except Exception:
        return st.secrets.get("GITHUB_PAT", ""), repo


def _gh_app_raw(path: str):
    req = urllib.request.Request(
        f"{_APP_API}/{path}",
        headers={"Authorization": f"Bearer {_pat()}", "Accept": "application/vnd.github.raw"},
    )
    try:
        with urllib.request.urlopen(req) as r:
            return r.status, r.read()
    except urllib.error.HTTPError as e:
        return e.code, b""


def _data_raw(path: str):
    pat, repo = _data_secret()
    req = urllib.request.Request(
        f"https://api.github.com/repos/{repo}/contents/{path}?ref=main",
        headers={"Authorization": f"Bearer {pat}", "Accept": "application/vnd.github.raw"},
    )
    try:
        with urllib.request.urlopen(req) as r:
            return r.status, r.read()
    except urllib.error.HTTPError as e:
        return e.code, b""


@st.cache_data(ttl=600, show_spinner=False)
def _product_master() -> pd.DataFrame:
    code, text = _gh_app_raw("reference/product_master.csv")
    if code != 200:
        return pd.DataFrame()
    return pd.read_csv(io.BytesIO(text), dtype=str)


@st.cache_data(ttl=600, show_spinner=False)
def _exclude_codes() -> set:
    """업로드감시 비판매 제외목록(상품코드). 없으면 빈 set."""
    code, text = _gh_app_raw("reference/upload_monitor_exclude.csv")
    if code != 200:
        return set()
    out = set()
    for row in csv.DictReader(io.StringIO(text.decode("utf-8-sig"))):
        sc = (row.get("상품코드") or "").strip()
        if sc:
            out.add(sc)
    return out


@st.cache_data(ttl=600, show_spinner="매출자료(소진율) 불러오는 중...")
def _sales_recent(months: int) -> pd.DataFrame:
    """data repo master/sales_YYYY-MM.parquet 최근 months개."""
    pat, repo = _data_secret()
    if not pat:
        return pd.DataFrame()
    code, body = _data_raw("master")
    if code != 200:
        # 디렉토리 목록은 JSON
        req = urllib.request.Request(
            f"https://api.github.com/repos/{repo}/contents/master?ref=main",
            headers={"Authorization": f"Bearer {pat}", "Accept": "application/vnd.github+json"},
        )
        try:
            items = json.load(urllib.request.urlopen(req))
        except urllib.error.HTTPError:
            return pd.DataFrame()
    else:
        req = urllib.request.Request(
            f"https://api.github.com/repos/{repo}/contents/master?ref=main",
            headers={"Authorization": f"Bearer {pat}", "Accept": "application/vnd.github+json"},
        )
        items = json.load(urllib.request.urlopen(req))
    import re
    mons = sorted(m.group(1) for it in items
                  for m in [re.match(r"sales_(\d{4}-\d{2})\.parquet$", it.get("name", ""))] if m)[-months:]
    parts = []
    for ym in mons:
        c, t = _data_raw(f"master/sales_{ym}.parquet")
        if c == 200:
            parts.append(pd.read_parquet(io.BytesIO(t)))
    return pd.concat(parts, ignore_index=True) if parts else pd.DataFrame()


@st.cache_data(ttl=600, show_spinner="매입현황(입고주기) 불러오는 중...")
def _buyin_recent(months: int) -> pd.DataFrame:
    pat, repo = _data_secret()
    if not pat:
        return pd.DataFrame()
    mons = buy_mod.list_months(pat, repo)[-months:]
    parts = [buy_mod.read_partition(pat, repo, m) for m in mons]
    parts = [p for p in parts if p is not None and len(p)]
    return pd.concat(parts, ignore_index=True) if parts else pd.DataFrame()


def _to_xlsx(df: pd.DataFrame, sheet="재고지능") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append([(v.item() if hasattr(v, "item") else v) for v in row.tolist()])
    for ci, col in enumerate(df.columns, start=1):
        if col in _CODE_COLS:
            for cell in ws[get_column_letter(ci)]:
                cell.number_format = "@"
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
st.title("🔮 재고 지능")
st.caption(
    "현재고(상품관리 최종재고·낱개)를 매출자료(전채널) 소진율로 나눠 **소진 예측일**을 구하고, "
    "매입현황 입고주기를 리드타임으로 보아 **재발주 시점**을 판정합니다. "
    "소진예측일 ≤ 리드타임 = 다음 입고 전 품절 위험(재발주 필요). "
    "매출 없이 재고만 묶인 상품은 '사장재고'로 분리합니다."
)

if not _data_secret()[0]:
    st.warning("매출·매입 데이터 접근용 PAT(st.secrets [data] 또는 GITHUB_PAT)가 없습니다.")
    st.stop()

with st.sidebar:
    st.subheader("⚙️ 설정")
    win = st.slider("소진율 산정 기간(개월)", 1, 12, 3, help="최근 N개월 매출 평균으로 일소진 산정")
    buy_win = st.slider("입고주기 산정 기간(개월)", 3, 24, 12)
    default_lead = st.number_input("기본 리드타임(일)", 1, 90, 14,
                                   help="매입 입고이력이 1회 이하라 입고주기를 못 구한 상품에 적용")
    today = st.date_input("기준일(NOW)", value=pd.Timestamp.now().date())

pm = _product_master()
if pm.empty:
    st.error("product_master를 불러오지 못했습니다.")
    st.stop()
sales = _sales_recent(win)
buyin = _buyin_recent(buy_win)
now = pd.Timestamp(today)

dep = so.depletion_rate(sales, months=win, now=now)
cad = so.restock_cadence(buyin, months=buy_win, now=now)
df = so.forecast(pm, dep, cad, now=now, default_lead_days=float(default_lead),
                 exclude_codes=_exclude_codes(), exclude_midcat=um.EXCLUDE_MIDCAT)

if df.empty:
    st.info("예측할 재고/판매 데이터가 없습니다.")
    st.stop()

summ = so.summarize(df)
c = summ["건수"]
k1, k2, k3, k4 = st.columns(4)
k1.metric("🔴 재발주 필요", f"{summ['재발주필요']}건")
k2.metric("🟡 곧 재발주", f"{c.get(so.B_SOON, 0)}건")
k3.metric("🟢 충분", f"{c.get(so.B_OK, 0)}건")
k4.metric("⚪ 사장재고", f"{c.get(so.B_DEAD, 0)}건",
          f"{summ['재고금액'].get(so.B_DEAD, 0)/1e8:.2f}억 묶임")

tab1, tab2, tab3 = st.tabs(["🔴🟡 재발주 대상", "⚪ 사장재고(과잉)", "전체"])


def _show(table: pd.DataFrame, key: str, show_dead_cols=False):
    if table.empty:
        st.info("해당 조건 상품이 없습니다.")
        return
    cols = ["구간", "관리코드", "상품명", "규격", "현재고(낱개)", "박스재고", "일소진(낱개)",
            "일소진(박스)", "일소진액(매입)", "월낱개", "월소진액(매입)", "소진예측일", "예상소진일자",
            "리드타임", "리드출처", "최근입고일", "입고횟수", "재고금액", "재발주필요"]
    view = table[cols].copy()
    st.dataframe(view, use_container_width=True, hide_index=True,
                 column_config={
                     "재고금액": st.column_config.NumberColumn(format="%d"),
                     "일소진액(매입)": st.column_config.NumberColumn(format="%d"),
                     "월소진액(매입)": st.column_config.NumberColumn(format="%d"),
                     "소진예측일": st.column_config.NumberColumn(format="%.1f"),
                     "일소진(박스)": st.column_config.NumberColumn(format="%.2f"),
                     "재발주필요": st.column_config.CheckboxColumn(),
                 })
    st.download_button("📥 XLSX 다운로드", _to_xlsx(view),
                       file_name=f"재고지능_{key}_{now.date()}.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       key=f"dl_{key}")


with tab1:
    reorder = df[df["구간"].isin([so.B_IMMINENT, so.B_SOON])].copy()
    cc1, cc2, cc3, cc4 = st.columns(4)
    fwd_only = cc1.toggle("현재고>0만", value=False,
                          help="끄면 현재고 0/마이너스(이미 소진·오버셀)도 포함")
    min_flow = cc2.number_input("월소진액(매입) ≥", 0, 100000000, 0, step=100000,
                                help="매입가 기준 월 빠지는 금액 — 상품 간 볼륨 비교")
    min_month = cc3.number_input("월 판매(낱개) ≥", 0, 100000, 0, step=50)
    min_value = cc4.number_input("재고금액 ≥", 0, 100000000, 0, step=100000)
    q = st.text_input("🔎 상품명·관리코드 검색", "")
    f = reorder
    if fwd_only:
        f = f[f["현재고(낱개)"] > 0]
    if min_flow:
        f = f[f["월소진액(매입)"] >= min_flow]
    if min_month:
        f = f[f["월낱개"] >= min_month]
    if min_value:
        f = f[f["재고금액"] >= min_value]
    if q.strip():
        s = q.strip()
        f = f[f["상품명"].str.contains(s, case=False, na=False)
              | f["관리코드"].astype(str).str.contains(s, na=False)]
    st.caption(f"{len(f)}건 — 재발주 필요(🔴) 먼저, 소진예측일 빠른 순. 현재고≤0 = 이미 소진/오버셀.")
    _show(f, "재발주")

with tab2:
    dead = df[df["구간"] == so.B_DEAD].sort_values("재고금액", ascending=False).copy()
    st.caption(f"{len(dead)}건 — 최근 {win}개월 매출 0인데 박스재고 보유(자본 묶임). 재고금액 큰 순.")
    _show(dead, "사장재고")

with tab3:
    bands = st.multiselect("구간", [so.B_IMMINENT, so.B_SOON, so.B_OK, so.B_DEAD],
                           default=[so.B_IMMINENT, so.B_SOON, so.B_OK, so.B_DEAD])
    g = df[df["구간"].isin(bands)]
    st.caption(f"{len(g)}건")
    _show(g, "전체")

st.caption(
    "ⓘ 현재고=최종재고(낱개, =박스×박스내품+낱개). 소진=매출자료 전채널 낱개(정산 진실·이미 낱개 분해). "
    "**소진액(매입)=소진 낱개×낱개 매입단가** — 박스 수량은 상품마다 가치가 달라, 매입원가로 환산해 볼륨을 비교(매입가 미기재 상품은 0). "
    "리드타임=매입현황 중앙 입고간격(발주→입고 실리드타임은 발주자료 적재 후 정밀화). "
    "박스재고는 흐름누적 오차가 있어 상품관리 앵커값 사용·음수는 0 처리(즉시 품절)."
)
