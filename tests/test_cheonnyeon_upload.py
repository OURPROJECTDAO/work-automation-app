"""천년경영업로드자동화 골든 테스트.

fixtures (PII 없음 — 파이프라인 사용 컬럼만 추출):
  baeju_20260604.xlsx        발주자료 7열 (상품수준)
  baemin_ship_20260604.xlsx  배민 배송비 (Z=관리코드, AL=배송비)
  sss_ship_20260604.xlsx     스스 배송비 슬림·평문 (AJ/AL/AO, r2헤더)
  classification_snapshot.csv / bm_commission.csv / sub_list.csv  기준데이터 스냅샷
golden: golden_260604.xlsx (마켓 전체/낱개 시트만)

기준데이터는 fixture로 명시 주입 → 라이브 reference 편집과 무관하게 결정적.
"""
import datetime
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from core.workflows import cheonnyeon_upload as cy

FX = Path(__file__).parent / "fixtures" / "cheonnyeon"
GOLD = Path(__file__).parent / "golden" / "cheonnyeon" / "golden_260604.xlsx"
RUN_DATE = datetime.date(2026, 6, 4)

FULL_COLS = [(1, "B"), (2, "C"), (3, "D"), (4, "E"), (5, "F"), (6, "G"), (7, "H")]
UNIT_COLS = FULL_COLS + [(8, "I"), (9, "J"), (10, "K")]


def _norm(v):
    if v in (None, ""):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return str(v).strip()


def _eq(a, b, tol=0.5):
    a, b = _norm(a), _norm(b)
    if a is None and b is None:
        return True
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return abs(a - b) <= tol
    return a == b


@pytest.fixture(scope="module")
def result():
    cls_map = cy.load_classification(
        pd.read_csv(FX / "classification_snapshot.csv", encoding="utf-8-sig", dtype=str))
    comm_map = cy.load_commission(
        pd.read_csv(FX / "bm_commission.csv", encoding="utf-8-sig", dtype=str))
    sub_map = cy.load_sub_list(
        pd.read_csv(FX / "sub_list.csv", encoding="utf-8-sig", dtype=str))
    baeju = cy.parse_baeju((FX / "baeju_20260604.xlsx").read_bytes())
    baemin_ws = openpyxl.load_workbook(FX / "baemin_ship_20260604.xlsx", data_only=True).active
    sss_ws = openpyxl.load_workbook(FX / "sss_ship_20260604.xlsx", data_only=True).active
    return cy.process(baeju, baemin_ws, sss_ws, cls_map, comm_map, sub_map, RUN_DATE)


@pytest.fixture(scope="module")
def golden():
    wb = openpyxl.load_workbook(GOLD, data_only=True)
    out = {}
    for name in cy.ALL_SHEETS + cy.UNIT_SHEETS:
        nc = 11 if name in cy.UNIT_SHEETS else 8
        if name not in wb.sheetnames:          # 골든에 없는 신규 시트(리테일 등) → 빈 목록
            out[name] = []
            continue
        out[name] = [r for r in wb[name].iter_rows(min_row=2, max_col=nc, values_only=True)
                     if r[1] not in (None, "")]
    wb.close()
    return out


@pytest.mark.parametrize("name", cy.ALL_SHEETS)
def test_full_sheets(result, golden, name):
    sheets, _ = result
    mine, gold = sheets[name], golden[name]
    assert len(mine) == len(gold), f"{name} 행수 {len(mine)} != {len(gold)}"
    for i, (m, g) in enumerate(zip(mine, gold), 2):
        for ci, key in FULL_COLS:
            assert _eq(m.get(key), g[ci]), f"{name} r{i} {key}: {m.get(key)!r} != {g[ci]!r}"


@pytest.mark.parametrize("name", cy.UNIT_SHEETS)
def test_unit_sheets(result, golden, name):
    _, units = result
    mine, gold = units[name], golden[name]
    assert len(mine) == len(gold), f"{name} 행수 {len(mine)} != {len(gold)}"
    for i, (m, g) in enumerate(zip(mine, gold), 2):
        for ci, key in UNIT_COLS:
            assert _eq(m.get(key), g[ci]), f"{name} r{i} {key}: {m.get(key)!r} != {g[ci]!r}"


def test_smartstore_shipping_join(result):
    """배송비 조인 검증: 스마트스토어 G = 배송비 AL*0.964 (대표 1건)."""
    sheets, _ = result
    row = next(r for r in sheets["스마트스토어전체"] if r["B"] == "20-90-11")
    assert abs(row["G"] - 2892) <= 1          # 3000 * 0.964
    assert abs(row["H"] - (row["F"] + row["G"]) / row["D"]) <= 0.5


def test_unit_split_formula(result):
    """낱개 J=D*개당수량, K=H*D/J 검증."""
    _, units = result
    for r in units["스마트스토어낱개"]:
        assert abs(r["J"] - r["D"] * r["I"]) <= 1e-6
        assert abs(r["K"] - r["H"] * r["D"] / r["J"]) <= 0.5


def test_detect_box_anomalies_synthetic():
    """박스코드 이상 탐지: 영문코드/낱개태그만 잡고 정상 박스(트레일링점 포함)는 무탐지."""
    sheets = {n: [] for n in cy.ALL_SHEETS}
    sheets["쿠팡전체"] = [
        {"B": "21-103", "C": "정상 박스상품"},                # 무탐지
        {"B": "29-30.", "C": "트레일링점 박스 24개입"},        # 무탐지(잡티, 영문 아님)
        {"B": "PC005875", "C": "[낱개1개][명가 메밀칩]"},      # 코드영문+상품명 → 높음
        {"B": "TU85G12EA-15-06-04", "C": "소분 참치"},        # 코드영문만 → 검토
        {"B": "99-99", "C": "[낱개]가짜 낱개태그"},           # 상품명만 → 검토
    ]
    anom = cy.detect_box_anomalies(sheets)
    by = {a["관리코드"]: a for a in anom}
    assert set(by) == {"PC005875", "TU85G12EA-15-06-04", "99-99"}
    assert by["PC005875"]["확신"] == "높음"
    assert by["TU85G12EA-15-06-04"]["확신"] == "검토"
    assert by["TU85G12EA-15-06-04"]["신호"] == "코드영문"
    assert by["99-99"]["신호"] == "상품명낱개태그"


def test_detect_box_anomalies_structure(result):
    """실 파이프라인 결과에 대해 list[dict] 스키마 보장(스모크)."""
    sheets, _ = result
    anom = cy.detect_box_anomalies(sheets)
    assert isinstance(anom, list)
    for a in anom:
        assert set(a) == {"시트", "관리코드", "상품명", "신호", "확신"}
        assert a["시트"] in cy.ALL_SHEETS


# ── 추가 판매처 매출통계(제이티·리테일) 병합 ──────────────────
_STATS_HTML = """<html><body><table border=1>
<tr><td>erp관리코드</td><td>어드민 옵션</td><td>총수량</td><td>평균단가</td>
<td>정산금액</td><td>판매처그룹</td><td>선결제택배비</td><td>옵션추가항목1</td></tr>
<tr><td><table><tr><td></td></tr></table></td>
<td><table><tr><td>스위트A 275ML*24 [24-174]</td></tr><tr><td>더블랙 275ML*24 [24-175]</td></tr></table></td>
<td>2</td><td>23,275</td><td>46,550</td><td>리테일앤인사이트</td><td>0</td>
<td><table><tr><td>24-174</td></tr><tr><td>24-175</td></tr></table></td></tr>
<tr><td>16-05</td><td>[롯데]런천미트(大)</td><td>5</td><td>9,936</td>
<td>248,400</td><td>리테일앤인사이트</td><td>0</td><td></td></tr>
<tr><td></td><td>코카콜라제로 355 [31-22-02]</td><td>1</td><td>20,200</td>
<td>20,200</td><td>제이티유통</td><td>0</td><td><table><tr><td>31-22-02</td></tr></table></td></tr>
</table></body></html>""".encode("utf-8")


def test_parse_sales_stats_basic():
    rows, skipped = cy.parse_sales_stats(_STATS_HTML)
    by = {r[0]: r for r in rows}
    # 정상 단일행 + 빈코드 단일코드 보정(코카콜라제로→31-22-02)
    assert set(by) == {"16-05", "31-22-02"}
    assert by["16-05"][2] == 5.0 and by["16-05"][4] == 248400.0      # 콤마 제거
    assert by["16-05"][5] == "리테일앤인사이트"
    assert by["31-22-02"][5] == "제이티유통"
    # 묶음(코드 2개) → 스킵 경고
    assert len(skipped) == 1 and skipped[0]["사유"].startswith("묶음")
    assert skipped[0]["후보코드"] == "24-174 | 24-175"


def test_stats_routing_retail_and_jt():
    """리테일앤인사이트→리테일전체, 제이티유통→제이티전체, H=F/D."""
    rows, _ = cy.parse_sales_stats(_STATS_HTML)
    empty = openpyxl.Workbook(); ews = empty.active; ews.append(["h"] * 45)
    empty2 = openpyxl.Workbook(); ews2 = empty2.active; ews2.append(["h"] * 45)
    sheets, units = cy.process(rows, ews, ews2, {}, {}, {}, RUN_DATE)
    assert "리테일전체" in sheets and "리테일낱개" in units
    retail = {r["B"]: r for r in sheets["리테일전체"]}
    assert "16-05" in retail
    assert abs(retail["16-05"]["H"] - 248400.0 / 5) <= 0.5          # F/D, 수수료 없음
    jt = {r["B"]: r for r in sheets["제이티전체"]}
    assert "31-22-02" in jt and abs(jt["31-22-02"]["H"] - 20200.0) <= 0.5


def test_optional_shipping_none():
    """배민·스스 미업로드(None) → 크래시 없이 스마트스토어·배민상회 G=선결제비 폴백."""
    baeju = [
        ("23-18", "웅진하늘보리", 1, 13700, 13700, "제이티유통", 0),
        ("20-90-11", "스마트상품", 3, 10000, 30000, "스마트스토어", 500),
        ("88-01", "배민상품", 2, 5000, 10000, "배민상회", 700),
    ]
    assert cy.open_baemin(None) is None and cy.open_sss(None) is None
    assert cy.process_baemin(None) == {} and cy.process_smartstore(None) == {}
    sheets, _ = cy.process(baeju, cy.open_baemin(None), cy.open_sss(None),
                           {}, {}, {}, RUN_DATE)
    assert sheets["제이티전체"]                                   # 발주자료만으로 분배
    assert sheets["스마트스토어전체"][0]["G"] == 500              # 선결제비 폴백
    assert sheets["배민상회전체"][0]["G"] == 700
