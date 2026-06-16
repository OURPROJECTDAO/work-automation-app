"""유형별매입현황(.xlsx) → 매입 이력(실입고·실매입가·리드타임 토대) 적재.

이력 엔진 (ADR 0018, intelligence-layer.md §3.3 / §4 purchases). 실입고 이벤트·실매입가·거래처.
- 저장: private repo work-automation-data : purchases/buyin_YYYY-MM.parquet (월 파티션).
- ★ 진짜 .xlsx(통파일은 72컬럼 풀 export·개별은 다른 컬럼) → openpyxl read_only 스트리밍 + **헤더명 매핑**(위치 시프트 무관).
- ★ PII 전제거: 사업자번호·거래처(약칭/상호)·우편번호·주소·휴대전화·전화·팩스·담당자. 거래처는 코드만 보존.
- 기준일 = 일자. 멱등 = 기준일 날짜구간 교체(고유 거래ID 없음 — orders/sales와 동일).
- 합계액 = 진실(수량×단가는 평균단가 반올림·할인·사은품으로 불일치 가능). 합계액<0 = 반품/조정 보존.

헤더명 매핑(공백 제거 후 정확 매칭. 통파일 '상품 관리코드' = 개별 '상품관리코드'):
  기준일←일자 · 거래처코드 · 관리코드←상품관리코드 · 상품명 · 규격 · 박스내품 · 박스 ·
  수량 · 단가 · 박스단가 · 공급가액 · 부가세 · 할인 · 합계액 · 대분류 · 중분류 · 소분류
"""
from __future__ import annotations

import base64
import io
import json
import re
import unicodedata
import urllib.error
import urllib.request

import openpyxl
import pandas as pd

PART_DIR = "purchases"
_PART_RE = re.compile(r"^buyin_(\d{4}-\d{2})\.parquet$")

# 정규화 헤더명 → 출력 컬럼명
_HDR_MAP = {
    "일자": "기준일",
    "거래처코드": "거래처코드",
    "상품관리코드": "관리코드",
    "상품명": "상품명",
    "규격": "규격",
    "박스내품": "박스내품",
    "박스": "박스",
    "수량": "수량",
    "단가": "단가",
    "박스단가": "박스단가",
    "공급가액": "공급가액",
    "부가세": "부가세",
    "할인": "할인",
    "합계액": "합계액",
    "대분류": "대분류",
    "중분류": "중분류",
    "소분류": "소분류",
}
COLS = ["기준일", "거래처코드", "관리코드", "상품명", "규격", "박스내품", "박스",
        "수량", "단가", "박스단가", "공급가액", "부가세", "할인", "합계액",
        "대분류", "중분류", "소분류"]
_STR_COLS = ("관리코드", "상품명", "규격", "대분류", "중분류", "소분류")
_NUM_COLS = ("박스내품", "박스", "수량", "단가", "박스단가", "공급가액", "부가세", "할인", "합계액")
_REQUIRED = ("일자", "상품관리코드", "수량", "단가", "합계액")


def _norm_hdr(v):
    return re.sub(r"\s+", "", unicodedata.normalize("NFC", str(v))) if v is not None else ""


def _nfc(v):
    return unicodedata.normalize("NFC", str(v)).strip() if pd.notna(v) else ""


def _num(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").strip()
    if s in ("", "-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _vendor_code(v):
    """거래처코드 → 정수 식별자(없으면 None). float(139.0)·str 모두 정규화."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, int):
        return v
    s = str(v).strip()
    if s in ("", "-"):
        return None
    try:
        f = float(s.replace(",", ""))
        return int(f) if f.is_integer() else None
    except ValueError:
        return s  # 비숫자 코드면 문자열 보존


def parse_purchases(file) -> pd.DataFrame:
    """유형별매입현황 .xlsx(경로/bytes) → 비-PII 매입 DataFrame(기준일·거래처코드·관리코드…).

    헤더명 매핑이라 위치 시프트(통파일 Source.Name 등)·컬럼 수와 무관.
    openpyxl read_only 스트리밍 — 필요 컬럼만 추출(대용량 통파일 메모리 안전).
    """
    src = io.BytesIO(file) if isinstance(file, (bytes, bytearray)) else file
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header = next(it)
        # 정규화 헤더 → 위치 인덱스 (중복 헤더는 첫 등장)
        pos = {}
        for i, h in enumerate(header):
            key = _norm_hdr(h)
            if key in _HDR_MAP and key not in pos:
                pos[key] = i
        missing = [k for k in _REQUIRED if k not in pos]
        if missing:
            raise ValueError("유형별매입현황 필수 헤더 누락: %s (헤더=%s)"
                             % (missing, [_norm_hdr(h) for h in header][:20]))
        idx = {out: pos[src_h] for src_h, out in _HDR_MAP.items() if src_h in pos}
        recs = []
        for r in it:
            rec = {out: (r[i] if i < len(r) else None) for out, i in idx.items()}
            recs.append(rec)
    finally:
        wb.close()
    out = pd.DataFrame(recs)
    for c in COLS:
        if c not in out.columns:
            out[c] = None
    out["거래처코드"] = pd.Series([_vendor_code(v) for v in out["거래처코드"]],
                              index=out.index, dtype="object")
    for c in _STR_COLS:
        out[c] = out[c].map(_nfc).astype("string")
    for c in _NUM_COLS:
        out[c] = out[c].map(_num)
    out["기준일"] = pd.to_datetime(out["기준일"], errors="coerce")
    out = out[out["기준일"].notna() & (out["관리코드"] != "")].copy()
    return out[COLS].sort_values(["기준일", "관리코드"]).reset_index(drop=True)


def split_by_month(df: pd.DataFrame) -> dict:
    return {ym: g.copy() for ym, g in df.groupby(df["기준일"].dt.strftime("%Y-%m"))}


def date_range_replace(existing, new: pd.DataFrame) -> pd.DataFrame:
    if existing is None or len(existing) == 0:
        return new.reset_index(drop=True)
    lo, hi = new["기준일"].min(), new["기준일"].max()
    keep = existing[(existing["기준일"] < lo) | (existing["기준일"] > hi)]
    return (pd.concat([keep, new], ignore_index=True)
              .sort_values(["기준일", "관리코드"]).reset_index(drop=True))


# ---- GitHub parquet R/W (work-automation-data) ----
def _gh(url, pat, method="GET", data=None, accept="application/vnd.github+json"):
    headers = {"Authorization": "Bearer " + pat, "User-Agent": "wa-app", "Accept": accept}
    if data is not None:
        headers["Content-Type"] = "application/json"
        data = json.dumps(data).encode("utf-8")
    return urllib.request.urlopen(urllib.request.Request(url, data=data, method=method, headers=headers))


def _part_path(ym):
    return "%s/buyin_%s.parquet" % (PART_DIR, ym)


def read_partition(pat, repo, ym) -> pd.DataFrame:
    url = "https://api.github.com/repos/%s/contents/%s" % (repo, _part_path(ym))
    try:
        with _gh(url + "?ref=main", pat, accept="application/vnd.github.raw") as r:
            return pd.read_parquet(io.BytesIO(r.read()))
    except urllib.error.HTTPError as e:
        if e.code == 404:
            return pd.DataFrame(columns=COLS)
        raise


def list_months(pat, repo) -> list:
    url = "https://api.github.com/repos/%s/contents/%s" % (repo, PART_DIR)
    try:
        items = json.load(_gh(url + "?ref=main", pat))
    except urllib.error.HTTPError as e:
        if e.code == 404:
            return []
        raise
    return sorted(m.group(1) for it in items for m in [_PART_RE.match(it.get("name", ""))] if m)


def read_all(pat, repo) -> pd.DataFrame:
    parts = [read_partition(pat, repo, ym) for ym in list_months(pat, repo)]
    parts = [p for p in parts if p is not None and len(p)]
    if not parts:
        return pd.DataFrame(columns=COLS)
    return pd.concat(parts, ignore_index=True).sort_values(["기준일", "관리코드"]).reset_index(drop=True)


def _write_partition(pat, repo, ym, df, msg):
    buf = io.BytesIO()
    df.to_parquet(buf, index=False)
    url = "https://api.github.com/repos/%s/contents/%s" % (repo, _part_path(ym))
    sha = None
    try:
        sha = json.load(_gh(url + "?ref=main", pat)).get("sha")
    except urllib.error.HTTPError as e:
        if e.code != 404:
            raise
    body = {"message": msg, "content": base64.b64encode(buf.getvalue()).decode("ascii")}
    if sha:
        body["sha"] = sha
    _gh(url, pat, method="PUT", data=body)


def ingest(df: pd.DataFrame, pat, repo) -> dict:
    """매입 DataFrame → 월별 파티션에 날짜구간 교체로 적재. 건드린 달만 재기록."""
    if df.empty:
        return {"rows": 0, "months": []}
    summary = []
    for ym, new in split_by_month(df).items():
        existing = read_partition(pat, repo, ym)
        before = len(existing)
        merged = date_range_replace(existing, new)
        _write_partition(pat, repo, ym, merged,
                         "data: buyin %s 적재 (%d→%d행)" % (ym, before, len(merged)))
        summary.append({"ym": ym, "before": before, "rows_in": len(new), "after": len(merged)})
    return {"rows": len(df), "months": summary}


def cadence_by_code(buyin) -> dict:
    """관리코드별 최근입고일 + 평균 입고주기(일) — 품절목록 E/F·재고 cadence.

    실입고(합계액>0 & 수량>0)·입고일 distinct 기준. 평균주기 = 연속 입고일 간격 평균(일, 입고 1회면 None).
    return {관리코드(NFC): {"최근입고일": Timestamp, "평균주기": float|None, "입고횟수": int}}.
    """
    if buyin is None or buyin.empty:
        return {}
    b = buyin.copy()
    b["_d"] = pd.to_datetime(b["기준일"], errors="coerce")
    q = pd.to_numeric(b["수량"], errors="coerce").fillna(0.0)
    t = pd.to_numeric(b["합계액"], errors="coerce").fillna(0.0)
    b = b[b["_d"].notna() & (q > 0) & (t > 0)].copy()
    if b.empty:
        return {}
    b["_code"] = b["관리코드"].map(_nfc)
    out = {}
    for code, g in b.groupby("_code"):
        if not code:
            continue
        days = sorted(set(g["_d"].dt.normalize()))
        last = days[-1]
        if len(days) >= 2:
            gaps = [(days[i] - days[i - 1]).days for i in range(1, len(days))]
            avg = float(sum(gaps) / len(gaps))
        else:
            avg = None
        out[code] = {"최근입고일": last, "평균주기": avg, "입고횟수": len(days)}
    return out
