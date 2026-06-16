"""당일 마진 점검 (두뇌① 탭D) — 당일 천년경영업로드 output + 송장출력 + master로
대략 실현마진 이상치를 즉시 탐지. intelligence-layer.md §6 ① 당일 렌즈 (사용자 설계 2026-06-15).

마진 3조각 (조인키 = 관리코드):
- 매출 = 천년경영 output **실제기입단가 × 수량** (= net·채널 수수료 적용·매출자료 proximate).
        raw '정산금액'(gross) 아님 — 쿠팡 정산×0.88·식봄 ×0.93·스마트 정산+선결제택배 → 실제기입단가가 net.
- 원가 = product_master **매입단가(낱개) × 낱개수량**.
- 택배 = **실제 물리 박스 단위 배분** — 송장출력을 박스(택배)로 재구성. 두 합포 시나리오 반영(2026-06-16):
    · 시나리오1(250/355 자동 합포분리): 이지어드민이 H열 한 행에 콤마로 2품목을 묶어 보냄 → 1박스,
      그 안 품목을 물류량(=수량÷박스내품)으로 분할. (★H열 콤마 다품목 전체 추출 — 둘째 품목 누락 버그 수정)
    · 시나리오2(175~200ml 30개입 합포가능 규격): 송장이 품목별로 쪼개져 보여도 같은 수령자끼리
      **최대 3팩까지 한 박스**로 물리 합포(맨 위 송장 1건). 합포가능 = reference/hapo_175_190.csv 관리코드.
      수령자 그룹 합포팩수 P → ceil(P/3) 박스, 팩수 비례로 품목에 택배 배분.
    · 그 외 일반 상품: 송장 1건 = 박스 1.
  관리코드 택배 = 배분박스수 × 채널 flat 택배단가. ★채널 물류량 비례배분 아님(과다계상 → 거짓 역마진).
- 이상 = 역마진(마진<0) OR 마진율 < 채널 baseline − buffer (탭C와 동일 기준).

★ 당일 raw 기반 **조기 트립와이어**. 정산 진실 = 매출자료 월정산(탭C). EasyAdmin/erp 정산은 raw(§2A).
★ PII: 송장출력 송장번호·수령자·주소는 박스 그룹키로만(원본 미저장). 상품명은 관리코드 추출에만.
"""
from __future__ import annotations

import io
import math
import re
import unicodedata

import pandas as pd

from core.intelligence.margin_erosion import EA_TO_CMM  # 송장 판매처(raw) → cmm 채널


def _nfc(v):
    return unicodedata.normalize("NFC", str(v)).strip() if v is not None and pd.notna(v) else ""


def _norm_code(c):
    """관리코드 정규화 — NFC + 트레일링 '.'/'-' 제거 (hapo 리스트 ↔ 추출코드 매칭용)."""
    return re.sub(r"[-.]+$", "", _nfc(c))


# 천년경영 output 시트 접두(전체/낱개 제거) → cmm 채널. 미매핑(멸치·11번가·셀러허브·제이티)=오프라인성 제외.
SHEET_TO_CMM = {
    "스마트스토어": "스마트스토어", "ESM": "ESM", "식봄": "식봄",
    "배민상회": "배민상회", "대용량": "배민상회",   # 대용량 = 배민대용량장보기
    "올웨이즈": "올웨이즈", "캐시노트": "캐시노트", "쿠팡": "쿠팡", "알리": "알리",
}

DEFAULT_FLAT = 2700  # 채널 flat 택배단가 기본값(인앱 편집). 문서화된 실택배비 표준.
HAPO_BOX_PACKS = 3   # 175~200ml 30개입 합포 한도 — 최대 3팩/박스 (사용자 확정 2026-06-16).

# 천년경영 전체: 0일자 1관리코드 2상품명 3주문수량 4평균단가 5정산금액 6선결제택배비 7실제기입단가
# 천년경영 낱개: ...3코드단위주문수량 ...7실제기입단가 8개당수량 9★기입수량 10★낱개단가
_C_CODE, _C_QTY, _C_REAL = 1, 3, 7
_N_PIECES, _N_UNIT = 9, 10

# 송장출력 상품명 → 관리코드 해소
_BR = re.compile(r"\[([^\[\]]+)\]")   # 가장 안쪽 [..] (이중 대괄호 [[code]..] 대응)
_PC = re.compile(r"^PC\d+$")           # PC낱개 → 상품코드
_QTY = re.compile(r"(\d+)\s*개\s*$")   # 상품명(조각) 끝 "N개"
# H열 다품목 구분: "...] N개, ..." — "개" 직후 콤마에서만 분리(상품명 내부 콤마 '190ml, 30개'는 보존)
_SPLIT = re.compile(r"(?<=개)\s*,\s*")


def _resolve_code(pname, valid_codes, pc_lookup):
    """송장 상품명(한 품목)에서 관리코드 추출. 모든 [..] 토큰을 후보로:
    PC낱개→상품코드(pc_lookup) · 바코드/이중괄호 접두는 앞 세그먼트를 떨궈가며 유효 관리코드(master) 대조.
    """
    if not pname:
        return None
    for raw in _BR.findall(pname):
        cand = raw.strip("[] ")
        if _PC.match(cand):
            r = pc_lookup.get(cand[2:], pc_lookup.get(cand))
            if r:
                return r
        c = _nfc(cand)
        if c in valid_codes:
            return c
        parts = c.split("-")               # 바코드 접두(GB130G12EA-·FH6-·BT10EA- 등) 떨구기
        for i in range(1, len(parts)):
            cc = "-".join(parts[i:])
            if cc in valid_codes:
                return cc
    return None


def _line_qty(seg):
    m = _QTY.search(seg or "")
    return float(m.group(1)) if m else 1.0


def _resolve_lines(pname, valid_codes, pc_lookup):
    """송장출력 H열 한 셀 → [(관리코드, 팩수)] 다품목.
    ★ H열 콤마 다품목(250/355 자동합포)을 '개' 뒤 콤마로 분리해 모든 품목 추출(둘째 누락 버그 수정).
    """
    out = []
    for seg in _SPLIT.split(str(pname or "")):
        code = _resolve_code(seg, valid_codes, pc_lookup)
        if code:
            out.append((code, _line_qty(seg)))
    return out


def _wb(file):
    from openpyxl import load_workbook
    if isinstance(file, (bytes, bytearray)):
        src = io.BytesIO(file)
    elif hasattr(file, "read"):
        src = io.BytesIO(file.read())
    else:
        src = file
    return load_workbook(src, data_only=True)   # ★ read_only 금지(천년경영 dim 오독, KB pitfalls)


def parse_invoice_shipping(file, box_lookup, pc_lookup=None, hapo_codes=None):
    """송장출력(.xlsx) → (택배 배분 dict, 채널 박스수 dict).

    물리 박스 = 택배 1건으로 재구성:
    - 합포가능 규격(hapo_codes, 175~200ml 30개입)만 든 송장 → (채널,수령자) 그룹으로 모아
      합포팩수 P → ceil(P/3) 박스, 팩수 비례 배분. (시나리오2 — 송장 분리돼도 물리 합포)
    - 그 외 송장 → 송장 1건=박스 1. 단일품목→풀박스, 합포(H열 다품목)→물류량 분할.
    return ({(채널, 관리코드): 배분박스수(소수)}, {채널: 박스(택배)수}).
    ★ 수령자·주소·송장번호는 그룹키로만(PII 미저장).
    """
    pc_lookup = pc_lookup or {}
    hapo = {_norm_code(c) for c in (hapo_codes or set())}
    valid_codes = set(box_lookup or {})
    wb = _wb(file)
    ws = wb[wb.sheetnames[0]]

    invoices: dict = {}   # 송장번호 → {ch, recv(수령자,주소), lines:[(code,packs)]}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r is None or len(r) < 10:
            continue
        ch = EA_TO_CMM.get(_nfc(r[3]))
        inv = r[9]
        if not ch or inv is None or _nfc(inv) == "":
            continue
        g = invoices.setdefault(_nfc(inv), {"ch": ch, "recv": (_nfc(r[5]), _nfc(r[6])), "lines": []})
        g["lines"].extend(_resolve_lines(r[7], valid_codes, pc_lookup))

    alloc: dict = {}
    ch_boxes: dict = {}
    hapo_grp: dict = {}   # (채널, 수령자) → [(code, packs)]

    for _inv, g in invoices.items():
        ch = g["ch"]
        lines = g["lines"]
        if lines and all(_norm_code(c) in hapo for c, _ in lines):
            # 시나리오2 — 합포송장: 수령자 그룹으로 미뤄 모음
            hapo_grp.setdefault((ch, g["recv"]), []).extend(lines)
            continue
        # 비합포 송장(시나리오1 포함) = 박스 1
        ch_boxes[ch] = ch_boxes.get(ch, 0) + 1
        if not lines:
            continue
        if len(lines) == 1:
            wts = {lines[0][0]: 1.0}
        else:  # H열 다품목(자동합포) → 물류량 분할
            raw: dict = {}
            for code, q in lines:
                bn = float(box_lookup.get(code, 1.0) or 1.0) or 1.0
                raw[code] = raw.get(code, 0.0) + q / bn
            tot = sum(raw.values()) or 1.0
            wts = {c: v / tot for c, v in raw.items()}
        for code, w in wts.items():
            alloc[(ch, code)] = alloc.get((ch, code), 0.0) + w

    # 시나리오2 — 합포 그룹: ceil(팩/3) 박스, 팩수 비례 배분
    for (ch, _recv), items in hapo_grp.items():
        total = sum(p for _, p in items)
        if total <= 0:
            continue
        nbox = math.ceil(total / HAPO_BOX_PACKS)
        ch_boxes[ch] = ch_boxes.get(ch, 0) + nbox
        bycode: dict = {}
        for code, p in items:
            bycode[code] = bycode.get(code, 0.0) + p
        for code, p in bycode.items():
            alloc[(ch, code)] = alloc.get((ch, code), 0.0) + nbox * (p / total)
    return alloc, ch_boxes


def parse_cheonnyeon_sales(file, box_lookup) -> pd.DataFrame:
    """천년경영업로드 output(.xlsx, 27시트) → (채널, 관리코드)별 net 매출·낱개수량.

    전체: 매출=실제기입단가×주문수량, 낱개수량=주문수량×박스내품.
    낱개: 매출=★낱개단가×★기입수량, 낱개수량=★기입수량.
    """
    wb = _wb(file)
    agg: dict = {}
    for sn in wb.sheetnames:
        base = sn.replace("전체", "").replace("낱개", "")
        ch = SHEET_TO_CMM.get(base)
        if ch is None:
            continue
        nat = sn.endswith("낱개")
        for r in wb[sn].iter_rows(min_row=2, values_only=True):
            if r is None or r[_C_CODE] is None:
                continue
            code = _nfc(r[_C_CODE])
            if not code:
                continue
            bn = float(box_lookup.get(code, 1.0) or 1.0) or 1.0
            if nat:
                pieces = float(r[_N_PIECES] or 0)
                net = float(r[_N_UNIT] or 0) * pieces
            else:
                qty = float(r[_C_QTY] or 0)
                pieces = qty * bn
                net = float(r[_C_REAL] or 0) * qty
            a = agg.setdefault((ch, code), {"매출": 0.0, "낱개수량": 0.0})
            a["매출"] += net
            a["낱개수량"] += pieces
    rows = [{"채널": ch, "관리코드": code, "매출": a["매출"], "낱개수량": a["낱개수량"]}
            for (ch, code), a in agg.items()]
    return pd.DataFrame(rows, columns=["채널", "관리코드", "매출", "낱개수량"])


def compute_daily_margin(sales_df: pd.DataFrame, box_alloc: dict, master_buy: dict,
                         name_lookup: dict, baseline_dict: dict,
                         flat_by_channel=None, buffer: float = 0.02) -> pd.DataFrame:
    """당일 (채널, 관리코드) 마진 + 이상 flag.

    box_alloc = {(채널,관리코드): 배분박스수} (parse_invoice_shipping). master_buy = 관리코드→매입단가(낱개).
    baseline_dict = 관리코드→{채널: 기준마진}. flat_by_channel = 채널→flat(결측=DEFAULT_FLAT).
    택배 = 배분박스수 × flat. 정렬 = 마진 오름차순(가장 손해 먼저). 이상행 거르기는 호출측.
    """
    cols = ["채널", "관리코드", "상품명", "매출", "낱개수량", "박스", "원가", "택배",
            "마진", "마진율", "기준마진", "역마진", "미달"]
    if sales_df is None or sales_df.empty:
        return pd.DataFrame(columns=cols)
    flat_by_channel = flat_by_channel or {}
    df = sales_df.copy()
    df["박스"] = [round(box_alloc.get((ch, c), 0.0), 2) for ch, c in zip(df["채널"], df["관리코드"])]
    df["택배"] = [box_alloc.get((ch, c), 0.0) * float(flat_by_channel.get(ch, DEFAULT_FLAT))
                 for ch, c in zip(df["채널"], df["관리코드"])]
    df["원가"] = [float(master_buy.get(c, 0.0)) * q for c, q in zip(df["관리코드"], df["낱개수량"])]
    df["마진"] = df["매출"] - df["원가"] - df["택배"]
    df["마진율"] = df["마진"] / df["매출"].where(df["매출"] > 0)
    df["상품명"] = df["관리코드"].map(lambda c: name_lookup.get(c, ""))
    df["기준마진"] = [(baseline_dict.get(c, {}) or {}).get(ch)
                   for c, ch in zip(df["관리코드"], df["채널"])]
    df["역마진"] = df["마진"] < 0
    df["미달"] = [(bm is not None) and pd.notna(mr) and (mr < bm - buffer)
                for bm, mr in zip(df["기준마진"], df["마진율"])]
    return df[cols].sort_values("마진").reset_index(drop=True)
