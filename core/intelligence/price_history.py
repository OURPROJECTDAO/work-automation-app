"""상품수정삭제로그 → 가격 변경 이력(매입단가/매출단가). 적재 + 역재생.

이력 엔진 첫 브릭 (ADR 0018, intelligence-layer.md Phase 1 1a).
- 저장: private repo work-automation-data : history/price_changes.parquet
- ⚠️ ERP 수정삭제로그는 조회일 기준 ~1년 롤링만 보관(3년 소급 불가) → 월 1회 재수신·dedup 누적이 핵심.
- dedup 키 = (상품코드, 수정항목, 수정일자). 재적재 멱등.
- 상품코드 = 정본 키(앞자리0 보존 문자열). 처리자(PII) 미저장.
"""
from __future__ import annotations

import base64
import io
import json
import unicodedata
import urllib.error
import urllib.request
import datetime as _dt

import pandas as pd

KEEP_ITEMS = ("매입단가", "매출단가")
HISTORY_PATH = "history/price_changes.parquet"
KEY = ["상품코드", "수정항목", "수정일자"]
COLS = ["상품코드", "관리코드", "상품명", "수정항목", "수정일자", "수정전", "수정후"]


def _nfc(v):
    return unicodedata.normalize("NFC", str(v)).strip() if v is not None else ""


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").strip()
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_price_log(file) -> pd.DataFrame:
    """수정삭제로그 .xlsx(경로/bytes/파일객체) → 매입단가·매출단가 변경만 정규화."""
    import openpyxl
    if isinstance(file, (bytes, bytearray)):
        file = io.BytesIO(file)
    wb = openpyxl.load_workbook(file, read_only=False, data_only=True)  # read_only 금지(pitfalls)
    ws = wb[wb.sheetnames[0]]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r is None or len(r) < 9:
            continue
        item = _nfc(r[6])
        if item not in KEEP_ITEMS:
            continue
        ts = r[4] if isinstance(r[4], _dt.datetime) else None
        out.append({
            "상품코드": _nfc(r[1]),
            "관리코드": _nfc(r[2]),
            "상품명": _nfc(r[3]),
            "수정항목": item,
            "수정일자": ts,
            "수정전": _num(r[7]),
            "수정후": _num(r[8]),
        })  # 처리자(PII)·사업장·수정위치·비고 제외
    df = pd.DataFrame(out, columns=COLS)
    df["수정일자"] = pd.to_datetime(df["수정일자"])
    for c in ("상품코드", "관리코드", "상품명", "수정항목"):
        df[c] = df[c].astype("string")
    return df.sort_values(["수정일자", "상품코드"]).reset_index(drop=True)


# ---- GitHub parquet R/W (work-automation-data) — 대시보드 store 패턴 ----
def _gh(url, pat, method="GET", data=None):
    headers = {"Authorization": "Bearer " + pat, "User-Agent": "wa-app",
               "Accept": "application/vnd.github+json"}
    if data is not None:
        headers["Content-Type"] = "application/json"
        data = json.dumps(data).encode("utf-8")
    return urllib.request.urlopen(urllib.request.Request(url, data=data, method=method, headers=headers))


def read_history(pat, repo) -> pd.DataFrame:
    url = "https://api.github.com/repos/%s/contents/%s" % (repo, HISTORY_PATH)
    try:
        meta = json.load(_gh(url + "?ref=main", pat))
        raw = base64.b64decode(meta["content"]) if meta.get("content") else _gh(meta["download_url"], pat).read()
        return pd.read_parquet(io.BytesIO(raw))
    except urllib.error.HTTPError as e:
        if e.code == 404:
            return pd.DataFrame(columns=COLS)
        raise


def ingest(new_df, pat, repo) -> dict:
    """기존 이력에 dedup-append 후 커밋. 멱등(KEY 중복 제거, keep=last)."""
    cur = read_history(pat, repo)
    before = len(cur)
    merged = pd.concat([cur, new_df], ignore_index=True)
    merged["수정일자"] = pd.to_datetime(merged["수정일자"])
    merged = (merged.drop_duplicates(KEY, keep="last")
                    .sort_values(["수정일자", "상품코드"]).reset_index(drop=True))
    buf = io.BytesIO()
    merged.to_parquet(buf, index=False)
    content = base64.b64encode(buf.getvalue()).decode("ascii")
    url = "https://api.github.com/repos/%s/contents/%s" % (repo, HISTORY_PATH)
    sha = None
    try:
        sha = json.load(_gh(url + "?ref=main", pat)).get("sha")
    except urllib.error.HTTPError as e:
        if e.code != 404:
            raise
    body = {"message": "data: price_changes 적재 (+%d행)" % (len(merged) - before), "content": content}
    if sha:
        body["sha"] = sha
    _gh(url, pat, method="PUT", data=body)
    return {"before": before, "added": len(merged) - before, "after": len(merged)}


# ---- 역재생: as-of 매입가/매출가 ----
def as_of_value(hist, 상품코드, 수정항목, when, current=None):
    """지정 시점(when)의 값. 앵커(current)는 호출자가 product_master에서 공급(낱개/박스 결정 — 검증 대기).

    규칙: when 이하 마지막 변경의 '수정후'. 그 이전이면 최초 변경의 '수정전'. 변경 없으면 current.
    """
    h = hist[(hist["상품코드"] == 상품코드) & (hist["수정항목"] == 수정항목)].sort_values("수정일자")
    if h.empty:
        return current
    le = h[h["수정일자"] <= pd.Timestamp(when)]
    if not le.empty:
        return float(le.iloc[-1]["수정후"])
    return float(h.iloc[0]["수정전"])
