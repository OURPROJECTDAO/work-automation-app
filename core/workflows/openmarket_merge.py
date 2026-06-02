"""
오픈마켓합포도서산간확인V7 → Python 재구현 (Phase 1)

[VBA 재현 주의사항]
- FasterCopyRows: ListSheet.Range("B1:B") — B1(헤더='주소')부터 읽으므로
  '주소' 자체가 키워드로 포함됨. '(상세주소 없음)' 등에 매칭 → 재현 필수.
- FasterCopyRows: 도서산간아님 예외 로직 없음 (해당 시트는 미사용).
- SortColumnBDescending: xlPinYin 정렬 → Python sort(ascending=False)로 근사.
- HighlightColumnC: ColorIndex 36(연노랑 #FFFF99) ↔ 35(연초록 #CCFFCC) 교대.
"""
from pathlib import Path
import io
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from core.base import Workflow, Step, WorkflowContext, normalize_kr
from core.workflows.registry import register

REF_DIR = Path(__file__).parent.parent.parent / "reference"

# VBA ColorIndex → HEX (엑셀 56색 팔레트)
_COLOR_HEX = {
    36: "FFFF99",  # 연노랑
    35: "CCFFCC",  # 연초록
}


class StepLoadInput(Step):
    """0. 입력(HTML-xls) 로드 + 기본 정제."""
    name = "입력_로드"
    def run(self, ctx: WorkflowContext) -> None:
        df = ctx.sheets['송장출력'].copy()
        df['송장번호'] = df['송장번호'].astype(str).str.replace(r'\.0$', '', regex=True)
        df['주소'] = df['주소'].apply(normalize_kr)
        df['상품명'] = df['상품명'].apply(normalize_kr)
        ctx.sheets['송장출력'] = df


class StepCopyDuplicates(Step):
    """1. 주소 중복 → 합포확인 (VBA: CopyDuplicatesToSummary)."""
    name = "합포_찾기"
    def run(self, ctx: WorkflowContext) -> None:
        df = ctx.sheets['송장출력']
        addr_counts = df['주소'].value_counts()
        dup_addrs = set(addr_counts[addr_counts > 1].index)
        df_dup = df[df['주소'].isin(dup_addrs)].copy()
        df_hapo = df_dup[['판매처', '수령자', '주소', '상품명', '송장번호']].rename(
            columns={'수령자': '수취인명'}
        )
        ctx.sheets['합포확인'] = df_hapo


class StepSortByAddress(Step):
    """2. 합포확인 주소 내림차순 정렬 (VBA: SortColumnBDescending, xlPinYin)."""
    name = "합포_정렬"
    def run(self, ctx: WorkflowContext) -> None:
        df = ctx.sheets['합포확인']
        df_sorted = df.sort_values('주소', ascending=False).reset_index(drop=True)
        ctx.sheets['합포확인'] = df_sorted


class StepColorGroups(Step):
    """3. 합포확인 색상 구분 (VBA: HighlightColumnC).

    같은 주소 그룹마다 ColorIndex 36(연노랑) ↔ 35(연초록) 교대.
    ctx.meta['hapo_colors'] = {row_idx: color_index} 로 저장.
    실제 셀 색칠은 _save() 에서 수행.
    """
    name = "합포_색상"

    def run(self, ctx: WorkflowContext) -> None:
        df = ctx.sheets['합포확인']
        colors: dict[int, int] = {}
        ci = 36          # VBA 초기값
        prev_addr = None
        for i, row in df.iterrows():
            addr = row['주소']
            if addr != prev_addr:
                if prev_addr is not None:
                    ci = 35 if ci == 36 else 36
                prev_addr = addr
            colors[i] = ci
        ctx.meta['hapo_colors'] = colors


class StepFilterProducts(Step):
    """4. 상품명 ↔ 필터링리스트 → 필터링확인 (VBA: FilterAndCopyRows)."""
    name = "필터링_확인"
    def run(self, ctx: WorkflowContext) -> None:
        from core.io_excel import load_csv_ref
        df = ctx.sheets['송장출력']
        fl = load_csv_ref(REF_DIR / 'filter_list.csv')
        keywords = [normalize_kr(k) for k in fl['상품명'].tolist() if k.strip()]
        mask = df['상품명'].apply(
            lambda p: any(kw in normalize_kr(str(p)) for kw in keywords)
        )
        ctx.sheets['필터링확인'] = df[mask].copy().reset_index(drop=True)


class StepDosanCheck(Step):
    """도서산간리스트 주소 매칭 → 지역확인 (VBA: FasterCopyRows).

    VBA 재현: B1(헤더='주소')부터 읽으므로 '주소' 키워드 포함.
    """
    name = "도서산간_확인"
    def run(self, ctx: WorkflowContext) -> None:
        from core.io_excel import load_csv_ref
        df = ctx.sheets['송장출력']
        ds = load_csv_ref(REF_DIR / 'dosan_list.csv')
        ds_kw = ['주소'] + [normalize_kr(k) for k in ds['주소'].tolist() if k.strip()]

        def is_dosan(addr):
            a = normalize_kr(str(addr))
            return any(kw in a for kw in ds_kw)

        ctx.sheets['지역확인'] = df[df['주소'].apply(is_dosan)].copy().reset_index(drop=True)


class StepUndeliveredCheck(Step):
    """미배송지리스트 주소 매칭 → 미배송지역확인 (VBA: mbCopyRows)."""
    name = "미배송지_확인"
    def run(self, ctx: WorkflowContext) -> None:
        from core.io_excel import load_csv_ref
        df = ctx.sheets['송장출력']
        mb = load_csv_ref(REF_DIR / 'undelivered_list.csv')
        mb_kw = [normalize_kr(k) for k in mb['미배송지 주소 리스트'].tolist() if k.strip()]

        def is_mb(addr):
            a = normalize_kr(str(addr))
            return any(kw in a for kw in mb_kw)

        ctx.sheets['미배송지역확인'] = df[df['주소'].apply(is_mb)].copy().reset_index(drop=True)


# ── 워크플로우 ────────────────────────────────────────────────────────────────

@register
class OpenmarketMergeWorkflow(Workflow):
    name = "오픈마켓_합포도서산간확인"
    steps = [
        StepLoadInput(),
        StepCopyDuplicates(),
        StepSortByAddress(),
        StepColorGroups(),
        StepFilterProducts(),
        StepDosanCheck(),
        StepUndeliveredCheck(),
    ]
    output_sheets = ["송장출력", "합포확인", "지역확인", "미배송지역확인", "필터링확인"]

    def _load(self, path: Path) -> dict:
        from core.io_excel import detect_and_load_input
        df = detect_and_load_input(path)
        return {'송장출력': df}

    def _save(self, ctx: WorkflowContext) -> Path:
        """기본 저장 후, 합포확인 시트에 그룹 색상 적용."""
        from core.io_excel import save_sheets
        out = ctx.output_dir / f"{self.name}_결과.xlsx"
        save_sheets(out, {k: ctx.sheets[k] for k in self.output_sheets if k in ctx.sheets})

        # 합포확인 색상 적용
        colors: dict = ctx.meta.get('hapo_colors', {})
        if colors:
            _apply_hapo_colors(out, '합포확인', colors)

        return out


def _apply_hapo_colors(xlsx_path: Path, sheet_name: str, colors: dict) -> None:
    """합포확인 시트 행별 배경색 적용.

    colors: {row_idx(0-based): color_index(35 or 36)}
    Excel 행번호 = row_idx + 2  (1행=헤더, 2행~=데이터)
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[sheet_name]
    n_cols = ws.max_column

    for row_idx, ci in colors.items():
        hex_color = _COLOR_HEX.get(ci, "FFFFFF")
        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        excel_row = row_idx + 2   # 헤더(1행) + 0-based offset
        for col in range(1, n_cols + 1):
            ws.cell(row=excel_row, column=col).fill = fill

    wb.save(xlsx_path)


def generate_invoice_xlsx(invoice_df: pd.DataFrame) -> bytes:
    """송장출력 시트 단독 xlsx (VBA SaveSheetToNewFile 복원).

    정제 후 '송장출력' 시트를 통째로 복사 → 단일시트 워크북.
    (원본 VBA: 송장출력 시트를 새 워크북에 복사 → ★★송장MMDD.xlsx 저장)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "송장출력"
    ws.append(list(invoice_df.columns))
    for _, row in invoice_df.iterrows():
        ws.append([row[c] for c in invoice_df.columns])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
