"""channel-margin-monitor — 채널 가격·마진 모니터 (코어 로직).

채널 상품관리 다운로드(라이브 리스팅)를 받아 상품별 마진율 계산 →
기준마진 대비 이탈(탐지) + 기준마진 달성 권장가 역산(100원 올림 → 기준마진 이상 보장).

판매자상품코드 4-tier 해석:
  박스(관리코드) / PC낱개(PC+상품코드) / 소분(변환코드-원코드) / 합포(코드1-CB-코드2).
매입가 = (코드해석 base) × N,  N = 판매자바코드(빈값/0→1, 분수 가능).
정산액 = 판매가net×(1-수수료) + 배송비×정산계수.
이익   = 정산액 - 매입가 - 실택배비.   마진율 = 이익/정산액.
권장가 = ⌈((매입가+실택배비)/(1-확정마진율) - 배송비×정산계수)/(1-수수료)⌉ (100원 올림).

reference: product_master.csv · baseline_margin.csv · sobun.csv · margin_floor.csv (app reference/).
공식·근거 = workflows/channel-margin-monitor.md. (검증: 2026-06-10 골든 705/706)
"""
from __future__ import annotations

import csv
import math
import random
import re
import unicodedata
import zipfile
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook

# ── 채널 config (채널 추가 = 여기에 한 세트) ────────────────────────────────
CHANNEL_CONFIG: dict[str, dict] = {
    "스마트스토어": {
        "key": "smartstore",       # 저장 파일명 reference/listing_<key>.csv
        "commission": 0.06,        # 판매수수료 → (1-수수료)=0.94 가 판매가에 곱
        "ship_settle": 0.967,      # 배송비 정산계수
        "real_ship": 2700,         # 실택배비 (단일)
        "baseline_col": "스마트스토어",  # baseline_margin.csv 의 채널 컬럼
        "apply_floor": True,       # 마진제한(하한 텍스트) 적용
        "sheet": None,             # None=첫 시트
        "header_row": 2,
        "data_start": 6,
        # 다운로드 컬럼 위치(1-indexed)
        "cols": {"상품번호": 1, "코드": 2, "상품명": 4, "판매가": 6,
                 "배송비": 41, "즉시할인": 58, "포인트": 69, "바코드": 78},
        "unitprice_use_col": 7,    # G 단위가격 사용여부: 양식 출력 시 비었으면 'N' 채움
    },
    "식봄": {
        "key": "sikbom",
        "commission": 0.07,        # 식봄 수수료 7% → (1-수수료)=0.93
        "ship_settle": 0.967,      # 배송비 정산계수 (스마트스토어 동일)
        "real_ship": 2700,         # 실택배비 (스마트스토어 기준 단일 — 골든 3000/3700 폐기)
        "ship_fee_const": 3000,    # 식봄 다운로드엔 '배송비명'뿐(숫자 없음) → income측 배송비 상수
        "baseline_col": "식봄",     # baseline_margin.csv 식봄 컬럼
        "apply_floor": True,
        "n_source": "ref",         # 합포량 N = hapo_multiplier(상품번호) — 다운로드 바코드 없음
        "sheet": "식봄붙여넣기",
        "header_row": 4,
        "data_start": 5,
        # 다운로드 컬럼(1-indexed). 정가=권장가 산출 시 정가≥판매단가 보존용. 즉시할인·포인트·배송비·바코드 없음
        "cols": {"상품번호": 1, "코드": 2, "상품명": 6, "판매가": 19, "정가": 16},
        # 가격 일괄변경 = 다운로드와 별개 '상품 일괄수정' 양식에 선택 행을 채워 넣는 append 방식
        "price_form": {
            "mode": "append",                       # 템플릿에 선택 행만 기입(스마트스토어=filter와 다름)
            "template": "sikbom_price_template.xlsx",  # reference/ 고정 양식
            "sheet": "(식봄)양식",
            "data_start": 7,                        # r1~3 안내·r4~6 헤더/설명
            "cols": {"상품번호": 1, "코드": 2, "상품명": 3, "정가": 4, "판매단가": 6},
            "fixed": {5: "n"},                      # E열 수량별 판매단가 설정 = n 고정
            "source": {"상품번호": "상품번호", "코드": "관리코드", "상품명": "상품명",
                       "정가": "정가", "판매단가": "권장가"},
            "price_field": "판매단가",
            "jeong_field": "정가",
        },
    },
    "캐시노트": {
        "key": "cashnote",
        "commission": 0.06,        # 6% (요율컬럼=6·천년경영 0.94·헤더 '6%기준' 근거. 골든 정산식 0.93은
                                   #   시트 내부 불일치로 미채택. 행사 차등수수료 0.88(12%)는 무시=단일 수수료, 사용자 확정 2026-06-11)
        "ship_settle": 0.967,      # 배송비 정산계수 (전 채널 동일)
        "real_ship": 2700,         # 실택배비 (스마트스토어 표준 단일 — 골든 3000/3700 미채택)
        "baseline_col": "캐시노트",  # baseline_margin.csv 캐시노트 컬럼
        "apply_floor": True,
        "n_source": "ref",         # 합포량 N = hapo_multiplier(상품번호) 채널무관 — 다운로드 바코드 없음
        "sheet": "상품",
        "header_row": 3,
        "data_start": 4,
        # 다운로드 컬럼(1-indexed). A=ID(상품번호)·E=입점사 관리 코드·C=상품명·N=판매 단가·O=할인 전 단가(정가).
        # 즉시할인·포인트·바코드 컬럼 없음(식봄형).
        "cols": {"상품번호": 1, "코드": 5, "상품명": 3, "판매가": 14, "정가": 15},
        # 배송비 = 배송정책코드(Y열=25) 조건부: DVP212991→3000, 그 외(DVP447716 등)→0. 골든 J식과 일치.
        "ship_fee_policy": {"col": 25, "map": {"DVP212991": 3000}, "default": 0},
        # 가격변경 양식(A=오퍼코드 OFR·D=옵션코드 SKU)이 다운로드 Q(17)·R(18)에만 있어 listing에 보존.
        "extra_cols": {"오퍼코드": 17, "옵션코드": 18},
        # 가격 일괄변경 = '(캐시노트)양식' append. F=수정·L=Y·N=9999 고정, G=판매단가(권장가)·H=할인전단가(≥판매단가).
        "price_form": {
            "mode": "append",
            "template": "cashnote_price_template.xlsx",   # reference/ 고정 양식(업로드 폼)
            "sheet": "(캐시노트)양식",
            "data_start": 4,                              # r1~3 그룹헤더/안내, r2=컬럼명
            "cols": {"오퍼코드": 1, "옵션코드": 4, "판매단가": 7, "할인전단가": 8, "관리코드": 15},
            "fixed": {6: "수정", 12: "Y", 14: 9999},       # F 변경타입·L 진열여부·N 재고수량
            "source": {"오퍼코드": "오퍼코드", "옵션코드": "옵션코드", "관리코드": "관리코드",
                       "할인전단가": "정가", "판매단가": "권장가"},
            "price_field": "판매단가",
            "jeong_field": "할인전단가",   # 무늬용 가짜 정가(표준 FAKE_JEONG). 마진은 항상 판매단가만 사용.
        },
    },
    "배민상회": {
        "key": "baemin",
        # 수수료 = 채널 단일값이 아니라 **상품별**(다운로드 BU열) → commission_source "download".
        #   commission = 수수료raw / commission_div + commission_add  (BU 4.5 → 4.5/100 + 0.03 = 0.075). 골든 J식과 일치.
        "commission_source": "download",
        "commission_field": "수수료raw",   # extra_cols로 캡처한 BU(73) 값
        "commission_div": 100,
        "commission_add": 0.03,           # 추가 고정수수료
        "ship_settle": 0.967,
        "real_ship": 2700,                # 스마트스토어 표준(골든 3000/3700 미채택)
        "baseline_col": "배민상회",
        "apply_floor": True,
        "n_source": "ref",                # 합포량 N = hapo_multiplier(상품번호) — 바코드 없음
        "sheet": "sheet1",
        "header_row": 2,
        "data_start": 3,
        # cols(1-indexed): A=상품번호·V(22)=관리코드(관리용 상품명)·B=상품명·X(24)=판매가·W(23)=소비자가(정가).
        # 즉시할인·포인트·바코드 컬럼 없음(식봄형).
        "cols": {"상품번호": 1, "코드": 22, "상품명": 2, "판매가": 24, "정가": 23},
        # 배송비 = 배송방법(BH=60) 조건부: "무료배송"→0, 그 외(택배배송 등)→3000. 골든 E식과 일치.
        "ship_fee_policy": {"col": 60, "map": {"무료배송": 0}, "default": 3000},
        # 상품별 수수료(BU=73) + 가격변경 양식용 옵션번호(R=18)·옵션명(U=21) listing 보존.
        "extra_cols": {"수수료raw": 73, "옵션번호": 18, "옵션명": 21},
        # 가격 일괄변경 = '(배민)양식' append. J=변경판매가(권장가). H=변경소비자가=무늬용 가짜(표준 FAKE_JEONG).
        "price_form": {
            "mode": "append",
            "template": "baemin_price_template.xlsx",   # reference/ 고정 양식
            "sheet": "(배민)양식",
            "data_start": 2,                            # r1=헤더
            "cols": {"상품번호": 1, "상품명": 2, "옵션번호": 3, "옵션명": 4, "관리코드": 5,
                     "현재소비자가": 7, "변경소비자가": 8, "현재판매가": 9, "변경판매가": 10},
            "source": {"상품번호": "상품번호", "상품명": "상품명", "옵션번호": "옵션번호",
                       "옵션명": "옵션명", "관리코드": "관리코드",
                       "현재소비자가": "정가", "현재판매가": "판매가"},
            "price_field": "변경판매가",
            "jeong_field": "변경소비자가",              # 무늬용 가짜 정가(표준 FAKE_JEONG)
        },
    },
    "쿠팡": {
        "key": "coupang",
        "commission": 0.12,               # 단일 12% (골든 ×0.88). 다운로드 수수료컬럼 없음
        "ship_settle": 0.967,
        "real_ship": 2700,                # 골든 3000/3700 미채택
        "ship_fee_const": 0,              # 배송비 항상 0(정산에 미반영)
        "baseline_col": "쿠팡",
        "apply_floor": True,
        "n_source": "ref",                # 합포 N = hapo_multiplier(옵션ID) — 키=상품번호=옵션ID
        "sheet": "data",
        "header_row": 3,
        "data_start": 4,
        # 키=옵션ID(C=3, 골든 조인키). 코드=업체상품코드(F=6). 상품명=쿠팡 노출 상품명(G=7).
        # 판매가=판매가격(J=10). 정가=할인율기준가(K=11). 즉시할인·포인트 없음.
        "cols": {"상품번호": 3, "코드": 6, "상품명": 7, "판매가": 10, "정가": 11},
        # E열(5)=바코드. 판매자택배 상품은 항상 공백, 값이 있으면 로켓그로스 → 판매자택배
        #   모니터 대상 아님(미매칭이 아니라 배송방식 차이) → parse 단계에서 행 자체 제외.
        "exclude_row_if_col_filled": 5,
        # 가격변경 = 다운로드의 '변경요청' 컬럼(P/Q)에 기입하는 filter형(스마트스토어식 원본편집).
        #   P(16)=변경 판매가(권장가), Q(17)=변경 할인율기준가(무늬용 가짜=표준 FAKE_JEONG). R/S(판매상태/재고)는 미기입.
        "price_form": {
            "mode": "filter",
            "write": {"판매가": 16, "정가": 17},
        },
    },
    "올웨이즈": {
        "key": "allways",
        "commission": 0.105,       # 10.5% 단일 (골든 정산 G=팀구매가×0.895 전건 검증)
        "ship_settle": 0.967,
        "real_ship": 2700,         # 스마트스토어 표준 (골든 3000·합포묶음+700 미채택)
        "ship_fee_const": 0,       # 배송비 항상 0 (골든 F 746/748=0, -700 2건 합포변칙 무시)
        "baseline_col": "올웨이즈",
        "apply_floor": True,
        "n_source": "ref",         # 합포 N = hapo_multiplier(올팜상품번호 ObjectId) — 골든 P열과 748/748 일치
        "sheet": None,             # 단일 시트
        "header_row": 1,
        "data_start": 4,           # r2=가이드·r3=예시 skip
        # cols(1-indexed): A=상품ID(키, ObjectId 문자열)·E(5)=판매자상품코드(관리코드)·B(2)=상품명·
        #   K(11)=팀구매가(판매가)·J(10)=개인구매가(정가). 즉시할인·포인트·배송비·바코드 컬럼 없음(식봄형).
        #   ※ 골든 D '일반가격(팀구매가*1.1)' 라벨은 거짓 — 실제 D=개인구매가(621/621 검증).
        "cols": {"상품번호": 1, "코드": 5, "상품명": 2, "판매가": 11, "정가": 10},
        # 가격변경 양식('(올웨이즈)양식')이 다운로드 전 컬럼 재업로드형(필수 多) → 양식 채울 컬럼 listing 보존.
        #   카테고리코드(C3)·판매상태(D4)·1옵션명/값(F6·G7)·재고수량(L12). 2옵션명/값(H/I)·옵션ID(M)은
        #   선택이고 단일옵션 상품은 항상 공백 → 미캡처(양식 H/I/M 공백 유효, stale 가드 거짓경고 방지).
        "extra_cols": {"카테고리코드": 3, "판매상태": 4, "옵션명1": 6, "옵션값1": 7, "재고수량": 12},
        # 가격 일괄변경 = '(올웨이즈)양식' append(식봄/캐시노트/배민형). 전 컬럼 재기입 + J/K 가격만 변경.
        #   K(팀구매가)=권장가(price_field), J(개인구매가)=무늬용 가짜(jeong_field=표준 FAKE_JEONG, 골든 O 가짜가격 대응).
        #   나머지 필수(상품명·카테고리·판매상태·1옵션명/값·재고)는 listing 보존값 그대로. H/I/M(선택)은 공백.
        "price_form": {
            "mode": "append",
            "template": "allways_price_template.xlsx",   # reference/ 고정 양식
            "sheet": "(올웨이즈)양식",
            "data_start": 4,                             # r1=헤더·r2=가이드·r3=예시
            "cols": {"상품ID": 1, "상품명": 2, "카테고리코드": 3, "판매상태": 4, "판매자상품코드": 5,
                     "옵션명1": 6, "옵션값1": 7, "개인구매가": 10, "팀구매가": 11, "재고수량": 12},
            "source": {"상품ID": "상품번호", "상품명": "상품명", "카테고리코드": "카테고리코드",
                       "판매상태": "판매상태", "판매자상품코드": "관리코드",
                       "옵션명1": "옵션명1", "옵션값1": "옵션값1", "재고수량": "재고수량"},
            "price_field": "팀구매가",      # K = 권장가
            "jeong_field": "개인구매가",     # J = 무늬용 가짜 정가(표준 FAKE_JEONG). 마진엔 미반영(표시용).
            "int_fields": ["카테고리코드", "재고수량"],   # 숫자 셀로 기입(텍스트 '103412' 방지)
        },
    },
    "알리": {
        "key": "ali",
        "commission": 0.09,        # 9% 단일 (골든 F=판매가×0.91 전건 검증)
        "ship_settle": 0.967,
        "real_ship": 2700,         # 스마트스토어 표준 (골든 3000·합포묶음+700 미채택)
        "ship_fee_const": 0,       # 배송비 항상 0 (골든 E 전건 0)
        "baseline_col": "알리",
        "apply_floor": True,
        "n_source": "ref",         # 합포 N = hapo_multiplier(알리상품번호) — 골든 N과 677/677 일치
        # 알리 전용 정제(매크로 ALI상품매크로V2 자동화): AliExpress 대량등록 export는
        #   카테고리별 다중시트(+각 *_hide 숨김시트, global_hide) + 다단헤더(r1 그룹·r2 라벨·r3 옵션필수·r4~5 설명/예시).
        #   보이는 카테고리 시트만 통합 → 헤더행(r2) 라벨로 4컬럼 추출(id·*제품 이름·*제품 소매 가격·SKU 코드).
        "consolidate": {
            "header_row": 2,
            "data_start": 5,          # 매크로 startRow=5. r5 '--' 예시행은 require_numeric_id로 제거.
            "skip_sheets": ["지침"],
            "require_numeric_id": True,  # 알리상품번호(16자리 숫자)만 — 예시 '--'/설명행 제외
            "labels": {               # 레코드필드 → r2 헤더 라벨
                "상품번호": "id",
                "상품명": "*제품 이름",
                "판매가": "*제품 소매 가격",
                "코드": "SKU 코드",
            },
        },
        # consolidate 채널은 cols/sheet 미사용(다중시트). 즉시할인·포인트·배송비·바코드·정가 없음.
        # 가격변경 미구현(AliExpress 가격/재고 업로드도 동일 다중시트 양식 → 별도 검증 후).
    },
    "esm": {
        "key": "esm",
        "commission": 0.175,       # 17.5% 단일 (골든 정산=가격×0.825=×(1−0.175) 정확 재현, 사용자 확정 2026-06-12).
                                   #   다운로드 O열 판매이용료(13/11/9%)는 카테고리 수수료라 미사용.
                                   #   골든 권장가 0.83·헤더 '17%기준'은 내부 불일치로 미채택(정산 우선).
        "ship_settle": 0.967,
        "real_ship": 2700,         # 스마트스토어 표준 단일 (골든 3000/3700·합포+700·ROUND 미채택)
        "baseline_col": "ESM",     # baseline_margin.csv ESM 컬럼
        "apply_floor": True,
        "n_source": "ref",         # 합포 N = hapo_multiplier(A 마스터상품번호·채널무관). 다운로드 바코드 없음.
                                   #   ⚠️ 골든 N은 #REF!(소스 N열 삭제) → N 자체는 골든 대조 불가(consolidation+입력으로 검증).
        "sheet": None,             # 다운로드 시트명 가변(worksheet1/합본) → 첫 시트
        "header_row": 1,
        "data_start": 2,
        # ESM(G마켓+옥션 통합) 다운로드는 1회 500상품 한도라 여러 배치로 받음. F(사이트)=='지마켓' 행만 모니터,
        #   A(마스터상품번호) 중복제거. parse가 자동 처리 → 수기 합치기·중복제거 불필요('신규만추가' 다회 or '전체교체' 합본).
        #   배송비 AI: '무료'→0, 그외 '3,000' 등 숫자(_num 콤마 허용). 즉시할인·포인트·바코드·정가 컬럼 없음.
        "include_row_if_col_value": {"col": 6, "value": "지마켓"},   # F=사이트
        "dedup_key": "상품번호",                                      # A 마스터상품번호(키=골든 조인+hapo N)
        "multi_file": True,        # 다운로드 500상품 한도 → 여러 배치. 업로더 다중파일 허용(이어붙이기+A중복제거 자동, 수기병합 불요)
        "cols": {"상품번호": 1, "코드": 5, "상품명": 3, "판매가": 9, "배송비": 35},
        # 가격변경 양식 키 = 사이트 상품번호(B, 다운로드 col2) — 모니터키 A(마스터)와 다름 → extra_cols로 listing 보존.
        "extra_cols": {"사이트상품번호": 2},
        # 가격변경 = append('(ESM)양식'): B=사이트상품번호·C=판매가(권장가). A=순번(데코). 정가 칸 없음(jeong 없음).
        #   템플릿 r1~5는 보호블록(다운로드 안내 '6행부터 입력') → data_start 6, 예시 r5 보존.
        "price_form": {
            "mode": "append",
            "template": "esm_price_template.xlsx",   # reference/ 고정 양식(클린 r1~5)
            "sheet": "(ESM)양식",
            "data_start": 6,
            "seq_col": 1,                            # A열 순번 1,2,3…
            "cols": {"사이트상품번호": 2, "판매가": 3},
            "source": {"사이트상품번호": "사이트상품번호"},
            "price_field": "판매가",                  # C = 권장가
            "int_fields": ["사이트상품번호"],          # B는 숫자셀(다운로드 동일)
            # jeong_field 없음 — 양식에 정가/할인전단가 칸 없음.
        },
    },
}


# 마진미달 판정 임계 (탐지 = 마진율 − 기준마진율 < 이 값). -0.01 = 기준보다 1%p↑ 낮음.
MARGIN_UNDER_THRESHOLD = -0.01

# 가격변경 양식 '무늬용 정가'(소비자가/할인전단가/정가) 표준 — 전 채널 공통:
# 권장가 × (1 + 랜덤 20~30%), 100원 반올림, >권장가. jeong_field 있는 채널은 기본 이 방식.
FAKE_JEONG = {"min_pct": 0.20, "max_pct": 0.30, "round": 100}


def _nfc(s) -> str:
    return unicodedata.normalize("NFC", str(s)).strip() if s not in (None, "") else ""


def _num(v, d: float = 0.0) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        if isinstance(v, str):                       # "3,000"(천단위 콤마) → 3000. "무료" 등 비숫자 → d (ESM 배송비)
            try:
                return float(v.replace(",", "").strip())
            except ValueError:
                return d
        return d


def _pid(v) -> str:
    """상품번호 정규화: 정수값 float(엑셀 숫자셀 46903.0)는 '46903'으로.

    캐시노트 등 다운로드의 상품번호(ID)가 숫자셀로 들어와 _nfc만 쓰면 '46903.0'이
    되어 hapo_multiplier 키('46903')·골든과 매칭 실패 → N=1 오류. 정수 float만 int화.
    """
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return _nfc(v)


def _deflo(s: str) -> str:
    """'510609.0' 같은 정수형 float 문자열 → '510609'. 그 외 원본 유지.

    구 listing에 숫자ID(옵션번호 등)가 엑셀 float로 저장돼 '510609.0'으로 남은 경우
    라운드트립에서 정수로 복원(재파싱 없이 양식 출력 정상화). 음수/일반 텍스트 불변.
    """
    return s[:-2] if re.fullmatch(r"-?\d+\.0", s) else s


def _strip_external_links(xlsx_bytes: bytes) -> bytes:
    """xlsx 패키지에서 외부 연결(externalLinks) 흔적을 제거 → 엑셀 '연결 업데이트' 경고 방지.

    일부 채널 양식 템플릿(배민)이 원본 마스터 통합문서를 가리키는 **고아 외부참조**를
    품고 있어, 그대로 저장하면 데이터(리터럴 값)는 멀쩡해도 열 때 외부링크 경고가 뜬다.
    수식이 외부참조를 실제로 쓰지 않으므로(전부 inlineStr/숫자) 안전하게 제거:
      ① xl/externalLinks/* 파트 ② workbook.xml <externalReferences>
      ③ workbook.xml.rels 의 externalLink 관계 ④ [Content_Types].xml override.
    외부링크가 없으면 무손실 no-op(바이트 그대로 반환).
    """
    zin = zipfile.ZipFile(BytesIO(xlsx_bytes))
    if not any(n.startswith("xl/externalLinks/") for n in zin.namelist()):
        return xlsx_bytes
    out = BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            n = item.filename
            if n.startswith("xl/externalLinks/"):
                continue
            data = zin.read(n)
            if n == "xl/workbook.xml":
                data = re.sub(rb"<externalReferences>.*?</externalReferences>", b"",
                              data, flags=re.S)
                data = data.replace(b"<externalReferences/>", b"")
            elif n == "xl/_rels/workbook.xml.rels":
                data = re.sub(rb"<Relationship[^>]*externalLink[^>]*/>", b"", data)
            elif n == "[Content_Types].xml":
                data = re.sub(rb"<Override[^>]*externalLink[^>]*/>", b"", data)
            zout.writestr(item, data)
    return out.getvalue()


def _pick_ws(wb, cfg):
    """cfg['sheet'] 가 있고 존재하면 그 시트, 아니면 첫 시트.

    채널 다운로드의 실제 시트명이 cfg와 다를 수 있어 첫 시트 폴백(예: 식봄 신규
    다운로드 시트명 != '식봄붙여넣기'). 식봄 다운로드는 단일 시트라 폴백 안전.
    """
    name = cfg.get("sheet")
    if name and name in wb.sheetnames:
        return wb[name]
    return wb[wb.sheetnames[0]]


# ── reference 로딩 ──────────────────────────────────────────────────────────
def load_references(ref_dir) -> dict:
    """app reference/ 에서 4종 로드 → dict."""
    ref_dir = Path(ref_dir)
    pm_by_mgmt: dict[str, dict] = {}
    pm_by_prod: dict[str, dict] = {}
    with open(ref_dir / "product_master.csv", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            mg, pr = _nfc(row.get("관리코드")), _nfc(row.get("상품코드"))
            if mg:
                pm_by_mgmt.setdefault(mg, row)
            if pr:
                pm_by_prod.setdefault(pr, row)

    def _load(name, key):
        d: dict[str, dict] = {}
        with open(ref_dir / name, encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                k = _nfc(row.get(key))
                if k:
                    d.setdefault(k, row)
        return d

    # 합포량(N) — 상품번호별 판매배수. 바코드 없는 채널 공용(마진율 예외). 파일 없으면 빈 dict.
    hapo: dict[str, float] = {}
    hp = ref_dir / "hapo_multiplier.csv"
    if hp.exists():
        with open(hp, encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                k = _nfc(row.get("상품번호"))
                if k and k not in hapo:
                    hapo[k] = _num(row.get("합포량"), 1.0)

    return {
        "pm_by_mgmt": pm_by_mgmt,
        "pm_by_prod": pm_by_prod,
        "sobun": _load("sobun.csv", "변환관리코드"),
        "baseline": _load("baseline_margin.csv", "관리코드"),
        "floor": _load("margin_floor.csv", "관리코드"),
        "hapo": hapo,
    }


# ── 코드 4-tier 해석 ────────────────────────────────────────────────────────
def resolve_code(code: str, refs: dict) -> tuple[str, float | None, float | None, str, str]:
    """returns (코드유형, base매입가, 재고, 규격, 비고).  base=None → 미매칭."""
    c = _nfc(code)
    if not c:
        return ("빈코드", None, None, "", "판매자상품코드 없음")
    pm_m, pm_p, sobun = refs["pm_by_mgmt"], refs["pm_by_prod"], refs["sobun"]
    # 1) 합포 (코드1-CB-코드2[-CB-코드3])
    if "-CB-" in c:
        prices, stocks, miss = [], [], []
        for p in c.split("-CB-"):
            r = pm_m.get(_nfc(p))
            if r:
                prices.append(_num(r["박스매입단가"]))
                stocks.append(_num(r["박스"]))
            else:
                miss.append(p)
        if miss:
            return ("합포", None, None, "", f"합포 구성코드 미등록: {','.join(miss)}")
        return ("합포", sum(prices) + 700, sum(stocks), "", "")
    # 2) 소분 (변환관리코드)
    if c in sobun:
        s = sobun[c]
        base = _nfc(s["원코드"])
        div = _num(s["내품나누기"], 0)
        r = pm_m.get(base)
        if not r or not div:
            return ("소분", None, None, _nfc(s.get("소분규격")), "소분 원코드 미등록/내품나누기 0")
        return ("소분", _num(r["박스매입단가"]) / div, _num(r["박스"]), _nfc(s.get("소분규격")), "")
    # 3) PC 낱개 (PC+상품코드)
    if c.upper().startswith("PC"):
        r = pm_p.get(_nfc(c[2:]))
        if not r:
            return ("낱개", None, None, "", f"상품코드 미등록: {c[2:]}")
        # 재고 = 그 상품코드 행의 **박스** 재고(낱개[15] 아님). 매입가는 낱개 매입단가 유지.
        return ("낱개", _num(r["매입단가"]), _num(r["박스"]), _nfc(r["규격"]), "")
    # 4) 박스 (관리코드)
    r = pm_m.get(c)
    if not r:
        return ("박스", None, None, "", "관리코드 미등록")
    return ("박스", _num(r["박스매입단가"]), _num(r["박스"]), _nfc(r["규격"]), "")


def _ceil100(x: float) -> int:
    return int(math.ceil(x / 100) * 100)


def fake_jeong(price: int, fake_cfg: dict | None = None) -> int:
    """무늬용 가짜 정가(소비자가/할인전단가/할인율기준가) — 전 채널 표준.

    권장가 × (1 + 랜덤 min~max%), round 단위 반올림, >권장가. fake_cfg로 %·단위 오버라이드.
    """
    fk = {**FAKE_JEONG, **(fake_cfg or {})}
    pct = random.uniform(fk["min_pct"], fk["max_pct"])
    unit = int(fk["round"])
    val = int(round(price * (1 + pct) / unit) * unit)
    return val if val > price else price + unit


def _ranges_desc(nums: list) -> list:
    """정렬 정수들 → 연속 구간 [(start,end)] 내림차순(아래부터 삭제용)."""
    if not nums:
        return []
    nums = sorted(set(nums))
    ranges, s, p = [], nums[0], nums[0]
    for n in nums[1:]:
        if n == p + 1:
            p = n
        else:
            ranges.append((s, p)); s = p = n
    ranges.append((s, p))
    return list(reversed(ranges))


# ── 다운로드 파싱 ───────────────────────────────────────────────────────────
def _consolidate_parse(wb, cfg: dict, con: dict) -> list[dict]:
    """알리 전용 정제: AliExpress 대량등록 export(카테고리별 다중시트+다단헤더)를
    매크로(ALI상품매크로V2: CopyDataFromAnotherWorkbook) 그대로 자동 통합.

    - 보이는 시트만(숨김 *_hide·global_hide 자동 제외), con['skip_sheets'] 제외.
    - con['header_row'] 행에서 라벨로 컬럼 위치 조회(시트마다 위치 달라도 안전).
    - con['data_start']부터 con['labels'] {레코드필드: 헤더라벨} 4종 추출.
    - require_numeric_id: 알리상품번호 비숫자(예시 '--' 행) 제외.
    배송비·즉시할인·포인트·정가·바코드 컬럼 없음 → 상수/0 처리(식봄형).
    """
    labels = con["labels"]
    hr, ds = con["header_row"], con["data_start"]
    skip = set(con.get("skip_sheets", []))
    req_num = con.get("require_numeric_id", False)
    ship_const = cfg.get("ship_fee_const")
    recs = []
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":      # _hide·global_hide(숨김) 제외 — 매크로 xlSheetVisible 동치
            continue
        if ws.title in skip:
            continue
        hmap: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            lbl = _nfc(ws.cell(hr, c).value)
            if lbl and lbl not in hmap:
                hmap[lbl] = c
        if not all(lab in hmap for lab in labels.values()):   # 라벨 누락 시트 skip(빈/비대상)
            continue
        for r in range(ds, ws.max_row + 1):
            raw_id = ws.cell(r, hmap[labels["상품번호"]]).value
            if raw_id in (None, ""):
                continue
            pid = _pid(raw_id)
            if req_num and not pid.isdigit():    # 예시행('--') 등 제외
                continue
            recs.append({
                "상품번호": pid,
                "코드": _nfc(ws.cell(r, hmap[labels["코드"]]).value),
                "상품명": _nfc(ws.cell(r, hmap[labels["상품명"]]).value),
                "판매가": _num(ws.cell(r, hmap[labels["판매가"]]).value),
                "배송비": float(ship_const) if ship_const is not None else 0.0,
                "즉시할인": 0.0, "포인트": 0.0, "정가": 0.0,
                "바코드": None, "오퍼코드": "", "옵션코드": "",
            })
    return recs


def parse_download(file, cfg: dict) -> list[dict]:
    """채널 상품관리 다운로드(.xlsx) → 레코드 리스트.

    채널별로 없는 컬럼(즉시할인·포인트·배송비·바코드)은 cfg['cols']에서 생략 가능
    → 0/상수/None 처리. 배송비 출처 3종(우선순위):
      ① cfg['ship_fee_const'] 상수(식봄) ② cfg['ship_fee_policy'] 배송정책코드 등
      컬럼값 조건부(캐시노트: DVP212991→3000, 그외→0) ③ cfg['cols']['배송비'] 숫자(스마트스토어).
    """
    src = BytesIO(file) if isinstance(file, (bytes, bytearray)) else file
    wb = load_workbook(src, data_only=True)  # read_only 금지(pitfalls)
    con = cfg.get("consolidate")
    if con:                                   # 알리: 카테고리별 다중시트+다단헤더 → 정제 통합(매크로 대체)
        return _consolidate_parse(wb, cfg, con)
    ws = _pick_ws(wb, cfg)
    col = cfg["cols"]
    ship_const = cfg.get("ship_fee_const")
    ship_policy = cfg.get("ship_fee_policy")  # {col, map, default} — 컬럼값 조건부 배송비(캐시노트)
    excl_col = cfg.get("exclude_row_if_col_filled")  # 그 컬럼에 값 있으면 행 제외(쿠팡 바코드=로켓그로스)
    incl = cfg.get("include_row_if_col_value")       # {col,value} 그 컬럼값==value 행만(ESM 사이트=지마켓)

    def _opt(r, key, default=0.0):
        c = col.get(key)
        return _num(ws.cell(r, c).value, default) if c else default

    def _ship(r):
        if ship_const is not None:
            return float(ship_const)
        if ship_policy:
            pol = _nfc(ws.cell(r, ship_policy["col"]).value)
            return float(ship_policy["map"].get(pol, ship_policy.get("default", 0)))
        return _opt(r, "배송비")

    recs = []
    for r in range(cfg["data_start"], ws.max_row + 1):
        pid = ws.cell(r, col["상품번호"]).value
        if pid in (None, ""):
            continue
        if excl_col is not None and ws.cell(r, excl_col).value not in (None, ""):
            continue                              # 로켓그로스(바코드 값 존재) → 판매자택배 모니터 제외
        if incl is not None and _nfc(ws.cell(r, incl["col"]).value) != incl["value"]:
            continue                              # ESM: 사이트(F)=='지마켓' 행만(옥션 제외)
        bc = col.get("바코드")
        rec = {
            "상품번호": _pid(pid),
            "코드": _nfc(ws.cell(r, col["코드"]).value),
            "상품명": _nfc(ws.cell(r, col["상품명"]).value),
            "판매가": _opt(r, "판매가"),
            "배송비": _ship(r),
            "즉시할인": _opt(r, "즉시할인"),
            "포인트": _opt(r, "포인트"),
            "정가": _opt(r, "정가"),
            "바코드": ws.cell(r, bc).value if bc else None,
            "오퍼코드": "", "옵션코드": "",          # 가격변경 양식 A/D용(캐시노트). 그 외 채널 공백
        }
        for name, c in cfg.get("extra_cols", {}).items():   # 다운로드 추가 컬럼 보존(OFR/SKU·옵션번호 등)
            rec[name] = _pid(ws.cell(r, c).value)           # 숫자ID(옵션번호) float '..0' 방지 — _pid 정수화
        recs.append(rec)
    dk = cfg.get("dedup_key")
    if dk:                                        # ESM: A(마스터상품번호) 중복제거(여러 배치 다운로드 합본·재업로드 안전)
        seen, uniq = set(), []
        for rec in recs:
            k = rec.get(dk)
            if k in seen:
                continue
            seen.add(k); uniq.append(rec)
        recs = uniq
    return recs


# ── 마진 계산 ───────────────────────────────────────────────────────────────
def compute(recs: list[dict], refs: dict, cfg: dict) -> list[dict]:
    settle, ship = cfg["ship_settle"], cfg["real_ship"]
    bcol = cfg["baseline_col"]
    apply_floor = cfg.get("apply_floor", True)
    comm_src = cfg.get("commission_source")   # "download"=상품별(BU) / None=단일 commission
    out = []
    for rec in recs:
        if comm_src == "download":             # 배민상회: 수수료raw/div + add (상품별)
            comm = _num(rec.get(cfg["commission_field"]), 0) / cfg.get("commission_div", 1) \
                + cfg.get("commission_add", 0.0)
        else:
            comm = cfg["commission"]
        rate = 1 - comm                        # 판매가net에 곱하는 정산비율
        typ, base, stock, spec, note = resolve_code(rec["코드"], refs)
        if cfg.get("n_source") == "ref":
            nv = refs.get("hapo", {}).get(rec["상품번호"], 1.0)  # 합포량(상품번호) 기본 1
            N = 1.0 if not nv else nv
        else:
            n_raw = _num(rec["바코드"], 0)
            N = 1.0 if n_raw == 0 else n_raw  # 빈값/0 → 1, 분수 허용
        row = {
            "상품번호": rec["상품번호"], "관리코드": rec["코드"], "상품명": rec["상품명"],
            "규격": spec, "코드유형": typ, "N": N, "재고": stock,
            "매입가": None, "판매가": rec["판매가"], "정가": rec.get("정가", 0), "배송비": rec["배송비"],
            "정산액": None, "마진율": None, "기준마진율": None, "탐지": None,
            "권장가": None, "제한": "", "비고": note,
        }
        # baseline 확정마진율 (판매자상품코드 직조인)
        bm = refs["baseline"].get(rec["코드"], {})
        bv = bm.get(bcol, "")
        base_margin = _num(bv, None) if bv not in (None, "") else None
        row["기준마진율"] = base_margin
        # 마진제한 텍스트
        fl = refs["floor"].get(rec["코드"]) if apply_floor else None
        if fl:
            row["제한"] = _nfc(fl.get("제한내용")) or _nfc(fl.get("비고"))

        if base is None:
            out.append(row)
            continue
        매입가 = base * N
        판매가net = rec["판매가"] - rec["즉시할인"] - rec["포인트"]
        정산액 = 판매가net * rate + rec["배송비"] * settle
        row["매입가"] = round(매입가)
        row["정산액"] = round(정산액)
        if 정산액 > 0:
            마진율 = (정산액 - 매입가 - ship) / 정산액
            row["마진율"] = round(마진율, 4)
            if base_margin is not None:
                row["탐지"] = round(마진율 - base_margin, 4)
        # 권장가 (기준마진 달성 판매가, 100원 올림 → 기준마진 이상 보장)
        if base_margin is not None and base_margin < 1:
            권장 = ((매입가 + ship) / (1 - base_margin) - rec["배송비"] * settle) / rate
            row["권장가"] = _ceil100(권장)
        out.append(row)
    return out


def _stats(rows: list[dict]) -> dict:
    margins = [r["마진율"] for r in rows if r["마진율"] is not None]
    return {
        "총건수": len(rows),
        "미매칭": sum(1 for r in rows if r["매입가"] is None),
        "미설정": sum(1 for r in rows if r["기준마진율"] is None and r["매입가"] is not None),
        "마진미달": sum(1 for r in rows if r["탐지"] is not None and r["탐지"] < MARGIN_UNDER_THRESHOLD),
        "제한상품": sum(1 for r in rows if r["제한"]),
        "평균마진율": round(sum(margins) / len(margins), 4) if margins else None,
    }


def compute_listing(recs: list[dict], channel: str, ref_dir) -> tuple[list[dict], dict]:
    """저장된 listing 레코드 + 채널 → (결과 레코드, 통계)."""
    if channel not in CHANNEL_CONFIG:
        raise ValueError(f"지원하지 않는 채널: {channel}")
    refs = load_references(ref_dir)
    rows = compute(recs, refs, CHANNEL_CONFIG[channel])
    return rows, _stats(rows)


def run(file, channel: str, ref_dir) -> tuple[list[dict], dict]:
    """다운로드(.xlsx) + 채널 → (결과 레코드, 통계)."""
    if channel not in CHANNEL_CONFIG:
        raise ValueError(f"지원하지 않는 채널: {channel}")
    recs = parse_download(file, CHANNEL_CONFIG[channel])
    return compute_listing(recs, channel, ref_dir)


# ── 저장 listing (연동데이터) 직렬화 / 병합 ──────────────────────────────────
LISTING_COLS = ["상품번호", "코드", "상품명", "판매가", "정가", "배송비", "즉시할인", "포인트", "바코드",
                "오퍼코드", "옵션코드"]


def recs_to_csv(recs: list[dict]) -> str:
    """parse_download 레코드 → CSV 텍스트 (저장용).

    LISTING_COLS + extra_cols로 들어온 추가 키(예 배민상회 수수료raw)를 자동 포함.
    """
    extra = []
    for r in recs:
        for k in r:
            if k not in LISTING_COLS and k not in extra:
                extra.append(k)
    cols = LISTING_COLS + extra
    buf = StringIO()
    w = csv.DictWriter(buf, fieldnames=cols, extrasaction="ignore")
    w.writeheader()
    for r in recs:
        bar = r.get("바코드")
        row = {k: r.get(k, "") for k in cols}
        row["바코드"] = "" if bar in (None, "") else bar
        w.writerow(row)
    return buf.getvalue()


def csv_text_to_recs(text: str) -> list[dict]:
    """저장 CSV 텍스트 → parse_download 호환 레코드 (추가 컬럼 포함, 구 CSV 하위호환)."""
    num_cols = {"판매가", "정가", "배송비", "즉시할인", "포인트"}
    recs = []
    for row in csv.DictReader(StringIO(text)):
        rec = {"상품번호": "", "코드": "", "상품명": "", "판매가": 0.0, "정가": 0.0,
               "배송비": 0.0, "즉시할인": 0.0, "포인트": 0.0, "바코드": "",
               "오퍼코드": "", "옵션코드": ""}
        for k, v in row.items():
            if k in num_cols:
                rec[k] = _num(v)
            elif k == "바코드":
                rec[k] = v or ""
            else:
                rec[k] = _deflo(_nfc(v))   # 구 listing의 '510609.0'(옵션번호 등) 정수 복원
        recs.append(rec)
    return recs


def merge_listing(existing: list[dict], new: list[dict]) -> tuple[list[dict], int]:
    """기존 + 신규: 기존 상품번호는 유지, 새 상품번호만 추가. → (병합, 추가건수)."""
    seen = {r["상품번호"] for r in existing}
    added = [r for r in new if r["상품번호"] and r["상품번호"] not in seen]
    return existing + added, len(added)


# ── 가격 일괄변경 (할인 우선 규칙) ───────────────────────────────────────────
def adjust_price(판매가: float, 즉시할인: float, 포인트: float,
                 target_net: float) -> tuple[int, int]:
    """net(=판매가-즉시할인-포인트)을 target_net으로 맞추는 (새 판매가, 새 즉시할인).

    ★ 할인 우선: 인상 시 즉시할인을 먼저 줄이고 모자라면 판매가를 올린다.
       인하 시 즉시할인을 먼저 늘리고 모자라면 판매가를 내린다. 포인트는 불변.
    """
    판매가, 즉시할인, 포인트 = float(판매가), float(즉시할인), float(포인트)
    cur_net = 판매가 - 즉시할인 - 포인트
    delta = target_net - cur_net
    new_price, new_disc = 판매가, 즉시할인
    if delta > 0:                                   # 인상: 할인 축소 우선
        cut = min(즉시할인, delta)
        new_disc = 즉시할인 - cut
        new_price = 판매가 + (delta - cut)
    elif delta < 0:                                 # 인하: 할인 확대 우선
        need = -delta
        room = max(cur_net, 0.0)                    # net 0까지만 할인 가능
        add = min(need, room)
        new_disc = 즉시할인 + add
        new_price = 판매가 - (need - add)
    return int(round(new_price)), int(round(new_disc))


def compute_new_prices(rows: list[dict], recs: list[dict],
                       pids: set) -> tuple[dict, list[str]]:
    """체크된 상품번호(pids) → {상품번호: (새 판매가, 새 즉시할인)} + 건너뛴 목록.

    target = 권장가(기준마진 달성가). 권장가 없는(미매칭/기준미설정) 상품은 skip.
    """
    rec_by = {r["상품번호"]: r for r in recs}
    row_by = {r["상품번호"]: r for r in rows}
    new_prices, skipped = {}, []
    for pid in pids:
        row, rec = row_by.get(pid), rec_by.get(pid)
        if not row or not rec or row.get("권장가") is None:
            skipped.append(pid)
            continue
        np_, nd_ = adjust_price(rec["판매가"], rec["즉시할인"], rec["포인트"],
                                row["권장가"])
        new_prices[pid] = (np_, nd_)
    return new_prices, skipped


def build_append_items(pf: dict, rows: list[dict], recs: list[dict],
                       pids) -> tuple[list[dict], list[dict], list[str]]:
    """append형 가격변경 양식의 (items, preview, skipped) 생성 — 채널 무관.

    pf['source'] {양식필드: 소스키}: row(우선)/rec 에서 값 추출.
    pf['price_field']: 권장가가 들어갈 양식필드(정수).
    pf['jeong_field']: (선택) 정가/할인전단가 필드 → max(소스값, 판매단가) 보장(정가≥판매가).
    권장가 없는(미매칭/기준 미설정) 상품은 skip.
    """
    row_by = {r["상품번호"]: r for r in rows}
    rec_by = {r["상품번호"]: r for r in recs}
    src = pf.get("source", {})
    price_f = pf["price_field"]
    jeong_f = pf.get("jeong_field")
    items, preview, skipped = [], [], []
    for pid in pids:
        ro = row_by.get(pid)
        if not ro or ro.get("권장가") is None:
            skipped.append(pid)
            continue
        merged = {**rec_by.get(pid, {}), **ro}       # row 우선
        price = int(ro["권장가"])
        it = {field: merged.get(key, "") for field, key in src.items()}
        it[price_f] = price
        if jeong_f:
            it[jeong_f] = fake_jeong(price, pf.get("jeong_fake"))   # 무늬용 가짜 정가(표준 FAKE_JEONG)
        items.append(it)
        cur = int(_num(ro.get("판매가")))
        preview.append({
            "상품명": ro.get("상품명"), "현재판매가": cur, "새판매단가": price,
            "정가": it.get(jeong_f) if jeong_f else "",
            "방향": "인상" if price > cur else ("인하" if price < cur else "유지"),
        })
    return items, preview, skipped


def build_price_form_append(template_xlsx: bytes, items: list[dict], pf: dict) -> bytes:
    """채널 '가격변경 양식' 템플릿에 선택 상품 행만 채워 append (식봄·캐시노트형).

    items: [{양식필드: 값}] — build_append_items 가 판매단가(=권장가)·정가/할인전단가까지
    계산해 넣는다. pf['cols'] {양식필드: 컬럼} 로 기입, pf['fixed'] {컬럼: 값} 고정값.
    템플릿의 기존/예시 데이터행은 모두 제거하고 data_start부터 기입.
    빈행 방지 위해 keep_last 초과 row_dimensions 정리(전역 pitfalls).
    """
    wb = load_workbook(BytesIO(template_xlsx))
    ws = wb[pf["sheet"]] if pf.get("sheet") else wb[wb.sheetnames[0]]
    cols = pf["cols"]
    start = pf["data_start"]
    fixed = pf.get("fixed", {})
    int_fields = set(pf.get("int_fields", []))       # 숫자로 기입할 양식필드(올웨이즈 카테고리코드·재고수량)
    seq_col = pf.get("seq_col")                       # (선택) 순번 컬럼 — 1,2,3… 기입(ESM A열 데코)
    if ws.max_row >= start:                          # 예시/기존 데이터행 제거
        ws.delete_rows(start, ws.max_row - start + 1)
    for i, it in enumerate(items):
        r = start + i
        if seq_col:
            ws.cell(r, int(seq_col)).value = i + 1
        for field, c in cols.items():
            if field in it:
                v = it[field]
                if field in int_fields and isinstance(v, str) and v.lstrip("-").isdigit():
                    v = int(v)
                ws.cell(r, c).value = v
        for c, val in fixed.items():                 # 고정값(예: 변경타입 '수정'·진열 'Y'·재고 9999)
            ws.cell(r, int(c)).value = val
    keep_last = start - 1 + len(items)
    for rr in [x for x in ws.row_dimensions if x > keep_last]:
        del ws.row_dimensions[rr]
    out = BytesIO()
    wb.save(out)
    return _strip_external_links(out.getvalue())  # 템플릿 딸린 고아 외부링크 제거(배민)


def build_bulk_price_xlsx(raw_xlsx: bytes, new_prices: dict,
                          cfg: dict) -> tuple[bytes, int, list[str]]:
    """원본 일괄변경 양식(raw_xlsx, 전체 컬럼) → 체크 상품 행만 가격 수정 후 남김.

    헤더(data_start 이전 행) 보존. new_prices 에 없는 데이터 행은 삭제.
    returns (xlsx bytes, 남긴 행수, 원본에 없던 상품번호 목록).
    """
    wb = load_workbook(BytesIO(raw_xlsx))           # 값+서식 보존 (read_only 금지)
    ws = _pick_ws(wb, cfg)
    col = cfg["cols"]
    c_pid, c_price, c_disc = col["상품번호"], col["판매가"], col["즉시할인"]
    c_unit = c_disc + 1                             # 즉시할인 단위(BG=BF+1)
    c_up = cfg.get("unitprice_use_col")             # 단위가격 사용여부(G) — 있을 때만
    start = cfg["data_start"]
    found, drop = set(), []
    for r in range(start, ws.max_row + 1):
        v = ws.cell(r, c_pid).value
        pid = _nfc(v) if v not in (None, "") else ""
        if pid and pid in new_prices:
            price, disc = new_prices[pid]
            ws.cell(r, c_price).value = price
            if disc and disc > 0:
                ws.cell(r, c_disc).value = disc
                ws.cell(r, c_unit).value = "원"
            else:
                ws.cell(r, c_disc).value = None
                ws.cell(r, c_unit).value = None
            if c_up and ws.cell(r, c_up).value in (None, ""):
                ws.cell(r, c_up).value = "N"        # 비었으면 N, 값 있으면 보존
            found.add(pid)
        else:
            drop.append(r)                          # 미체크 행 + 빈행 → 삭제(업로드 시 빈행 방지)
    for a, b in _ranges_desc(drop):                 # 연속 구간 묶어 아래부터 삭제
        ws.delete_rows(a, b - a + 1)
    # delete_rows가 남기는 빈 row_dimensions(빈 <row> 요소 유발) 정리 — 마지막 데이터행 이후 제거
    keep_last = (start - 1) + len(found)
    for rr in [x for x in ws.row_dimensions if x > keep_last]:
        del ws.row_dimensions[rr]
    out = BytesIO()
    wb.save(out)
    missing = [p for p in new_prices if p not in found]
    return _strip_external_links(out.getvalue()), len(found), missing


def append_rows_to_raw(raw_xlsx: bytes, src_xlsx: bytes,
                       pids: set, cfg: dict) -> bytes:
    """저장 원본(raw)에 src 양식의 신규 상품번호(pids) 행을 값으로 추가."""
    tgt = load_workbook(BytesIO(raw_xlsx))
    tws = _pick_ws(tgt, cfg)
    src = load_workbook(BytesIO(src_xlsx), data_only=True)
    sws = _pick_ws(src, cfg)
    c_pid = cfg["cols"]["상품번호"]
    start = cfg["data_start"]
    src_rows = {}
    for r in range(start, sws.max_row + 1):
        v = sws.cell(r, c_pid).value
        pid = _nfc(v) if v not in (None, "") else ""
        if pid:
            src_rows[pid] = r
    ncol = max(tws.max_column, sws.max_column)
    dest = tws.max_row + 1
    for pid in pids:
        sr = src_rows.get(pid)
        if not sr:
            continue
        for c in range(1, ncol + 1):
            tws.cell(dest, c).value = sws.cell(sr, c).value
        dest += 1
    out = BytesIO()
    tgt.save(out)
    return out.getvalue()


def _col_letter(n: int) -> str:
    """1→A, 16→P, 27→AA."""
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _sheet_part(zin: zipfile.ZipFile, cfg: dict) -> str:
    """cfg['sheet'] 이름 → 해당 worksheet xml 경로(xl/worksheets/sheetN.xml). 못 찾으면 첫 시트."""
    names = zin.namelist()
    want = cfg.get("sheet")
    try:
        wbxml = zin.read("xl/workbook.xml").decode("utf-8")
        rels = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        rid = None
        for m in re.finditer(r'<sheet\b[^>]*?/>', wbxml):
            tag = m.group(0)
            nm = re.search(r'name="([^"]*)"', tag)
            ridm = re.search(r'r:id="([^"]*)"', tag)
            if nm and ridm and (want is None or nm.group(1) == want):
                rid = ridm.group(1); break
        if rid:
            rm = re.search(rf'<Relationship\b[^>]*Id="{re.escape(rid)}"[^>]*/>', rels)
            if rm:
                tgt = re.search(r'Target="([^"]*)"', rm.group(0)).group(1)
                tgt = tgt.lstrip("/")
                if not tgt.startswith("xl/"):
                    tgt = "xl/" + tgt
                if tgt in names:
                    return tgt
    except KeyError:
        pass
    ws = sorted(n for n in names if re.match(r"xl/worksheets/sheet\d+\.xml$", n))
    return ws[0]


def _read_sst(zin: zipfile.ZipFile) -> list[str]:
    """sharedStrings.xml → 인덱스별 해소 문자열(모든 <t> 텍스트 연결, XML 언이스케이프)."""
    import html
    try:
        s = zin.read("xl/sharedStrings.xml").decode("utf-8")
    except KeyError:
        return []
    out = []
    for si in re.findall(r"<si>(.*?)</si>", s, re.S):
        txt = "".join(re.findall(r"<t[^>]*>(.*?)</t>", si, re.S))
        out.append(html.unescape(txt))
    return out


def _cell_in_row(row_xml: str, ref: str) -> str | None:
    """행 청크에서 셀(<c r="ref"...>...</c> 또는 <c r="ref".../>) 원문 반환. 없으면 None."""
    m = re.search(rf'<c r="{ref}"[^>]*?/>|<c r="{ref}"[^>]*?>.*?</c>', row_xml, re.S)
    return m.group(0) if m else None


def _cell_text(cell_xml: str, sst: list[str]) -> str:
    """셀 원문 → 표시 문자열(t="s"=sst 조회 / inlineStr / 숫자 / 빈칸)."""
    import html
    if cell_xml is None:
        return ""
    if 't="s"' in cell_xml:
        m = re.search(r"<v>(.*?)</v>", cell_xml, re.S)
        if m:
            i = int(m.group(1))
            return sst[i] if 0 <= i < len(sst) else ""
        return ""
    if 't="inlineStr"' in cell_xml:
        return html.unescape("".join(re.findall(r"<t[^>]*>(.*?)</t>", cell_xml, re.S)))
    m = re.search(r"<v>(.*?)</v>", cell_xml, re.S)
    return html.unescape(m.group(1)) if m else ""


def _set_num_cell(row_xml: str, ref: str, value: int) -> str:
    """행 청크의 ref 셀을 숫자값으로 기입(t 속성 없는 네이티브형 `<c r s><v>n</v></c>`). 스타일 보존.

    셀이 없으면 직전 존재 셀 뒤에 삽입(컬럼 순서). 쿠팡 변경요청 컬럼은 보통 빈 셀로 존재.
    """
    existing = _cell_in_row(row_xml, ref)
    col = re.match(r"[A-Z]+", ref).group(0)
    if existing is not None:
        sm = re.search(r'\ss="(\d+)"', existing)
        style = f' s="{sm.group(1)}"' if sm else ""
        new_cell = f'<c r="{ref}"{style}><v>{value}</v></c>'
        return row_xml.replace(existing, new_cell, 1)
    # 부재 → 컬럼 순서 삽입: 같은 행에서 ref보다 작은 마지막 셀 뒤
    want = _col_idx(col)
    last = None
    for m in re.finditer(r'<c r="([A-Z]+)\d+"[^>]*?(?:/>|>.*?</c>)', row_xml, re.S):
        if _col_idx(m.group(1)) < want:
            last = m
        else:
            break
    new_cell = f'<c r="{ref}"><v>{value}</v></c>'
    if last:
        return row_xml[:last.end()] + new_cell + row_xml[last.end():]
    return row_xml.replace("</row>", new_cell + "</row>", 1)


def _col_idx(letters: str) -> int:
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - 64)
    return n


def _renumber_row(row_xml: str, old: int, new: int) -> str:
    """행 청크의 모든 r 참조(<row r=> 와 <c r=Col###>)의 행번호 old→new."""
    return re.sub(rf'r="([A-Z]*){old}"', rf'r="\g<1>{new}"', row_xml)


def _inline_cells_to_shared(row_xml: str, sst_blocks: list, text2idx: dict) -> str:
    """행 청크의 inlineStr 셀을 t="s"(sharedStrings 참조)로 변환. 스타일 s= 보존.
    빈 inlineStr는 빈 스타일셀로. 엑셀 '값만 붙여넣기'가 하는 정규화를 코드가 수행 —
    원본 raw가 inlineStr(=openpyxl 오염/구 스냅샷)이어도 출력은 항상 네이티브 → 쿠팡 업로드 호환.
    sst_blocks(verbatim <si> 리스트)·text2idx(텍스트→인덱스)를 in-place 확장."""
    from xml.sax.saxutils import escape, unescape
    pat = re.compile(r'<c r="([A-Z]+\d+)"((?: s="\d+")?) t="inlineStr"><is>(.*?)</is></c>', re.S)

    def repl(m):
        ref, sattr, inner = m.group(1), m.group(2), m.group(3)
        text = unescape("".join(re.findall(r"<t[^>]*>(.*?)</t>", inner, re.S)))
        if text == "":
            return f'<c r="{ref}"{sattr}/>'           # 빈 셀
        idx = text2idx.get(text)
        if idx is None:
            idx = len(sst_blocks)
            text2idx[text] = idx
            sp = ' xml:space="preserve"' if text != text.strip() else ""
            sst_blocks.append(f"<si><t{sp}>{escape(text)}</t></si>")
        return f'<c r="{ref}"{sattr} t="s"><v>{idx}</v></c>'

    return pat.sub(repl, row_xml)


def build_filter_price_xlsx(raw_xlsx: bytes, rows: list[dict], pids,
                            cfg: dict) -> tuple[bytes, list[dict], list[str], list[str]]:
    """원본 다운로드(조회 + '변경요청' 컬럼형, 쿠팡)에서 선택 옵션만 남기고 변경요청 컬럼 기입.

    pf['write'] = {판매가: col, 정가: col}. P(판매가)=권장가, Q(정가/할인율기준가)=가짜정가(FAKE_JEONG).
    R/S(판매상태/재고)는 미기입(가격만 변경). 미선택 행 삭제. 키 = cfg['cols']['상품번호'](쿠팡 옵션ID).

    ★ 두 단계 네이티브 보장:
      (1) zip레벨 수술 — openpyxl load→save 금지(전 셀 inlineStr 변질 → 쿠팡 거부). 헤더행 보존 +
          선택 데이터행만 남겨 연속 재번호 + P/Q 숫자 기입. styles/mergeCells/네임스페이스/XML선언 원본 유지.
      (2) **inlineStr→sharedStrings 정규화** — 원본 raw가 inlineStr(구 스냅샷·openpyxl 오염·캐시
          stale)이어도 남긴 행 셀을 t="s"로 변환하고 sharedStrings.xml을 재구성 → **raw 상태와 무관하게
          출력은 항상 네이티브**(엑셀 '값만 붙여넣기' 동치). 쿠팡 업로더(POI 엄격형) 호환.
    """
    pf = cfg["price_form"]
    wp_col = _col_letter(pf["write"]["판매가"])         # P(16)
    wj_idx = pf["write"].get("정가")
    wj_col = _col_letter(wj_idx) if wj_idx else None      # Q(17)
    key_col = _col_letter(cfg["cols"]["상품번호"])        # C(3) 옵션ID
    start = cfg["data_start"]
    row_by = {r["상품번호"]: r for r in rows}
    targets, skipped = {}, []
    for pid in pids:
        ro = row_by.get(pid)
        if not ro or ro.get("권장가") is None:
            skipped.append(pid)
            continue
        price = int(ro["권장가"])
        jeong = fake_jeong(price, pf.get("jeong_fake")) if wj_col else None
        targets[pid] = (price, jeong, ro)

    zin = zipfile.ZipFile(BytesIO(raw_xlsx))
    sheet = _sheet_part(zin, cfg)
    sst = _read_sst(zin)                                  # 키 조회용(원본 인덱스 해소)
    sml = zin.read(sheet).decode("utf-8")

    # sharedStrings.xml verbatim <si> + 인덱스 맵 (inlineStr→t=s 변환 시 확장)
    try:
        sst_raw = zin.read("xl/sharedStrings.xml").decode("utf-8")
        m_open = re.search(r"<sst\b[^>]*>", sst_raw)
        sst_open = m_open.group(0)
        sst_decl = sst_raw[:m_open.start()] or '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
        sst_blocks = re.findall(r"<si>.*?</si>", sst_raw, re.S)
        has_sst = True
    except KeyError:
        sst_open = ('<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
                    'count="0" uniqueCount="0">')
        sst_decl = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
        sst_blocks = []
        has_sst = False
    text2idx = {}
    for i, si in enumerate(sst_blocks):
        t = "".join(re.findall(r"<t[^>]*>(.*?)</t>", si, re.S))
        from xml.sax.saxutils import unescape as _un
        text2idx.setdefault(_un(t), i)

    sd = re.search(r"<sheetData[^>]*>", sml)
    sd_close = sml.index("</sheetData>")
    prefix, body, suffix = sml[:sd.end()], sml[sd.end():sd_close], sml[sd_close:]
    row_chunks = re.findall(r"<row\b[^>]*>.*?</row>", body, re.S)

    out_rows, found, new_idx = [], set(), start
    for rx in row_chunks:
        r_old = int(re.search(r'<row r="(\d+)"', rx).group(1))
        if r_old < start:
            out_rows.append(rx)                          # 헤더행(안내·그룹·컬럼명) 그대로
            continue
        key = _cell_text(_cell_in_row(rx, f"{key_col}{r_old}"), sst).strip()
        key = _deflo(key)
        if key not in targets:
            continue                                     # 미선택 행 삭제
        price, jeong, _ = targets[key]
        found.add(key)
        rx = _inline_cells_to_shared(rx, sst_blocks, text2idx)   # ★ inlineStr→sharedStrings
        rx = _set_num_cell(rx, f"{wp_col}{r_old}", price)
        if wj_col and jeong is not None:
            rx = _set_num_cell(rx, f"{wj_col}{r_old}", jeong)
        out_rows.append(_renumber_row(rx, r_old, new_idx))
        new_idx += 1

    last_row = new_idx - 1
    new_sml = prefix + "".join(out_rows) + suffix
    new_sml = re.sub(r'(<dimension ref="[A-Z]+\d+:[A-Z]+)\d+("\s*/>)',
                     rf"\g<1>{last_row}\g<2>", new_sml)

    # sharedStrings.xml 재구성(count=시트 t=s 참조수 / uniqueCount=si 개수)
    n_refs = new_sml.count('t="s"')
    if "count=" in sst_open:
        sst_open2 = re.sub(r'count="\d+"', f'count="{n_refs}"', sst_open)
    else:
        sst_open2 = sst_open[:-1] + f' count="{n_refs}">'
    if "uniqueCount=" in sst_open2:
        sst_open2 = re.sub(r'uniqueCount="\d+"', f'uniqueCount="{len(sst_blocks)}"', sst_open2)
    else:
        sst_open2 = sst_open2[:-1] + f' uniqueCount="{len(sst_blocks)}">'
    sst_new = sst_decl + sst_open2 + "".join(sst_blocks) + "</sst>"

    out = BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        wrote_sst = False
        for item in zin.infolist():
            fn = item.filename
            if fn == sheet:
                data = new_sml.encode("utf-8")
            elif fn == "xl/sharedStrings.xml":
                data = sst_new.encode("utf-8"); wrote_sst = True
            else:
                data = zin.read(fn)
            zout.writestr(item, data)
        if not wrote_sst and has_sst is False and sst_blocks:
            # 원본에 sst 파트가 없었으나 새로 생성해야 하는 경우(쿠팡에선 미발생)
            zout.writestr("xl/sharedStrings.xml", sst_new.encode("utf-8"))

    prev = []
    for pid in pids:
        if pid not in found:
            continue
        price, jeong, ro = targets[pid]
        cur = int(_num(ro.get("판매가")))
        prev.append({"상품명": ro.get("상품명", ""), "현재판매가": cur, "새판매가": price,
                     "정가": jeong, "권장가(net)": ro.get("권장가"),
                     "방향": "인상" if price > cur else ("인하" if price < cur else "유지")})
    missing = [p for p in targets if p not in found]
    return out.getvalue(), prev, skipped, missing
