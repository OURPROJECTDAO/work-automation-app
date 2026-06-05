"""
온누리양식_발주서 워크플로우.

입력: 발주서 xlsx (관리코드, 총 주문 수량 등 포함)
처리: SKU 참조 → 합계:판매가(부가세 포함) 계산
출력: 원본파일명(확인).xlsx

수식: 합계 = 공급가(VAT포함) × 수량 + ceil(수량 / 최대합포수량) × 배송비
참조: reference/sku_list.csv

[수정 2026-06-04] _save 방식 변경: openpyxl save → zipfile 직접 조작
  openpyxl save 시 sharedString(t="s") → inlineStr(t="inlineStr") 변환 발생,
  일부 외부 시스템에서 헤더 인식 불가 → zipfile로 sheet1.xml의 합계 열만 패치,
  원본 sharedStrings 구조 완전 유지.
"""
import io
import math
import re
import zipfile
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from core.base import Workflow, Step, WorkflowContext
from core.workflows.registry import register

_REF = Path(__file__).parent.parent.parent / "reference"
_COL_TOTAL = "합계 : 판매가(부가세 포함)"
_COL_CODE = "관리코드"
_COL_QTY = "총 주문 수량"


def _col_num_to_letter(n: int) -> str:
    """열 번호(1-based)를 열 문자로 변환. 예: 1→A, 7→G."""
    result = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def _col_letter_to_num(letters: str) -> int:
    """열 문자를 번호(1-based)로 변환. 예: A→1, G→7."""
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - 64)
    return n


def _insert_cell_into_row(content: str, row_num: int, col_letter: str,
                          cell_ref: str, int_val: int) -> str:
    """해당 <row>에 열 순서를 지켜 새 셀 삽입. 행을 못 찾으면 그대로 반환."""
    target_col = _col_letter_to_num(col_letter)
    new_cell = f'<c r="{cell_ref}"><v>{int_val}</v></c>'
    row_pat = re.search(rf'(<row r="{row_num}"[^>]*>)(.*?)(</row>)',
                        content, re.DOTALL)
    if not row_pat:
        return content
    head, body, tail = row_pat.group(1), row_pat.group(2), row_pat.group(3)
    insert_at = len(body)
    for cm in re.finditer(r'<c r="([A-Z]+)\d+"', body):
        if _col_letter_to_num(cm.group(1)) > target_col:
            insert_at = cm.start()
            break
    new_body = body[:insert_at] + new_cell + body[insert_at:]
    return content[:row_pat.start()] + head + new_body + tail + content[row_pat.end():]


def _patch_column_values(sheet_xml: bytes, col_letter: str, values: list) -> bytes:
    """
    sheet1.xml에서 특정 열의 데이터 행(2행 이후) 값만 수정.
    - 원본 sharedStrings 구조(t="s") 변경 없음, 기존 셀 스타일(s) 보존
    - 값 있는 셀(<c ...>...</c>): 값만 교체
    - 자체닫힘 빈 셀(<c .../>): 값 있는 셀로 교체  [버그픽스 2026-06-05]
    - 셀 자체가 없는 행: 열 순서 맞춰 새 셀 삽입      [버그픽스 2026-06-05]
    (기존 정규식은 </c> 가진 셀만 매칭 → 빈 합계 셀을 못 잡거나,
     자체닫힘 셀에서 다음 행까지 과매칭하여 데이터 손상시킴.)
    """
    content = sheet_xml.decode("utf-8")
    for i, val in enumerate(values):
        if val is None:
            continue
        row_num = i + 2  # 행1=헤더, 행2부터 데이터
        cell_ref = f"{col_letter}{row_num}"
        int_val = int(val)

        # 값셀(>...</c>) 또는 자체닫힘(/>) 둘 다 매칭.
        # 템퍼드 패턴으로 첫 </c> 를 넘어 다음 셀까지 먹지 않게 함.
        m = re.search(
            rf'<c r="{re.escape(cell_ref)}"([^>]*?)(?:/>|>(?:(?!</c>).)*?</c>)',
            content, re.DOTALL)
        if m:
            s_match = re.search(r's="(\d+)"', m.group(1))
            s_attr = f' s="{s_match.group(1)}"' if s_match else ""
            new_cell = f'<c r="{cell_ref}"{s_attr}><v>{int_val}</v></c>'
            content = content[: m.start()] + new_cell + content[m.end():]
        else:
            content = _insert_cell_into_row(
                content, row_num, col_letter, cell_ref, int_val)
    return content.encode("utf-8")


class LoadSKU(Step):
    """SKU 참조 CSV 로드 → ctx.meta['sku'] (관리코드 인덱스)."""
    name = "load_sku"

    def run(self, ctx: WorkflowContext) -> None:
        sku = pd.read_csv(
            _REF / "sku_list.csv",
            encoding="utf-8-sig",
            dtype={"관리코드": str, "원코드": str},
        )
        sku.drop_duplicates("관리코드", keep="first", inplace=True)
        ctx.meta["sku"] = sku.set_index("관리코드")


class CalcTotal(Step):
    """합계:판매가 = 공급가 × 수량 + ceil(수량/최대합포) × 배송비."""
    name = "calc_total"

    def run(self, ctx: WorkflowContext) -> None:
        df = ctx.sheets["발주서"]
        sku = ctx.meta["sku"]

        def _calc(row):
            code = str(row[_COL_CODE]) if row[_COL_CODE] is not None else ""
            if not code or code not in sku.index:
                return None
            qty = int(row[_COL_QTY])
            s = sku.loc[code]
            price = int(s["공급가(VAT 포함)"])
            max_bundle = int(s["배송비 부과 규칙 (규격 기준)"])
            ship = int(s["배송비"])
            n = math.ceil(qty / max_bundle)
            return price * qty + n * ship

        df[_COL_TOTAL] = df.apply(_calc, axis=1)


@register
class OnnuriOrderWorkflow(Workflow):
    """온누리양식 발주서 처리 워크플로우."""

    name = "온누리양식_발주서"
    steps = [LoadSKU(), CalcTotal()]
    output_sheets = ["발주서"]

    def _load(self, path: Path) -> dict:
        """openpyxl로 읽어 원본 셀 값(문자열 우편번호 등) 보존."""
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return {"발주서": pd.DataFrame()}
        return {"발주서": pd.DataFrame(rows[1:], columns=rows[0])}

    def _save(self, ctx: WorkflowContext) -> Path:
        """
        zipfile 직접 조작으로 합계 컬럼만 패치 → sharedString 원본 구조 완전 유지.

        기존 openpyxl save 방식은 모든 문자열 셀을 sharedString(t="s")에서
        inlineStr(t="inlineStr")으로 변환하여 외부 시스템 헤더 인식 불가 문제 발생.
        """
        stem = ctx.input_path.stem
        out = ctx.output_dir / f"{stem}(확인).xlsx"

        df = ctx.sheets["발주서"]

        # 합계 컬럼 위치 파악
        wb = load_workbook(ctx.input_path, read_only=True)
        ws = wb.active
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        wb.close()
        try:
            col_idx = header.index(_COL_TOTAL) + 1  # 1-based
        except ValueError:
            raise ValueError(f"'{_COL_TOTAL}' 컬럼을 발주서 시트에서 찾지 못했습니다.")

        col_letter = _col_num_to_letter(col_idx)
        total_values = df[_COL_TOTAL].tolist()

        # zipfile로 xlsx 직접 조작 (sharedStrings 원본 유지)
        out_buf = io.BytesIO()
        with zipfile.ZipFile(ctx.input_path, "r") as zin:
            with zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.namelist():
                    data = zin.read(item)
                    if item == "xl/worksheets/sheet1.xml":
                        data = _patch_column_values(data, col_letter, total_values)
                    zout.writestr(zin.getinfo(item), data)

        out.write_bytes(out_buf.getvalue())
        return out
