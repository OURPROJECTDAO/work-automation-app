"""
발주서 출력 업무 워크플로우.

입력 : 판매처상품매출통계 .xls (HTML), 상품관리 .xlsx (일일 재고)
처리 : Phase1(정제+분류) → Phase2(재고대조+출력)
출력 : 발주자료 아카이브 xlsx, 최종결과물 xlsx (물류팀+품절목록)
참조 : reference/logistics_classification.csv
       reference/unit_list.csv
       reference/spec_master.csv
       reference/product_master.csv  ← 연동데이터관리에서 매일 갱신
"""
import io
import math
import re
import unicodedata
from datetime import datetime, timezone, timedelta

_KST = timezone(timedelta(hours=9))  # 한국 표준시 UTC+9
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

_REF = Path(__file__).parent.parent.parent / "reference"

# erp관리코드 패턴: XX-XX-XX (각 부분 정확히 2자리) — 구버전 합포 fallback용
_ERP_CODE_RE = re.compile(r'\d{2}-\d{2}-\d{2}')
# 어드민옵션 안의 [erp코드] (영숫자·하이픈·점, 포맷 무관) — 합포 감지·분리 2순위 기준
_BRACKET_CODE_RE = re.compile(r'\[([A-Z0-9][A-Z0-9.\-]*)\]')
# 합포 셀의 중첩테이블(tableGridA) 행 경계 보존용 구분자 (parse → split 1순위)
_SENTINEL = "\x1f"

# ───────────────────────────────────────────────
# Reference 로딩
# ───────────────────────────────────────────────

def load_classification() -> pd.DataFrame:
    return pd.read_csv(_REF / "logistics_classification.csv",
                       encoding="utf-8-sig", dtype=str)

def load_unit_list() -> pd.DataFrame:
    return pd.read_csv(_REF / "unit_list.csv",
                       encoding="utf-8-sig", dtype={"관리코드": str, "원코드": str})


def unit_origin_map(unit_df: pd.DataFrame = None) -> dict:
    """낱개/소분 코드 → 원코드(박스 관리코드) 매핑 (NFC·자기자신 제외).

    ★ 매입현황(buyin cadence)·상품관리(박스재고)는 전부 **박스 관리코드** 기준 —
      낱개/소분 코드로 조회하면 무조건 미매칭(공백)이라 원코드로 치환해야 함.
      재고 대조(reconcile_stock)가 쓰는 원코드와 동일 기준.
    """
    if unit_df is None:
        unit_df = load_unit_list()
    out = {}
    for _, r in unit_df.iterrows():
        if pd.isna(r.get("관리코드")) or pd.isna(r.get("원코드")):
            continue
        k = unicodedata.normalize("NFC", str(r["관리코드"]).strip())
        v = unicodedata.normalize("NFC", str(r["원코드"]).strip())
        if k and v and k != v:
            out[k] = v
    return out

def load_spec_master() -> pd.DataFrame:
    return pd.read_csv(_REF / "spec_master.csv",
                       encoding="utf-8-sig", dtype={"관리코드": str})

def load_product_master() -> pd.DataFrame:
    path = _REF / "product_master.csv"
    if not path.exists():
        return pd.DataFrame()
    return pd.read_csv(path, encoding="utf-8-sig", dtype=str)

def get_product_master_updated() -> str:
    meta = _REF / "product_master_updated.txt"
    if meta.exists():
        return meta.read_text(encoding="utf-8").strip()
    return ""

# ───────────────────────────────────────────────
# Phase 1 — 정제 + 분류
# ───────────────────────────────────────────────

_COLS = ["erp관리코드", "어드민옵션", "총수량", "평균단가",
         "정산금액", "판매처그룹", "선결제택배비", "옵션추가항목1"]


def _extract_cell(td) -> list:
    """셀 텍스트 추출. 중첩 tableGridA가 있으면 행별 텍스트 리스트, 없으면 [단일 텍스트].

    ERP 매출통계 HTML은 합포(다른 상품)를 셀 안 중첩 테이블(tableGridA)의
    여러 <tr>로 표현한다(코드/상품명이 행 분리). pd.read_html은 이를 이어붙여
    뭉개므로, 여기서 각 행을 분리 보존해 합포 분리의 정확한 경계로 쓴다.
    """
    inner = td.xpath('.//table[contains(@class,"tableGridA")]')
    if inner:
        subs = [" ".join(r.itertext()).strip() for r in inner[0].xpath('.//tr')]
        return subs if subs else [""]
    return [" ".join(td.itertext()).strip()]


def _parse_via_lxml(text: str):
    """중첩테이블 행 경계를 _SENTINEL로 보존해 파싱. 실패 시 None."""
    from lxml import html as _lh
    doc = _lh.fromstring(text)
    # 본 데이터 테이블 = tableGridA(셀 내부용) 아닌 테이블 중 'erp관리코드' 포함·행 최다
    main, best = None, 0
    for tb in doc.xpath("//table"):
        if "tableGridA" in (tb.get("class") or ""):
            continue
        trs = tb.xpath("./tbody/tr | ./tr")
        if "erp관리코드" in "".join(tb.itertext()) and len(trs) > best:
            main, best = tb, len(trs)
    if main is None:
        return None
    trs = main.xpath("./tbody/tr | ./tr")
    parsed = []
    for tr in trs:
        tds = tr.xpath("./td")
        parsed.append([_extract_cell(td) for td in tds]
                      if len(tds) == len(_COLS) else None)
    header_idx = None
    for i, vc in enumerate(parsed):
        if vc and vc[0] and vc[0][0].strip() == "erp관리코드":
            header_idx = i
            break
    if header_idx is None:
        return None
    out = []
    for vc in parsed[header_idx + 1:]:
        if vc is None:
            continue
        rec = {}
        for ci, col in enumerate(_COLS):
            cell = vc[ci]
            if len(cell) > 1:                 # 합포: 중첩 행들 → sentinel 보존
                rec[col] = _SENTINEL.join(cell)
            else:                             # 단일 셀: 빈값은 NaN(fill_management_code 호환)
                v = cell[0] if cell else ""
                rec[col] = v if v.strip() != "" else None
        out.append(rec)
    return pd.DataFrame(out, columns=_COLS)


def _parse_via_readhtml(text: str):
    """구 방식(pd.read_html). 중첩 경계 보존 못 함 — lxml 실패 시 폴백."""
    df = pd.read_html(io.StringIO(text))[0]
    header_idx = None
    for i, row in df.iterrows():
        if str(row.iloc[0]).strip() == "erp관리코드":
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("헤더 행(erp관리코드)을 찾을 수 없습니다")
    data = df.iloc[header_idx + 1:].copy()
    data.columns = _COLS
    return data.reset_index(drop=True)


def parse_sales_report(file_bytes: bytes) -> pd.DataFrame:
    """HTML-xls 매출통계 파싱. 노이즈 행 제거 후 실제 데이터 반환.

    합포(다른 상품) 셀의 중첩테이블 행 경계를 _SENTINEL로 보존한다
    (split_multiproduct_cells 1순위). lxml 파싱 실패 시 구 read_html로 폴백.
    """
    text = file_bytes.decode("utf-8").replace("\ufeff", "")
    try:
        data = _parse_via_lxml(text)
    except Exception:
        data = None
    if data is None or data.empty:
        data = _parse_via_readhtml(text)

    for col in ["총수량", "평균단가", "정산금액", "선결제택배비"]:
        data[col] = pd.to_numeric(
            data[col].astype(str).str.replace(",", "", regex=False),
            errors="coerce",
        )
    return data


def split_multiproduct_cells(df: pd.DataFrame) -> pd.DataFrame:
    """
    Step 0a : 서로 다른 상품 합포 케이스 분리.

    pd.read_html이 HTML xls의 합포 행을 파싱할 때 두 상품 이상의
    어드민옵션·옵션추가항목1이 한 셀에 이어붙여지는 경우가 있음.
    예) 어드민옵션='코카콜라355...[31-03-05]코카콜라제로355...[31-22-02]'
        어드민옵션='명가 꽈배기 참깨[MGG1EA-67-74]명가 꽈배기 흑당[MGG1EA-67-74-01]'

    감지·분리 1순위 = parse가 남긴 _SENTINEL(중첩테이블 행 경계 = 원본 진실).
      - erp관리코드/어드민옵션/옵션추가항목1 어디든 sentinel이 있으면 그 경계로 분할.
      - 코드 = erp행값 → (없으면) 옵션추가항목1행값 → (없으면) 상품명 속 [대괄호코드] 순.
      - 대괄호도 옵션추가항목1도 필요 없어, 카프리썬처럼 코드끼리 붙고 옵션1이 빈 케이스도 잡음.
    2순위 = 어드민옵션의 [erp코드] 대괄호(포맷 무관·N개).
    3순위 = 옵션추가항목1의 숫자코드(XX-XX-XX) 2개(구버전 호환).
    처리 : 총수량·정산금액 ÷ N(상품 수), 상품명을 각 상품 경계로 분할,
           각 행에 해당 erp코드 직접 기입. 선결제비·판매처·평균단가는 복사.
    """
    new_rows = []
    changed = False

    for _, row in df.iterrows():
        erp = str(row.get("erp관리코드") or "")
        admin = str(row.get("어드민옵션") or "")
        opt1 = str(row.get("옵션추가항목1") or "")

        codes = None
        segs = None

        # 1순위: sentinel (중첩테이블 행 경계 = ground truth)
        if _SENTINEL in erp or _SENTINEL in admin or _SENTINEL in opt1:
            eparts = erp.split(_SENTINEL)
            nparts = admin.split(_SENTINEL)
            oparts = opt1.split(_SENTINEL)
            n = max(len(eparts), len(nparts), len(oparts))
            pick = lambda parts, k: (parts[k] if k < len(parts) else "")
            cand_codes, cand_segs = [], []
            for k in range(n):
                nm = pick(nparts, k).strip()
                code = pick(eparts, k).strip() or pick(oparts, k).strip()
                if not code:  # 코드가 상품명 대괄호에만 있는 경우
                    bm = [m for m in _BRACKET_CODE_RE.finditer(nm)
                          if any(c.isdigit() for c in m.group(1))]
                    if bm:
                        code = bm[0].group(1)
                cand_codes.append(code)
                cand_segs.append(nm)
            if sum(1 for c in cand_codes if c) >= 2:
                codes, segs = cand_codes, cand_segs
            else:
                # sentinel이 있으나 실제 합포가 아님(이상 케이스) → sentinel만 정리해 단일화
                r = row.copy()
                r["erp관리코드"] = (erp.replace(_SENTINEL, "") or None)
                r["어드민옵션"] = admin.replace(_SENTINEL, " ").strip()
                r["옵션추가항목1"] = (opt1.replace(_SENTINEL, "") or None)
                new_rows.append(r)
                changed = True
                continue

        # 2순위: 어드민옵션 [erp코드] (숫자를 포함하는 코드만)
        if codes is None:
            bm = [m for m in _BRACKET_CODE_RE.finditer(admin)
                  if any(c.isdigit() for c in m.group(1))]
            if len(bm) >= 2:
                codes = [m.group(1) for m in bm]
                bounds = [0] + [m.end() for m in bm]
                segs = [admin[bounds[k]:bounds[k + 1]].strip() for k in range(len(bm))]

        # 3순위(구버전 호환): 옵션추가항목1 숫자코드 2개
        if codes is None:
            ncodes = _ERP_CODE_RE.findall(opt1)
            if len(ncodes) < 2:
                new_rows.append(row)
                continue
            codes = ncodes[:2]
            sm = re.search(r'\[' + re.escape(codes[0]) + r'\]', admin)
            if sm:
                segs = [admin[:sm.end()].strip(), admin[sm.end():].strip()]
            else:
                segs = [admin, admin]

        changed = True
        n = len(codes)
        qty = float(row.get("총수량") or 0)
        price = float(row.get("정산금액") or 0)

        for code, seg in zip(codes, segs):
            new_row = row.copy()
            new_row["어드민옵션"] = seg
            new_row["총수량"] = qty / n
            new_row["정산금액"] = price / n
            new_row["옵션추가항목1"] = code
            new_row["erp관리코드"] = code
            new_rows.append(new_row)

    if not changed:
        return df

    result = pd.DataFrame(new_rows)
    return result[df.columns].reset_index(drop=True)


def fill_management_code(df: pd.DataFrame) -> pd.DataFrame:
    """Step 0b : erp관리코드 공백 → 옵션추가항목1에서 채움."""
    df = df.copy()
    mask = df["erp관리코드"].isna() & df["옵션추가항목1"].notna()
    df.loc[mask, "erp관리코드"] = df.loc[mask, "옵션추가항목1"].astype(str)
    return df


def split_merged_cells(df: pd.DataFrame) -> pd.DataFrame:
    """Step 1 : 총수량 NaN 행(병합셀 두 번째 행) → 위 행의 총수량·정산금액 ÷ 2."""
    df = df.copy().reset_index(drop=True)
    i = 1
    while i < len(df):
        if pd.isna(df.at[i, "총수량"]) and not pd.isna(df.at[i - 1, "총수량"]):
            for col in ["총수량", "정산금액"]:
                half = (df.at[i - 1, col] or 0) / 2
                df.at[i - 1, col] = half
                df.at[i, col] = half
            for col in ["판매처그룹", "선결제택배비", "평균단가"]:
                df.at[i, col] = df.at[i - 1, col]
        i += 1
    return df


def enrich_classification(df: pd.DataFrame,
                          cls_df: pd.DataFrame) -> tuple[pd.DataFrame, list]:
    """Step 2 : 구분 매핑. 미분류 코드 목록 반환."""
    cls_map = dict(zip(cls_df["관리코드"].astype(str), cls_df["구분"]))
    df = df.copy()
    df["구분"] = df["erp관리코드"].astype(str).map(cls_map)
    unmatched = (
        df[df["구분"].isna() & df["erp관리코드"].notna()]
        [["erp관리코드", "어드민옵션"]]
        .drop_duplicates()
        .to_dict("records")
    )
    return df, unmatched


def deduplicate(df: pd.DataFrame) -> pd.DataFrame:
    """Step 3 : 관리코드 기준 중복 제거 + 총수량·정산금액·선결제비 합산."""
    agg = {
        "어드민옵션": "first",
        "총수량": "sum",
        "평균단가": "first",
        "정산금액": "sum",
        "판매처그룹": "first",
        "선결제택배비": "sum",
        "옵션추가항목1": "first",
        "구분": "first",
    }
    return (
        df.groupby("erp관리코드", sort=False)
        .agg(agg)
        .reset_index()
    )


def enrich_spec(df: pd.DataFrame, spec_df: pd.DataFrame) -> pd.DataFrame:
    """Step 4 : 규격 추가."""
    spec_map = dict(zip(spec_df["관리코드"].astype(str), spec_df["규격"]))
    df = df.copy()
    df["규격"] = df["erp관리코드"].astype(str).map(spec_map).fillna("")
    return df


def run_phase1(sales_bytes: bytes, cls_df=None, spec_df=None):
    """
    Phase 1 전체 실행.
    반환 : (result_df, unmatched, pre_cls_df, archive_df)
      - archive_df : 정제+셀나누기까지만 된 8열 스냅샷 (중복제거 전, 발주자료 아카이브용)
      - unmatched 비어있으면 result_df 완성, pre_cls_df=None
      - unmatched 있으면 result_df=None, pre_cls_df로 GATE A 처리 후 재실행
    """
    if cls_df is None:
        cls_df = load_classification()
    if spec_df is None:
        spec_df = load_spec_master()

    df = parse_sales_report(sales_bytes)
    df = split_multiproduct_cells(df)  # Step 0a: 다른 상품 합포 분리 (fill 전)
    df = fill_management_code(df)      # Step 0b: erp코드 채우기
    df = split_merged_cells(df)        # Step 1: 같은 상품 합포(NaN 방식) 분리
    # 유효 행만 (총수량, erp관리코드 모두 있는 것)
    df = df[df["총수량"].notna() & df["erp관리코드"].notna()].copy()

    # 발주자료 아카이브 스냅샷 (중복제거·구분·규격 전, 원본 8열)
    archive_df = df[_COLS].copy()

    df2, unmatched = enrich_classification(df, cls_df)
    if unmatched:
        return None, unmatched, df, archive_df   # pre_cls_df + archive

    df2 = deduplicate(df2)
    df2 = enrich_spec(df2, spec_df)
    out_cols = ["구분", "규격", "erp관리코드", "어드민옵션", "총수량",
                "평균단가", "정산금액", "판매처그룹", "선결제택배비", "옵션추가항목1"]
    df2 = df2[[c for c in out_cols if c in df2.columns]]
    return df2, [], None, archive_df


def resume_phase1_after_gate(pre_cls_df: pd.DataFrame,
                             cls_df=None, spec_df=None):
    """GATE A 통과 후 Phase 1 재시작 (parse 단계 생략)."""
    if cls_df is None:
        cls_df = load_classification()
    if spec_df is None:
        spec_df = load_spec_master()

    df, unmatched = enrich_classification(pre_cls_df, cls_df)
    if unmatched:
        return None, unmatched, pre_cls_df

    df = deduplicate(df)
    df = enrich_spec(df, spec_df)
    out_cols = ["구분", "규격", "erp관리코드", "어드민옵션", "총수량",
                "평균단가", "정산금액", "판매처그룹", "선결제택배비", "옵션추가항목1"]
    df = df[[c for c in out_cols if c in df.columns]]
    return df, [], None


# ───────────────────────────────────────────────
# Phase 2 — 재고 대조 + 출력
# ───────────────────────────────────────────────

def reconcile_stock(df: pd.DataFrame,
                    pm_df: pd.DataFrame,
                    unit_df: pd.DataFrame):
    """
    재고 대조.
    pm_df : 상품관리 (col4=관리코드, col14=박스재고)
    unit_df : 낱개처리목록
    반환 : (df_with_stock, unmatched_units)
    """
    # 상품관리 맵 (관리코드 → 박스재고)
    stock_map = {}
    if not pm_df.empty:
        codes = pm_df.iloc[:, 4].astype(str).str.strip()
        stocks = pd.to_numeric(pm_df.iloc[:, 14], errors="coerce").fillna(0)
        stock_map = dict(zip(codes, stocks))

    # 낱개 맵 (낱개코드 → {원코드, 입력값})
    unit_map = {}
    for _, r in unit_df.iterrows():
        if pd.notna(r["관리코드"]) and pd.notna(r["원코드"]):
            unit_map[str(r["관리코드"]).strip()] = {
                "원코드": str(r["원코드"]).strip(),
                "입력값": float(r["입력값"]) if pd.notna(r["입력값"]) else 1.0,
            }

    df = df.copy()
    df["_재고박스"] = 0.0
    df["_필요수량"] = pd.to_numeric(df["총수량"], errors="coerce").fillna(0).astype(float)
    df["_낱개"] = False
    df["_원코드미매칭"] = False

    for idx, row in df.iterrows():
        code = str(row["erp관리코드"]).strip()
        qty = float(row["총수량"]) if pd.notna(row["총수량"]) else 0

        if code in unit_map:
            원코드 = unit_map[code]["원코드"]
            배수 = unit_map[code]["입력값"]
            재고 = stock_map.get(원코드, None)
            if 재고 is None:
                df.at[idx, "_원코드미매칭"] = True
                재고 = 0
            df.at[idx, "_재고박스"] = 재고
            df.at[idx, "_필요수량"] = qty * 배수
            df.at[idx, "_낱개"] = True
        else:
            df.at[idx, "_재고박스"] = stock_map.get(code, 0)

    df["재고"] = (df["_재고박스"] - df["_필요수량"]).round(0).astype(int)
    df["총수량표시"] = df.apply(
        lambda r: (f"낱{int(r['총수량'])}" if r["_낱개"]
                   else str(int(r["총수량"])) if pd.notna(r["총수량"])
                   else ""),
        axis=1,
    )

    unmatched_units = (
        df[df["_원코드미매칭"]][["erp관리코드", "어드민옵션"]]
        .drop_duplicates()
        .to_dict("records")
    )
    return df, unmatched_units


def run_phase2(phase1_df: pd.DataFrame, pm_df=None, unit_df=None):
    """
    Phase 2 전체 실행.
    반환 : (logistics_df, unmatched_units, stockout_df)
    """
    if pm_df is None:
        pm_df = load_product_master()
    if unit_df is None:
        unit_df = load_unit_list()

    sections = []
    for 구분 in ["선물세트", "식품", "음료"]:
        sec = phase1_df[phase1_df["구분"] == 구분].copy()
        sec = sec.sort_values("규격", ascending=False)
        sections.append(sec)

    combined = pd.concat(sections, ignore_index=True)
    combined, unmatched_units = reconcile_stock(combined, pm_df, unit_df)

    if unmatched_units:
        return None, unmatched_units, combined

    stockout = (
        combined[combined["재고"] < 0]
        [["erp관리코드", "어드민옵션", "총수량표시", "재고"]]
        .copy()
    )
    stockout.columns = ["관리코드", "상품명", "발주수량", "현재고"]
    stockout = stockout.reset_index(drop=True)
    return combined, [], stockout


# ───────────────────────────────────────────────
# Excel 생성
# ───────────────────────────────────────────────

_THIN = Side(border_style="thin", color="999999")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_RIGHT  = Alignment(horizontal="right",  vertical="center")

# 타이틀 바
_TITLE_FILL = PatternFill("solid", fgColor="2F5496")
_TITLE_FONT = Font(color="FFFFFF", bold=True, size=12)
# 섹션 헤더
_SEC_FILL = PatternFill("solid", fgColor="D9E1F2")
_SEC_FONT = Font(color="1F3864", bold=True, size=11)
# 구분 카테고리 색
_GUBUN_FILL = {
    "선물세트": PatternFill("solid", fgColor="FFF3CD"),
    "식품":     PatternFill("solid", fgColor="E2EFF7"),
    "음료":     PatternFill("solid", fgColor="E3F1E5"),
}
_GUBUN_FONT = {
    "선물세트": Font(bold=True, color="7A5C00"),
    "식품":     Font(bold=True, color="1F4E64"),
    "음료":     Font(bold=True, color="1F5128"),
}
# 품절 / 낱개 / 일반 재고
_OUT_FILL = PatternFill("solid", fgColor="FDE7E7")
_OUT_FONT = Font(bold=True, color="C00000")
_OK_FONT  = Font(bold=True, color="222222")
_NAT_FONT = Font(bold=True, color="0B5394")    # 낱개 총수량
_QTY_FONT = Font(color="222222")

_WEEKDAY = ["월", "화", "수", "목", "금", "토", "일"]


def _korean_date(dt) -> str:
    return f"{dt.year}년 {dt.month}월 {dt.day}일 {_WEEKDAY[dt.weekday()]}요일"


def _write_border(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = _BORDER


def _merge_consecutive(ws, col: int, start_row: int, end_row: int):
    """동일 값이 연속되는 구간을 병합."""
    vals = [ws.cell(row=r, column=col).value for r in range(start_row, end_row + 1)]
    i = 0
    while i < len(vals):
        j = i + 1
        while j < len(vals) and vals[j] == vals[i]:
            j += 1
        if j - i > 1:
            r_s = start_row + i
            r_e = start_row + j - 1
            ws.merge_cells(
                start_row=r_s, start_column=col,
                end_row=r_e,   end_column=col
            )
            ws.cell(row=r_s, column=col).alignment = _CENTER
        i = j


def generate_archive_xlsx(archive_df: pd.DataFrame) -> bytes:
    """발주자료 아카이브 xlsx (복사붙여넣기 시트, 원본 8열·중복제거 전)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "복사붙여넣기"

    # 실제 워크플로우 헤더 (어드민 옵션 = 띄어쓰기 포함)
    display_header = ["erp관리코드", "어드민 옵션", "총수량", "평균단가",
                      "정산금액", "판매처그룹", "선결제택배비", "옵션추가항목1"]
    ws.append(display_header)
    for _, row in archive_df.iterrows():
        ws.append([row.get(c, "") for c in _COLS])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_result_xlsx(logistics_df: pd.DataFrame,
                         stockout_df: pd.DataFrame, cadence: dict = None,
                         unit_df: pd.DataFrame = None) -> bytes:
    """최종결과물 xlsx : 물류팀 + 품절목록."""
    wb = Workbook()

    # ── 물류팀 시트 ──────────────────────────────
    ws = wb.active
    ws.title = "물류팀"

    today = _korean_date(datetime.now(_KST))

    # Row 1 : 타이틀
    ws.append(["", "멸치+오픈마켓", "", today, "", "재고"])
    for c in range(1, 7):
        cell = ws.cell(row=1, column=c)
        cell.fill = _TITLE_FILL
        cell.font = _TITLE_FONT
        cell.alignment = _CENTER

    current_row = 2
    for 구분 in ["선물세트", "식품", "음료"]:
        section = logistics_df[logistics_df["구분"] == 구분]
        if section.empty:
            continue

        # 섹션 헤더
        ws.append(["구분", "규격", "erp관리코드", "어드민 옵션", "총수량", "재고"])
        for c in range(1, 7):
            cell = ws.cell(row=current_row, column=c)
            cell.fill = _SEC_FILL
            cell.font = _SEC_FONT
            cell.alignment = _CENTER
        current_row += 1
        section_data_start = current_row

        for _, r in section.iterrows():
            재고_val = int(r["재고"])
            qty_disp = r.get("총수량표시", "")
            is_nat = str(qty_disp).startswith("낱")
            ws.append([
                r.get("구분", ""),
                r.get("규격", ""),
                r.get("erp관리코드", ""),
                r.get("어드민옵션", ""),
                qty_disp,
                재고_val,
            ])
            # 구분(A) 색
            ws.cell(row=current_row, column=1).fill = _GUBUN_FILL.get(구분, PatternFill())
            ws.cell(row=current_row, column=1).font = _GUBUN_FONT.get(구분, Font(bold=True))
            # 규격(B) 가운데
            ws.cell(row=current_row, column=2).alignment = _CENTER
            # 어드민옵션(D) 좌측
            ws.cell(row=current_row, column=4).alignment = _LEFT
            # 총수량(E) 가운데, 낱개=파랑
            ce = ws.cell(row=current_row, column=5)
            ce.alignment = _CENTER
            ce.font = _NAT_FONT if is_nat else _QTY_FONT
            # 재고(F) 우측, 품절=빨강+분홍
            cf = ws.cell(row=current_row, column=6)
            cf.alignment = _RIGHT
            if 재고_val < 0:
                cf.fill = _OUT_FILL
                cf.font = _OUT_FONT
            else:
                cf.font = _OK_FONT
            current_row += 1

        section_data_end = current_row - 1

        if section_data_end >= section_data_start:
            # A열 : 섹션 전체 병합 (구분 동일)
            if section_data_end > section_data_start:
                ws.merge_cells(
                    start_row=section_data_start, start_column=1,
                    end_row=section_data_end,     end_column=1
                )
            ws.cell(row=section_data_start, column=1).alignment = _CENTER
            # B열 : 연속 동일 규격 병합
            _merge_consecutive(ws, col=2,
                               start_row=section_data_start,
                               end_row=section_data_end)

    # 전체 테두리 + C열 숨김
    last_row = ws.max_row
    _write_border(ws, 1, last_row, 1, 6)
    ws.column_dimensions["C"].hidden = True

    # 컬럼 너비
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 17
    ws.column_dimensions["D"].width = 34
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 8
    ws.row_dimensions[1].height = 22

    # A·B·F 글자크기 9 · D(어드민옵션) 자동줄바꿈 해제 (물류팀 인쇄 가독성)
    # 물류팀 ws 한정 후처리 — C/D/E 폰트·품절목록 시트는 그대로.
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
        for ci in (1, 2, 6):          # A 구분 · B 규격 · F 재고
            cell = row[ci - 1]
            f = cell.font
            cell.font = Font(name=f.name, size=9, bold=f.bold, italic=f.italic,
                             color=f.color, underline=f.underline, strike=f.strike)
        dcell = row[3]                # D 어드민옵션 — wrap만 끄고 정렬 보존
        a = dcell.alignment
        dcell.alignment = Alignment(horizontal=a.horizontal, vertical=a.vertical,
                                    wrap_text=False, text_rotation=a.text_rotation,
                                    indent=a.indent)

    # ── 품절목록 시트 ─────────────────────────────
    ws2 = wb.create_sheet("품절목록")
    cadence = cadence or {}
    # 낱개/소분 코드는 매입현황에 없음 → 원코드(박스)로 치환해 cadence 조회 (재고와 동일 기준)
    try:
        _omap = unit_origin_map(unit_df) if cadence else {}
    except Exception:
        _omap = {}
    ws2.append(["관리코드", "상품명", "발주수량", "현재고", "최근 입고일", "평균매입주기(일)", "입고횟수(1년)"])
    for c in range(1, 8):
        cell = ws2.cell(row=1, column=c)
        cell.fill = _SEC_FILL
        cell.font = _SEC_FONT
        cell.alignment = _CENTER

    r_idx = 2
    for _, r in stockout_df.iterrows():
        _code = unicodedata.normalize("NFC", str(r["관리코드"]).strip())
        _info = cadence.get(_code) or cadence.get(_omap.get(_code, "")) or {}
        _last = _info.get("최근입고일")
        _last_s = _last.strftime("%Y-%m-%d") if (_last is not None and pd.notna(_last)) else ""
        _avg = _info.get("평균주기")
        _avg_v = round(_avg) if _avg is not None else ""
        _cnt = _info.get("입고횟수")
        _cnt_v = int(_cnt) if _cnt else ""
        ws2.append([r["관리코드"], r["상품명"], r["발주수량"], int(r["현재고"]), _last_s, _avg_v, _cnt_v])
        cf = ws2.cell(row=r_idx, column=4)
        cf.font = _OUT_FONT
        cf.fill = _OUT_FILL
        cf.alignment = _RIGHT
        ws2.cell(row=r_idx, column=3).alignment = _CENTER
        ws2.cell(row=r_idx, column=5).alignment = _CENTER
        ws2.cell(row=r_idx, column=6).alignment = _CENTER
        ws2.cell(row=r_idx, column=7).alignment = _CENTER
        r_idx += 1

    _write_border(ws2, 1, ws2.max_row, 1, 7)
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 38
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 10
    ws2.column_dimensions["E"].width = 13
    ws2.column_dimensions["F"].width = 14
    ws2.column_dimensions["G"].width = 11

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
