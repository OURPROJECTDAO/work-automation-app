"""
송장번호 일괄입력 (invoice-fill) 워크플로우 — 채널별 송장 채우기.

원리 (전 채널 공통, 템플릿 양식만 다름):
  입력 = 채널 '송장번호 일괄입력 템플릿' (처리전; 송장번호 빈칸)
         + 공통 송장 마스터(송장출력 .xlsx)
  처리 = VLOOKUP(처리전.<match_col> ⟷ 마스터.<master_key>) → 택배사·송장번호 (첫 매칭)
  N/A  = 매칭 실패 행 →
           ① 동일 배송지에 송장 채워진 행(박스) 존재 → 합포장 후보로 사용자 확인
                (같은 주소 박스 여럿이면 사용자가 어느 박스인지 선택; 확정 시 그 송장 복사)
           ② 없음(독립개체) 또는 사용자가 합포장 부정 → N/A 유지
  전달 = 잔존 N/A 행 전체 삭제 후 원본 양식 그대로 출력 + N/A 건수 보고

채널 추가 시: CHANNEL_CONFIG 에 한 줄 등록(아래 스키마 참조). 양식이 같으면 그걸로 끝.
출력은 반드시 원본 양식(.xls/.xlsx) — 모든 판매처가 원본 양식만 허용.

[채널 설정 스키마]
  format        : "xls"(OLE2 BIFF, xlrd/xlwt) | "xlsx"(openpyxl)
  match_col     : 처리전(채널 템플릿)에서 VLOOKUP 키로 쓰는 컬럼명
  master_key    : 송장 마스터에서 매칭되는 컬럼명
  courier       : 택배사 열에 일괄 기입할 택배사 (없으면 lookup 택배사 fallback)
  courier_col   : 택배사 컬럼명
  invoice_col   : 송장번호(운송장번호) 컬럼명 — 숫자 형식으로 기입
  addr_col      : 합포장 동일주소 판정 컬럼명
  recv_col      : 수령인 표시 컬럼명
  has_guide_row : 헤더 다음 안내문 행 유무 (식봄 .xls = True, 데이터 r2~ / 없으면 데이터 r1~)
"""
import io
import unicodedata
from collections import OrderedDict

import openpyxl


def nfc(s) -> str:
    if s is None:
        return ""
    return unicodedata.normalize("NFC", str(s)).strip()


def to_invoice_number(v):
    """송장번호를 숫자(int)로. 전부 숫자면 int, 아니면 원본 문자열 유지.
    openpyxl 등이 float로 읽어 '....0' 꼬리표가 붙은 경우도 정수화."""
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return int(s) if s.isdigit() else (s if s else "")


def decrypt_if_needed(file_bytes: bytes, cfg: dict) -> bytes:
    """암호 걸린 xlsx면 cfg['password']로 복호화. 평문이면 그대로(미리 푼 파일도 허용).
    배민상회 등 일부 채널은 다운로드 파일에 항상 열기 암호가 걸려 있다(msoffcrypto-tool)."""
    if cfg.get("format") != "xlsx":
        return file_bytes
    import msoffcrypto
    try:
        off = msoffcrypto.OfficeFile(io.BytesIO(file_bytes))
        if not off.is_encrypted():
            return file_bytes
    except Exception:
        return file_bytes  # 평문 xlsx(ZIP)는 OfficeFile 생성 실패 가능 → 그대로 사용
    buf = io.BytesIO()
    off.load_key(password=cfg.get("password") or "")
    off.decrypt(buf)
    return buf.getvalue()


# ── 채널 설정 ─────────────────────────────────────────────
CHANNEL_CONFIG = {
    "식봄": {
        "format": "xls",
        "match_col": "상품주문번호",
        "master_key": "주문번호",
        "courier": "한진택배",
        "courier_col": "택배사",
        "invoice_col": "송장번호",
        "addr_col": "배송지",
        "recv_col": "수취인명(받는사람)",
        "has_guide_row": True,
    },
    "올웨이즈": {
        "format": "xlsx",
        "match_col": "주문아이디",
        "master_key": "주문번호",
        "courier": "한진택배",
        "courier_col": "택배사",
        "invoice_col": "운송장번호",
        "addr_col": "주소",
        "recv_col": "수령인",
        "has_guide_row": False,
    },
    "배민상회": {
        "format": "xlsx",
        "password": "qwer",          # 배민 다운로드 파일은 항상 암호 걸림
        "match_col": "주문번호",
        "master_key": "주문번호",
        "courier": "한진택배",
        "courier_col": "*택배사",      # 헤더에 * 접두사
        "invoice_col": "*송장번호",     # 헤더에 * 접두사 (트래킹번호 등 다른 * 필드는 미수정)
        "addr_col": "도로명 주소",
        "recv_col": "받는분",
        "has_guide_row": False,
    },
    # "캐시노트": {...},  # 샘플 받으면 추가
}

# 송장 마스터(송장출력) 표준 컬럼
MASTER_COLS = ["상태", "관리번호", "발주일", "판매처", "주문번호",
               "수령자", "주소", "상품명", "택배사", "송장번호"]


# ── IO: 처리전 템플릿 파싱 ────────────────────────────────
def parse_template(file_bytes: bytes, cfg: dict) -> dict:
    """채널 포맷에 따라 처리전 템플릿 파싱. format='xls'|'xlsx'."""
    if cfg.get("format", "xls") == "xlsx":
        return _parse_template_xlsx(file_bytes)
    return _parse_template_xls(file_bytes)


def _parse_template_xls(file_bytes: bytes) -> dict:
    """채널 송장 템플릿(.xls OLE2) 파싱. r0=헤더, r1=안내문, r2+=데이터.
    셀 타입을 보존해 출력 시 원본 양식 재현.
    """
    import xlrd
    book = xlrd.open_workbook(file_contents=file_bytes)
    sh = book.sheet_by_index(0)
    header = [sh.cell_value(0, c) for c in range(sh.ncols)]
    guide = [sh.cell_value(1, c) for c in range(sh.ncols)] if sh.nrows > 1 else [""] * sh.ncols
    rows, types = [], []
    for r in range(2, sh.nrows):
        rows.append({header[c]: sh.cell_value(r, c) for c in range(sh.ncols)})
        types.append([sh.cell_type(r, c) for c in range(sh.ncols)])
    return {"format": "xls", "sheet_name": sh.name, "header": header,
            "guide": guide, "rows": rows, "types": types}


def _parse_template_xlsx(file_bytes: bytes) -> dict:
    """채널 송장 템플릿(.xlsx) 파싱. r0=헤더, r1+=데이터(안내문 행 없음).
    출력은 원본 .xlsx를 in-place 편집하므로 인덱스 정렬 유지를 위해 빈 행도 건너뛰지 않는다.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active
    rit = ws.iter_rows(values_only=True)
    header = list(next(rit))
    rows = []
    for r in rit:
        rows.append({header[c]: (r[c] if c < len(r) else None) for c in range(len(header))})
    sheet_name = ws.title
    wb.close()
    return {"format": "xlsx", "sheet_name": sheet_name, "header": header,
            "guide": [], "rows": rows, "types": None}


def parse_master(file_bytes: bytes) -> list:
    """공통 송장 마스터(.xlsx, 시트 '송장출력' 또는 첫 시트) → list[dict]."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb["송장출력"] if "송장출력" in wb.sheetnames else wb[wb.sheetnames[0]]
    rit = ws.iter_rows(values_only=True)
    hdr = list(next(rit))
    out = [{hdr[i]: row[i] for i in range(len(hdr))} for row in rit if any(v is not None for v in row)]
    wb.close()
    return out


# ── IO: 처리후 출력 ───────────────────────────────────────
def write_template(orig_bytes: bytes, parsed: dict, rows: list,
                   keep_idx: list, cfg: dict) -> bytes:
    """살아남은 행(keep_idx)만, 원본 양식(.xls/.xlsx) 그대로 출력.
    송장번호 = 숫자 형식, 택배사 = 채널 courier 일괄.
    """
    if cfg.get("format", "xls") == "xlsx":
        return _write_template_xlsx(orig_bytes, parsed, rows, keep_idx, cfg)
    return _write_template_xls(parsed, rows, keep_idx, cfg)


def _write_template_xls(parsed: dict, rows: list, keep_idx: list, cfg: dict) -> bytes:
    """.xls(OLE2) 양식 재작성(xlwt). 시트명·헤더·안내문·셀타입 보존."""
    import xlrd
    import xlwt
    header = parsed["header"]
    song_col = header.index(cfg["invoice_col"])
    courier_col_name = cfg.get("courier_col", "택배사")
    courier = cfg.get("courier")
    wbk = xlwt.Workbook(encoding="utf-8")
    sh = wbk.add_sheet(parsed["sheet_name"])
    for c, v in enumerate(header):
        sh.write(0, c, v)
    for c, v in enumerate(parsed["guide"]):
        sh.write(1, c, v)
    out_r = 2
    for i in keep_idx:
        row = rows[i]
        types = parsed["types"][i]
        for c in range(len(header)):
            if c == song_col:
                sh.write(out_r, c, to_invoice_number(row["_송장"]))   # 송장번호 = 숫자 형식
            elif header[c] == courier_col_name:
                sh.write(out_r, c, courier or row.get("_택배사") or "")  # 택배사 일괄
            else:
                val = row.get(header[c])
                if types[c] == xlrd.XL_CELL_NUMBER:
                    sh.write(out_r, c, val)                 # 숫자 보존
                else:
                    sh.write(out_r, c, "" if val is None else val)
        out_r += 1
    buf = io.BytesIO()
    wbk.save(buf)
    return buf.getvalue()


def _write_template_xlsx(orig_bytes: bytes, parsed: dict, rows: list,
                         keep_idx: list, cfg: dict) -> bytes:
    """.xlsx 양식을 원본 그대로 in-place 편집(openpyxl). 서식 완전 보존.
    데이터행(parsed['rows'][i]) = 엑셀 행 (base + i). 안내문 행 없으면 base=2(헤더=1).
    """
    wb = openpyxl.load_workbook(io.BytesIO(orig_bytes))
    ws = wb[parsed["sheet_name"]] if parsed["sheet_name"] in wb.sheetnames else wb.active
    header = parsed["header"]
    inv_c = header.index(cfg["invoice_col"]) + 1          # 1-based
    cour_c = header.index(cfg.get("courier_col", "택배사")) + 1
    courier = cfg.get("courier")
    base = 3 if cfg.get("has_guide_row") else 2           # 데이터 시작 엑셀 행(1-based)

    # 1) 모든 데이터 행에 송장번호·택배사 기입
    for i, row in enumerate(rows):
        er = base + i
        song = row.get("_송장")
        ws.cell(er, inv_c).value = to_invoice_number(song) if song else None
        ws.cell(er, cour_c).value = courier or row.get("_택배사") or ws.cell(er, cour_c).value

    # 2) 잔존 N/A 행 삭제(아래에서 위로 — 인덱스 밀림 방지)
    keep = set(keep_idx)
    for i in range(len(rows) - 1, -1, -1):
        if i not in keep:
            ws.delete_rows(base + i, 1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 로직 ─────────────────────────────────────────────────
def build_master_lookup(master_rows, master_key="주문번호"):
    """마스터 key → (택배사, 송장번호). VLOOKUP 의미상 '첫 매칭'만 유지(분할배송 첫 박스)."""
    lk = OrderedDict()
    for m in master_rows:
        k = nfc(m.get(master_key))
        if not k:
            continue
        if k not in lk:
            lk[k] = (str(m.get("택배사") or "").strip(),
                     str(m.get("송장번호") or "").strip())
    return lk


def vlookup_fill(before_rows, master_lookup, channel):
    """STEP1: VLOOKUP. 각 처리전 행에 _송장/_택배사/_status(matched|na) 부여."""
    mc = CHANNEL_CONFIG[channel]["match_col"]
    out = []
    for row in before_rows:
        r = dict(row)
        hit = master_lookup.get(nfc(r.get(mc)))
        if hit:
            r["_택배사"], r["_송장"], r["_status"] = hit[0], hit[1], "matched"
        else:
            r["_택배사"], r["_송장"], r["_status"] = "", None, "na"
        out.append(r)
    return out


def _recv(r, cfg):
    return str(r.get(cfg["recv_col"]) or r.get("수취인") or "").strip()


def find_consolidation_candidates(rows, cfg):
    """STEP2: N/A 행별 합포장 후보. 동일 배송지(NFC+trim)에 송장 채워진 박스가 있으면 후보.
    반환: (candidates, independents)
      candidates: [{na_index, na_row, boxes:[{송장,택배사,수취인,상품명}...]}]
      independents: [na_index ...]  (동일주소 박스 없음 → N/A 유지)
    """
    addr_col = cfg["addr_col"]
    addr_boxes = {}
    for r in rows:
        if r["_status"] == "matched" and r["_송장"]:
            addr = nfc(r.get(addr_col))
            addr_boxes.setdefault(addr, OrderedDict())
            if r["_송장"] not in addr_boxes[addr]:
                addr_boxes[addr][r["_송장"]] = {
                    "송장": r["_송장"], "택배사": r["_택배사"],
                    "수취인": _recv(r, cfg), "상품명": str(r.get("상품명") or "").strip(),
                }
    candidates, independents = [], []
    for i, r in enumerate(rows):
        if r["_status"] != "na":
            continue
        boxes = list(addr_boxes.get(nfc(r.get(addr_col)), {}).values())
        (candidates.append({"na_index": i, "na_row": r, "boxes": boxes})
         if boxes else independents.append(i))
    return candidates, independents


def apply_decisions(rows, decisions):
    """STEP3: 합포장 결정 반영. decisions: {na_index: 송장 or None}.
    송장 → 그 박스 송장/택배사 복사(status=consolidated), None → N/A 유지.
    """
    song_to_courier = {}
    for r in rows:
        if r["_송장"]:
            song_to_courier.setdefault(r["_송장"], r["_택배사"])
    for i, song in decisions.items():
        if song:
            rows[i]["_송장"] = song
            rows[i]["_택배사"] = song_to_courier.get(song, rows[i]["_택배사"])
            rows[i]["_status"] = "consolidated"
    return rows


def finalize(rows):
    """STEP4: 잔존 N/A 행 제외. 반환 (keep_idx, na_count, na_rows)."""
    keep_idx, na = [], []
    for i, r in enumerate(rows):
        if r["_status"] == "na" or not r["_송장"]:
            na.append(r)
        else:
            keep_idx.append(i)
    return keep_idx, len(na), na
