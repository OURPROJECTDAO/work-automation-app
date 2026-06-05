"""
천년경영업로드자동화 (cheonnyeon-upload)
─────────────────────────────────────────────────────────────
원본 `천년경영업로드자동화V15.xlsm`(VBA OC_CheonnyeonFullRun 37단계 + 저장)의 Python 재구현.

입력 (매 실행 3개 업로드):
  - ★★발주자료{mmdd}.xlsx  : logistics 발주자료 아카이브. 앞 7열만 사용
                              [erp관리코드, 어드민옵션, 총수량, 평균단가, 정산금액, 판매처그룹, 선결제택배비]
  - 배민주문{yyyymmdd}.xlsx : 배민 배송비. Z(26)=관리코드, AL(38)=배송비
  - 스스주문{yyyymmdd}.xlsx : 스마트스토어 배송비 (암호 1323). r2=헤더, AJ(36)=묶음번호, AL(38)=배송비합계, AO(41)=상품코드

기준데이터 (reference/, 고정):
  - logistics_classification.csv  : 멸치쇼핑 분류표 (관리코드→구분). **발주서출력업무와 공유**
  - bm_commission.csv             : 배민상회 수수료율 (관리코드→수수료율)
  - sub_list.csv                  : 소분목록 (낱개코드→낱개개수·원코드)

출력: {yymmdd}.xlsx — 마켓플레이스 14 전체 + 13 낱개 시트.
"""
from __future__ import annotations
import datetime

_KST = datetime.timezone(datetime.timedelta(hours=9))  # 한국 표준시 UTC+9
import io
from collections import Counter
from pathlib import Path

import msoffcrypto
import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

_REF = Path(__file__).resolve().parent.parent.parent / "reference"
_SSS_PASSWORD = "1323"

# ── 분배/수식 스펙 ────────────────────────────────────────────
GROUP_TO_SHEET = {
    "ESM": "ESM전체", "스마트스토어": "스마트스토어전체", "11번가": "11번가전체",
    "배민상회": "배민상회전체", "식봄": "식봄전체", "올웨이즈": "올웨이즈전체",
    "배민대용량": "대용량전체", "캐시노트": "캐시노트전체", "쿠팡": "쿠팡전체",
    "알리익스프레스": "알리전체", "셀러허브": "셀러허브전체", "제이티유통": "제이티전체",
}
ALL_SHEETS = ["멸치식품", "멸치음료", "ESM전체", "스마트스토어전체", "11번가전체",
              "식봄전체", "배민상회전체", "올웨이즈전체", "대용량전체", "캐시노트전체",
              "쿠팡전체", "알리전체", "셀러허브전체", "제이티전체"]
NO_G = {"멸치식품", "멸치음료", "ESM전체", "11번가전체", "식봄전체"}
FULL_TO_UNIT = {
    "멸치식품": "멸치낱개", "멸치음료": "멸치낱개", "스마트스토어전체": "스마트스토어낱개",
    "ESM전체": "ESM낱개", "11번가전체": "11번가낱개", "식봄전체": "식봄낱개",
    "배민상회전체": "배민상회낱개", "올웨이즈전체": "올웨이즈낱개", "대용량전체": "대용량낱개",
    "캐시노트전체": "캐시노트낱개", "쿠팡전체": "쿠팡낱개", "알리전체": "알리낱개",
    "셀러허브전체": "셀러허브낱개", "제이티전체": "제이티낱개",
}
UNIT_SHEETS = ["멸치낱개", "스마트스토어낱개", "ESM낱개", "11번가낱개", "식봄낱개",
               "배민상회낱개", "올웨이즈낱개", "대용량낱개", "캐시노트낱개", "쿠팡낱개",
               "알리낱개", "셀러허브낱개", "제이티낱개"]
_FULL_HEADER = ["일자", "관리코드", "상품명", "주문수량", "평균단가",
                "정산금액", "선결제택배비", "실제기입단가"]
_UNIT_HEADER = ["일자", "관리코드", "상품명", "코드단위주문수량", "평균단가", "정산금액",
                "선결제택배비", "실제기입단가", "개당수량", "★기입수량", "★낱개단가"]


# ── 헬퍼 ──────────────────────────────────────────────────────
def _num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _code(v):
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


# ── 기준데이터 로더 ───────────────────────────────────────────
def load_classification(cls_df: pd.DataFrame | None = None) -> dict:
    if cls_df is None:
        cls_df = pd.read_csv(_REF / "logistics_classification.csv",
                             encoding="utf-8-sig", dtype=str)
    return {_code(c): str(g).strip()
            for c, g in zip(cls_df["관리코드"], cls_df["구분"]) if _code(c)}


def load_commission(comm_df: pd.DataFrame | None = None) -> dict:
    if comm_df is None:
        comm_df = pd.read_csv(_REF / "bm_commission.csv",
                              encoding="utf-8-sig", dtype=str)
    m = {}
    for c, rate in zip(comm_df["관리코드"], comm_df["수수료율"]):
        c = _code(c)
        if c and c not in m:                # 첫 출현 우선
            m[c] = _num(rate)
    return m


def load_sub_list(sub_df: pd.DataFrame | None = None) -> dict:
    if sub_df is None:
        sub_df = pd.read_csv(_REF / "sub_list.csv",
                             encoding="utf-8-sig", dtype=str)
    m = {}
    for c, cnt, orig in zip(sub_df["관리코드"], sub_df["낱개개수"], sub_df["원코드"]):
        c = _code(c)
        if c and c not in m:                # 첫 출현 우선
            m[c] = (_num(cnt), _code(orig))
    return m


# ── 입력 파싱 ─────────────────────────────────────────────────
def parse_baeju(file_bytes: bytes):
    """발주자료 앞 7열 → row 튜플 리스트."""
    ws = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True).active
    rows = []
    for r in ws.iter_rows(min_row=2, max_col=7, values_only=True):
        if r[0] in (None, ""):
            continue
        rows.append(r)
    return rows


def open_baemin(file_bytes: bytes):
    return openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True).active


def open_sss(file_bytes: bytes, password: str = _SSS_PASSWORD):
    """스스주문 워크시트 반환. 암호(1323) 걸려 있으면 복호화, 평문이면 그대로 연다."""
    off = msoffcrypto.OfficeFile(io.BytesIO(file_bytes))
    try:
        encrypted = off.is_encrypted()
    except Exception:
        encrypted = True
    if encrypted:
        off.load_key(password=password)
        buf = io.BytesIO()
        off.decrypt(buf)
        return openpyxl.load_workbook(buf, data_only=True).active
    return openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True).active


# ── 배송비조사 정제 ──────────────────────────────────────────
def process_baemin(ws) -> dict:
    """Z(코드) dedup, AL 합산 (Module7_1)."""
    agg = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        code = _code(r[25]) if len(r) > 25 else ""
        if not code:
            continue
        al = _num(r[37]) if len(r) > 37 else None
        agg[code] = (agg.get(code, 0.0) or 0.0) + (al or 0.0)
    return agg


def process_smartstore(ws) -> dict:
    """묶음번호(AJ) 개수로 AL 분할(6_8) → 상품코드(AO) dedup 합산(7)."""
    rows = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        aj = r[35] if len(r) > 35 else None
        al = _num(r[37]) if len(r) > 37 else None
        ao = _code(r[40]) if len(r) > 40 else ""
        if not ao and aj is None and al is None:
            continue
        rows.append({"aj": aj, "al": al, "ao": ao})
    cnt = Counter(x["aj"] for x in rows if x["aj"] is not None)
    for x in rows:
        if x["aj"] is not None and cnt[x["aj"]] > 0 and x["al"] is not None:
            x["al"] = x["al"] / cnt[x["aj"]]
    agg = {}
    for x in rows:
        if not x["ao"]:
            continue
        agg[x["ao"]] = (agg.get(x["ao"], 0.0) or 0.0) + (x["al"] or 0.0)
    return agg


# ── 메인 파이프라인 ──────────────────────────────────────────
def process(baeju_rows, baemin_ws, sss_ws, cls_map, comm_map, sub_map,
            run_date: datetime.date):
    sheets = {s: [] for s in ALL_SHEETS}

    for r in baeju_rows:                                   # 분배 + 구분 룩업
        code = _code(r[0])
        if not code:
            continue
        group = str(r[5]).strip() if r[5] is not None else ""
        gubun = cls_map.get(code, "")
        if group == "멸치":
            target = ("멸치식품" if gubun in ("식품", "선물세트")
                      else "멸치음료" if gubun == "음료" else None)
        else:
            target = GROUP_TO_SHEET.get(group)
        if target is None:
            continue
        sheets[target].append({
            "B": code, "C": r[1] if r[1] is not None else "",
            "D": _num(r[2]), "E": _num(r[3]), "F": _num(r[4]),
            "G": None if target in NO_G else _num(r[6]),
        })

    bm_ship = process_baemin(baemin_ws)
    ss_ship = process_smartstore(sss_ws)

    for name, rows in sheets.items():                      # 상품명(C) dedup, D/F/G 합산
        merged, order = {}, []
        for row in rows:
            key = row["C"]
            if key in merged:
                m = merged[key]
                m["D"] = (m["D"] or 0) + (row["D"] or 0)
                m["F"] = (m["F"] or 0) + (row["F"] or 0)
                m["G"] = (m["G"] or 0) + (row["G"] or 0)
            else:
                merged[key] = dict(row); order.append(key)
        sheets[name] = [merged[k] for k in order]

    for row in sheets["스마트스토어전체"]:                  # 배송비 조인
        if row["B"] in ss_ship:
            row["G"] = ss_ship[row["B"]] * 0.964
    for row in sheets["배민상회전체"]:
        if row["B"] in bm_ship:
            row["G"] = bm_ship[row["B"]]
        row["I_rate"] = comm_map.get(row["B"])

    def Hf(factor):
        return lambda r: (r["F"] * factor) / r["D"] if r["D"] else None
    simple = {"멸치식품": 1.0, "멸치음료": 1.0, "ESM전체": 1.0, "11번가전체": 1.0,
              "올웨이즈전체": 1.0, "셀러허브전체": 1.0, "제이티전체": 1.0,
              "식봄전체": 0.93, "캐시노트전체": 0.94, "쿠팡전체": 0.88, "알리전체": 0.91}
    for name, f in simple.items():
        for r in sheets[name]:
            r["H"] = Hf(f)(r)
    for r in sheets["스마트스토어전체"]:
        r["H"] = ((r["F"] + (r["G"] or 0)) / r["D"]) if r["D"] else None
    for r in sheets["대용량전체"]:
        r["H"] = ((r["F"] * 0.93 + (r["G"] or 0) * 0.967) / r["D"]) if r["D"] else None
    for r in sheets["배민상회전체"]:
        i = r.get("I_rate") or 0
        r["H"] = ((r["F"] * (1 - i - 0.03) + (r["G"] or 0) * 0.967) / r["D"]) if r["D"] else None

    run_dt = datetime.datetime(run_date.year, run_date.month, run_date.day)
    for rows in sheets.values():
        for r in rows:
            r["A"] = run_dt

    units = {s: [] for s in UNIT_SHEETS}                    # 낱개 이동 + 변환
    for full, rows in sheets.items():
        unit, keep = FULL_TO_UNIT[full], []
        for r in rows:
            (units[unit] if r["B"] in sub_map else keep).append(
                dict(r) if r["B"] in sub_map else r)
        sheets[full] = keep
    for unit, rows in units.items():
        for r in rows:
            cnt, orig = sub_map[r["B"]]
            r["B"] = orig
            r["I"] = cnt
            j = (r["D"] or 0) * (cnt or 0)
            r["J"] = j
            r["K"] = (r["H"] * r["D"] / j) if j else None

    for name in sheets:                                     # 빈 상품명 제거
        sheets[name] = [r for r in sheets[name] if r["C"] not in (None, "")]
    return sheets, units


# ── 출력 xlsx ─────────────────────────────────────────────────
_HDR_FILL = PatternFill("solid", fgColor="2F5496")
_HDR_FONT = Font(color="FFFFFF", bold=True)


def generate_output_xlsx(sheets: dict, units: dict, run_date: datetime.date) -> bytes:
    wb = Workbook(); wb.remove(wb.active)
    full_keys = ["A", "B", "C", "D", "E", "F", "G", "H"]
    unit_keys = full_keys + ["I", "J", "K"]

    def write(ws, header, keys, rows):
        ws.append(header)
        for c in range(1, len(header) + 1):
            cell = ws.cell(1, c); cell.fill = _HDR_FILL; cell.font = _HDR_FONT
            cell.alignment = Alignment(horizontal="center")
        for r in rows:
            ws.append([r.get(k) for k in keys])
        for c in range(1, len(header) + 1):
            ws.column_dimensions[chr(64 + c)].width = 14
        ws.freeze_panes = "A2"

    for name in ALL_SHEETS:
        write(wb.create_sheet(name), _FULL_HEADER, full_keys, sheets[name])
    for name in UNIT_SHEETS:
        write(wb.create_sheet(name), _UNIT_HEADER, unit_keys, units[name])

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def run(baeju_bytes: bytes, baemin_bytes: bytes, sss_bytes: bytes,
        run_date: datetime.date | None = None, sss_password: str = _SSS_PASSWORD):
    """엔드투엔드: 3파일 bytes → (출력 xlsx bytes, 통계 dict)."""
    if run_date is None:
        run_date = datetime.datetime.now(_KST).date()
    sheets, units = process(
        parse_baeju(baeju_bytes), open_baemin(baemin_bytes),
        open_sss(sss_bytes, sss_password),
        load_classification(), load_commission(), load_sub_list(), run_date)
    out = generate_output_xlsx(sheets, units, run_date)
    stats = {name: len(sheets[name]) for name in ALL_SHEETS if sheets[name]}
    stats.update({name: len(units[name]) for name in UNIT_SHEETS if units[name]})
    return out, stats, sheets, units
