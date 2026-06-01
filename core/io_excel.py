"""
Excel/CSV/HTML 읽기·쓰기 헬퍼.
한글 인코딩 규칙:
  - xlsx/xlsm: openpyxl → 내부 UTF-8이라 문제없음
  - csv: UTF-8-sig(BOM) → Excel에서 직접 열어도 안 깨짐
  - HTML-format .xls: 마켓플레이스(스마트스토어/쿠팡 등)가 내보내는 형식.
    UTF-8 + <feff> BOM 문자열 제거 후 pd.read_html()
"""
from pathlib import Path
import io as _io
import pandas as pd
import openpyxl


def load_sheets(path: Path) -> dict[str, pd.DataFrame]:
    """xlsm/xlsx 의 모든 시트를 DataFrame dict 로 로드. 한글 시트명 OK."""
    wb = openpyxl.load_workbook(path, read_only=True, keep_vba=True, data_only=True)
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            result[name] = pd.DataFrame()
        else:
            result[name] = pd.DataFrame(rows[1:], columns=rows[0])
    wb.close()
    return result


def load_html_xls(path: Path) -> pd.DataFrame:
    """마켓플레이스 HTML-format .xls 로드.
    마켓플레이스(스마트스토어/쿠팡/G마켓 등)가 내보내는 발주 파일은
    실제로는 HTML 테이블인데 .xls 확장자를 씁니다.
    파일 앞에 <feff> BOM 문자열이 붙어 있어 제거 후 파싱합니다.
    컬럼: 상태/관리번호/발주일/판매처/주문번호/수령자/주소/상품명/택배사/송장번호
    """
    raw = path.read_bytes()
    # UTF-8로 디코딩 후 BOM 문자열/BOM 유니코드 제거
    html = raw.decode('utf-8').replace('\ufeff', '').replace('<feff>', '')
    tables = pd.read_html(_io.StringIO(html), flavor='lxml', header=0)
    df = tables[0]
    # 송장번호는 숫자로 파싱될 수 있지만 문자열로 유지
    if '송장번호' in df.columns:
        df['송장번호'] = df['송장번호'].astype(str).str.replace(r'\.0$', '', regex=True)
    return df


def detect_and_load_input(path: Path) -> pd.DataFrame:
    """입력 파일 형식을 자동 감지해서 로드.
    - .xlsx/.xlsm: openpyxl → 첫 번째 시트 반환
    - .xls: HTML-format 시도 (마켓플레이스 발주 파일)
    """
    suffix = path.suffix.lower()
    if suffix in ('.xlsx', '.xlsm'):
        sheets = load_sheets(path)
        return list(sheets.values())[0]
    elif suffix == '.xls':
        return load_html_xls(path)
    else:
        raise ValueError(f"지원하지 않는 파일 형식: {suffix}")


def save_sheets(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    """sheets dict 를 xlsx 로 저장. 빈 DataFrame 도 시트로 포함."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)


def load_csv_ref(path: Path) -> pd.DataFrame:
    """UTF-8-sig(BOM) csv 로드. 참조 데이터(도서산간/필터링/미배송) 전용."""
    return pd.read_csv(path, encoding="utf-8-sig", dtype=str).fillna("")
